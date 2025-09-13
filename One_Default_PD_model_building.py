

import pandas as pd
import numpy as np
from pandasql import sqldf
import model_building_functions_toobox

pysqldf = lambda q: sqldf(q, globals())


####model building   ***#reading raw data
    
df = pd.read_csv(r'D:\HA_study_plan_tasks\banking_financial_scorecard_dev\model_1\data\model_raw_data.csv')


# contents the whole data :
def df_contents(df: pd.DataFrame):
    """
    Create PROC CONTENTS-style metadata tables for a pandas DataFrame.
    Returns:
        meta_df   : dataset-level metadata (1 row)
        vars_df   : variable-level catalog (one row per column)
    """
    # --- dataset-level (like "Data Set Attributes") ---
    meta = {
        "n_rows": [len(df)],
        "n_columns": [df.shape[1]],
        "memory_bytes": [df.memory_usage(deep=True).sum()],
        "index_name": [df.index.name],
        "index_type": [type(df.index).__name__],
        "column_names": [list(df.columns)],
    }
    meta_df = pd.DataFrame(meta)

    # --- variable-level (like "Variables") ---
    rows = []
    for col in df.columns:
        s = df[col]
        # best-effort basic stats
        non_null = s.notna().sum()
        nulls = s.isna().sum()
        nunique = s.nunique(dropna=True)

        # Simple type mapping (SAS has numeric/char; here we show dtype + a coarse class)
        pd_dtype = str(s.dtype)
        if pd.api.types.is_numeric_dtype(s):
            kind = "numeric"
            vmin = s.min(skipna=True)
            vmax = s.max(skipna=True)
            mean = s.mean(skipna=True)
            sample = s.dropna().iloc[0] if non_null else np.nan
        elif pd.api.types.is_datetime64_any_dtype(s):
            kind = "datetime"
            vmin = s.min(skipna=True)
            vmax = s.max(skipna=True)
            mean = pd.NaT
            sample = s.dropna().iloc[0] if non_null else pd.NaT
        elif pd.api.types.is_bool_dtype(s):
            kind = "boolean"
            vmin = s.min(skipna=True)
            vmax = s.max(skipna=True)
            mean = s.mean(skipna=True)
            sample = s.dropna().iloc[0] if non_null else np.nan
        elif pd.api.types.is_categorical_dtype(s):
            kind = "categorical"
            vmin = vmax = mean = np.nan
            sample = s.dropna().iloc[0] if non_null else np.nan
        else:
            kind = "string"
            vmin = vmax = mean = np.nan
            sample = s.dropna().iloc[0] if non_null else np.nan

        # memory for this column
        mem_bytes = s.memory_usage(deep=True)

        rows.append({
            "varnum": len(rows) + 1,          # position like SAS VARNUM
            "name": col,
            "kind": kind,                      # coarse SAS-like class
            "pandas_dtype": pd_dtype,          # precise pandas/NumPy dtype
            "non_null": int(non_null),
            "nulls": int(nulls),
            "n_unique": int(nunique),
            "memory_bytes": int(mem_bytes),
            "min": vmin,
            "max": vmax,
            "mean": mean,
            "example_non_null": sample,
        })

    vars_df = pd.DataFrame(rows)

    # sort by varnum to match SAS default
    vars_df = vars_df.sort_values("varnum", kind="stable").reset_index(drop=True)
    return meta_df, vars_df

#call the function:
meta_df, vars_df= df_contents(df)  #output the vars_df to excel for checking: 
    
#lst all numeric  attributes as below:
cn_list=["ALL0100",	"ALL0200",	"ALL0216",	"ALL0317",	"ALL0336",	"ALL0400",	"ALL0416",	"ALL0436",	"ALL0446",	"ALL0448",	"ALL0700",	"ALL1380",	"ALL2000",	"ALL2001",	"ALL2002",	"ALL2005",	"ALL2106",	"ALL2136",	"ALL2176",	"ALL2326",	"ALL2327",	"ALL2357",	"ALL2380",	"ALL2387",	"ALL2427",	"ALL2428",	"ALL2480",	"ALL2707",	"ALL2870",	"ALL2907",	"ALL2937",	"ALL2967",	"ALL2990",	"ALL3517",	"ALL4080",	"ALL5047",	"ALL5070",	"ALL5743",	"ALL8026",	"ALL8164",	"ALL8167",	"ALL8325",	"ALX5839",	"AUA0416",	"AUA1300",	"AUA2358",	"AUA5820",	"AUA6160",	"BCA0416",	"BCA5030",	"BCA5070",	"BCA5740",	"BCA6220",	"BCC0436",	"BCC0446",	"BCC1360",	"BCC3342",	"BCC3345",	"BCC3423",	"BCC3512",	"BCC5020",	"BCC5320",	"BCC5421",	"BCC5620",	"BCC5627",	"BCC7117",	"BCC8120",	"BCX3421",	"BCX3422",	"BRC0416",	"BRC1300",	"BRC5320",	"BRC5830",	"BRC7140",	"COL3210",	"COL5060",	"COL8165",	"FIP0300",	"FIP2358",	"IQT9415",	"IQT9420",	"IQT9425",	"IQT9426",	"International",	"MTA6160",	"MTF2358",	"MTF8166",	"MTX5839",	"PIL5020",	"REV0300",	"REV1360",	"REV2320",	"REV3421",	"REV5020",	"REV5620",	"REV5627",	"REV8160",	"RTI5020",	"RTI5820",	"RTR3422",	"ResponseScore",	"RiskScore",	"STU5820",	"VANTAGE_V3_SCORE",	"anual_income",	"r_xnm3",	"r_xnm4",	"r_xnm5",	"r_xnm6",	"r_xnm7",	"r_xnm8",	"r_xnm11",	"r_xnm12",	"r_xnm13",	"r_xnm14",	"r_xnm15",	"r_xnm16",	"r_xnm17",	"r_xnm18",	"r_xnm19",	"r_xnm20",	"r_xnm26",	"r_xnm27",	"r_xnm28",	"r_xnm29",	"r_xnm30",	"r_xnm31",	"r_xnm32",	"r_xnm33",	"r_xnm34",	"r_xnm35",	"r_xnm36",	"r_xnm37",	"r_xnm38",	"r_xnm39",	"r_xnm40",	"r_xnm41",	"r_xnm42",	"r_xnm43",	"r_xnm44",	"r_xnm45",	"r_xnm46",	"r_xnm47",	"r_xnm48",	"r_xnm49",	"r_xnm50",	"r_xnm51",	"r_xnm52",	"r_xnm53",	"r_xnm54",	"r_xnm55",	"r_xnm57",	"r_xnm58",	"r_xnm59",	"r_xnm60",	"r_xnm61",	"r_xnm62",	"r_xnm63",	"r_xnm64",	"r_xnm65",	"r_xnm66",	"r_xnm67",	"r_xnm68",	"r_xnm69",	"r_xnm70",	"r_xnm71",	"r_xnm72",	"r_xnm73",	"r_xnm74",	"r_xnm75",	"r_xnm76",	"r_xnm77",	"r_xnm78",	"r_xnm79",	"r_xnm80",	"r_xnm81",	"r_xnm82",	"r_xnm83",	"r_xnm84",	"r_xnm85",	"r_xnm86",	"r_xnm87",	"r_xnm89",	"r_xnm90",	"r_xnm91",	"r_xnm92",	"r_xnm93",	"r_xnm94",	"r_xnm95",	"r_xnm96",	"r_xnm97",	"r_xnm98",	"r_xnm99",	"r_xnm100",	"r_xnm101",	"r_xnm102",	"r_xnm103",	"r_xnm104",	"r_xnm105",	"r_xnm106",	"r_xnm107",	"r_xnm108",	"r_xnm109",	"r_xnm110",	"r_xnm111",	"r_xnm112",	"r_xnm113",	"r_xnm114",	"r_xnm115",	"r_xnm116",	"r_xnm117",	"r_xnm118",	"r_xnm119",	"r_xnm120",	"r_xnm121",	"r_xnm122",	"r_xnm123",	"r_xnm124",	"r_xnm125",	"r_xnm126",	"r_xnm127",	"r_xnm128",	"r_xnm129",	"r_xnm130",	"r_xnm131",	"r_xnm132",	"r_xnm133",	"r_xnm134",	"r_xnm135",	"r_xnm136",	"r_xnm137",	"r_xnm138",	"r_xnm139",	"r_xnm140",	"r_xnm141",	"r_xnm142",	"r_xnm143",	"r_xnm144",	"r_xnm145",	"r_xnm146",	"r_xnm147",	"r_xnm148",	"r_xnm149",	"r_xnm150",	"r_xnm151",	"r_xnm152",	"r_xnm153",	"r_xnm154",	"r_xnm155",	"r_xnm156",	"r_xnm157",	"r_xnm158",	"r_xnm159",	"r_xnm160",	"r_xnm161",	"r_xnm162",	"r_xnm163",	"r_xnm164",	"r_xnm165",	"r_xnm166",	"r_xnm167",	"r_xnm168",	"r_xnm169",	"r_xnm170",	"r_xnm171",	"r_xnm172",	"r_xnm173",	"r_xnm174",	"r_xnm175",	"r_xnm176",	"r_xnm177",	"r_xnm178",	"r_xnm179",	"r_xnm180",	"r_xnm181",	"r_xnm182",	"r_xnm183",	"r_xnm184",	"r_xnm185",	"r_xnm186",	"r_xnm187",	"r_xnm188",	"r_xnm189",	"r_xnm190",	"r_xnm191",	"r_xnm192",	"r_xnm193",	"r_xnm194",	"r_xnm195",	"r_xnm196",	"r_xnm197",	"r_xnm198",	"r_xnm199",	"r_xnm200",	"r_xnm201",	"r_xnm202",	"r_xnm203",	"r_xnm204",	"r_xnm205",	"r_xnm206",	"r_xnm207",	"r_xnm208",	"r_xnm209",	"r_xnm210",	"r_xnm211",	"r_xnm212",	"r_xnm213",	"r_xnm214",	"r_xnm215",	"r_xnm216",	"r_xnm217",	"r_xnm218",	"r_xnm219",	"r_xnm220",	"r_xnm221",	"r_xnm222",	"r_xnm223",	"r_xnm224",	"r_xnm225",	"r_xnm226",	"r_xnm227",	"r_xnm228",	"r_xnm229",	"r_xnm230",	"r_xnm231",	"r_xnm232",	"r_xnm233",	"r_xnm234",	"r_xnm235",	"r_xnm236",	"r_xnm237",	"r_xnm238",	"r_xnm239",	"r_xnm240",	"r_xnm241",	"r_xnm242",	"r_xnm243",	"r_xnm244",	"r_xnm245",	"r_xnm246",	"r_xnm247",	"r_xnm248",	"r_xnm249",	"r_xnm250",	"r_xnm251",	"r_xnm252",	"r_xnm253",	"r_xnm254",	"r_xnm255",	"r_xnm256",	"r_xnm257",	"r_xnm258",	"r_xnm259",	"r_xnm260",	"r_xnm261",	"r_xnm262",	"r_xnm263",	"r_xnm264",	"r_xnm265",	"r_xnm266",	"r_xnm267",	"r_xnm268",	"r_xnm269",	"r_xnm270",	"r_xnm271",	"r_xnm272",	"r_xnm273",	"r_xnm274",	"r_xnm275",	"r_xnm276",	"r_xnm277",	"r_xnm278",	"r_xnm279",	"r_xnm280",	"r_xnm281",	"r_xnm282",	"r_xnm283",	"r_xnm284",	"r_xnm285",	"r_xnm286",	"r_xnm287",	"r_xnm288",	"r_xnm289",	"r_xnm290",	"r_xnm291",	"r_xnm292",	"r_xnm293",	"r_xnm294",	"r_xnm295",	"r_xnm296",	"r_xnm297",	"r_xnm298",	"r_xnm299",	"r_xnm300",	"r_xnm301",	"r_xnm302",	"r_xnm303",	"r_xnm304",	"r_xnm305",	"r_xnm306",	"r_xnm307",	"r_xnm308",	"r_xnm309",	"r_xnm310",	"r_xnm311",	"r_xnm312",	"r_xnm313",	"r_xnm314",	"r_xnm315",	"r_xnm316",	"r_xnm317",	"r_xnm318",	"r_xnm319",	"r_xnm320",	"r_xnm321",	"r_xnm322",	"r_xnm323",	"r_xnm324",	"r_xnm325",	"r_xnm326",	"r_xnm327",	"r_xnm328",	"r_xnm329",	"r_xnm330",	"r_xnm331",	"r_xnm332",	"r_xnm333",	"r_xnm334",	"r_xnm335",	"r_xnm336",	"r_xnm337",	"r_xnm338",	"r_xnm339",	"r_xnm340",	"r_xnm341",	"r_xnm342",	"r_xnm343",	"r_xnm344",	"r_xnm345",	"r_xnm346",	"r_xnm347",	"r_xnm348",	"r_xnm349",	"r_xnm350",	"r_xnm351",	"r_xnm352",	"r_xnm353",	"r_xnm354",	"r_xnm355",	"r_xnm356",	"r_xnm357",	"r_xnm358",	"r_xnm359",	"r_xnm360",	"r_xnm361",	"r_xnm362",	"r_xnm363",	"r_xnm364",	"r_xnm365",	"r_xnm366",	"r_xnm367",	"r_xnm368",	"r_xnm369",	"r_xnm370",	"r_xnm371",	"r_xnm372",	"r_xnm373",	"r_xnm374",	"r_xnm375",	"r_xnm376",	"r_xnm377",	"r_xnm378",	"r_xnm379",	"r_xnm380",	"r_xnm381",	"r_xnm382",	"r_xnm383",	"r_xnm384",	"r_xnm385",	"r_xnm386",	"r_xnm387",	"r_xnm388",	"r_xnm389",	"r_xnm390",	"r_xnm391",	"r_xnm392",	"r_xnm393",	"r_xnm394",	"r_xnm395",	"r_xnm396",	"r_xnm397",	"r_xnm398",	"r_xnm399",	"r_xnm400",	"r_xnm401",	"r_xnm402",	"r_xnm403",	"r_xnm404",	"r_xnm405",	"r_xnm406",	"r_xnm407",	"r_xnm408",	"r_xnm409",	"r_xnm410",	"r_xnm411",	"r_xnm412",	"r_xnm413",	"r_xnm414",	"r_xnm415",	"r_xnm416",	"r_xnm417",	"r_xnm418",	"r_xnm419",	"r_xnm420",	"r_xnm422",	"r_xnm423",	"r_xnm424",	"r_xnm425",	"r_xnm426",	"r_xnm427",	"r_xnm428",	"r_xnm429",	"r_xnm430",	"r_xnm431",	"r_xnm432",	"r_xnm433",	"r_xnm434",	"r_xnm435",	"r_xnm436",	"r_xnm437",	"r_xnm438",	"r_xnm439",	"r_xnm440",	"r_xnm441",	"r_xnm442",	"r_xnm443",	"r_xnm444",	"r_xnm445",	"r_xnm446",	"r_xnm447",	"r_xnm448",	"r_xnm449",	"r_xnm450",	"r_xnm451",	"r_xnm452",	"r_xnm453",	"r_xnm454",	"r_xnm455",	"r_xnm456",	"r_xnm457",	"r_xnm458",	"r_xnm459",	"r_xnm460",	"r_xnm461",	"r_xnm462",	"r_xnm463",	"r_xnm464",	"r_xnm465",	"r_xnm466",	"r_xnm467",	"r_xnm468",	"r_xnm469",	"r_xnm470",	"r_xnm471",	"r_xnm472",	"r_xnm473",	"r_xnm474",	"r_xnm475",	"r_xnm476",	"r_xnm477",	"r_xnm478",	"r_xnm479",	"r_xnm480",	"r_xnm481",	"r_xnm482",	"r_xnm483",	"r_xnm484",	"r_xnm485",	"r_xnm486",	"r_xnm487",	"r_xnm488",	"r_xnm489",	"r_xnm490",	"r_xnm491",	"r_xnm492",	"r_xnm493",	"r_xnm494",	"r_xnm495",	"r_xnm496",	"r_xnm497",	"r_xnm498",	"r_xnm499",	"r_xnm500",	"r_xnm501",	"r_xnm502",	"r_xnm503",	"r_xnm504",	"r_xnm505",	"r_xnm506",	"r_xnm507",	"r_xnm508",	"r_xnm509",	"r_xnm510",	"r_xnm511",	"r_xnm512",	"r_xnm513",	"r_xnm514",	"r_xnm515",	"r_xnm516",	"r_xnm517",	"r_xnm518",	"r_xnm519",	"r_xnm520",	"r_xnm521",	"r_xnm522",	"r_xnm523",	"r_xnm524",	"r_xnm525",	"r_xnm526",	"r_xnm527",	"r_xnm528",	"r_xnm529",	"r_xnm530",	"r_xnm531",	"r_xnm532",	"r_xnm533",	"r_xnm534",	"r_xnm535",	"r_xnm536",	"r_xnm537",	"r_xnm538",	"r_xnm539",	"r_xnm540",	"r_xnm541",	"r_xnm542",	"r_xnm543",	"r_xnm544",	"r_xnm545",	"r_xnm546",	"r_xnm547",	"r_xnm548",	"r_xnm549",	"r_xnm550",	"r_xnm551",	"r_xnm552",	"r_xnm553",	"r_xnm554",	"r_xnm555",	"r_xnm556",	"r_xnm557",	"r_xnm558",	"r_xnm559",	"r_xnm560",	"r_xnm561",	"r_xnm562",	"r_xnm563",	"r_xnm564",	"r_xnm565",	"r_xnm566",	"r_xnm567",	"r_xnm568",	"r_xnm569",	"r_xnm570",	"r_xnm571",	"r_xnm572",	"r_xnm573",	"r_xnm574",	"r_xnm575",	"r_xnm576",	"r_xnm577",	"r_xnm578",	"r_xnm579",	"r_xnm580",	"r_xnm581",	"r_xnm582",	"r_xnm583",	"r_xnm584",	"r_xnm585",	"r_xnm586",	"r_xnm587",	"r_xnm588",	"r_xnm589",	"r_xnm590",	"r_xnm591",	"r_xnm592",	"r_xnm593",	"r_xnm594",	"r_xnm595",	"r_xnm596",	"r_xnm597",	"r_xnm598",	"r_xnm599",	"r_xnm600",	"r_xnm601",	"r_xnm602",	"r_xnm603",	"r_xnm604",	"r_xnm605",	"r_xnm606",	"r_xnm607",	"r_xnm608",	"r_xnm609",	"r_xnm610",	"r_xnm611",	"r_xnm612",	"r_xnm613",	"r_xnm614",	"r_xnm615",	"r_xnm616",	"r_xnm617",	"r_xnm618",	"r_xnm619",	"r_xnm620",	"r_xnm621",	"r_xnm622",	"r_xnm623",	"r_xnm624",	"r_xnm625",	"r_xnm626",	"r_xnm627",	"r_xnm628",	"r_xnm629",	"r_xnm630",	"r_xnm631",	"r_xnm632",	"r_xnm633",	"r_xnm634",	"r_xnm635",	"r_xnm636",	"r_xnm637",	"r_xnm638",	"r_xnm639",	"r_xnm640",	"r_xnm641",	"r_xnm642",	"r_xnm643",	"r_xnm644",	"r_xnm645",	"r_xnm646",	"r_xnm647",	"r_xnm648",	"r_xnm649",	"r_xnm650",	"r_xnm651",	"r_xnm652",	"r_xnm653",	"r_xnm654",	"r_xnm655",	"r_xnm656",	"r_xnm657",	"r_xnm658",	"r_xnm659",	"r_xnm660",	"r_xnm661",	"r_xnm662",	"r_xnm663",	"r_xnm664",	"r_xnm665",	"r_xnm666",	"r_xnm667",	"r_xnm668",	"r_xnm669",	"r_xnm670",	"r_xnm671",	"r_xnm672",	"r_xnm673",	"r_xnm674",	"r_xnm675",	"r_xnm676",	"r_xnm677",	"r_xnm678",	"r_xnm679",	"r_xnm680",	"r_xnm681",	"r_xnm682",	"r_xnm683",	"r_xnm684",	"r_xnm685",	"r_xnm686",	"r_xnm687",	"r_xnm688",	"r_xnm689",	"r_xnm690",	"r_xnm691",	"r_xnm692",	"r_xnm693",	"r_xnm694",	"r_xnm695",	"r_xnm696",	"r_xnm697",	"r_xnm698",	"r_xnm699",	"r_xnm700",	"r_xnm701",	"r_xnm702",	"r_xnm703",	"r_xnm704",	"r_xnm705",	"r_xnm706",	"r_xnm707",	"r_xnm708",	"r_xnm709",	"r_xnm710",	"r_xnm711",	"r_xnm712",	"r_xnm713",	"r_xnm714",	"r_xnm715",	"r_xnm716",	"r_xnm717",	"r_xnm718",	"r_xnm719",	"r_xnm720",	"r_xnm721",	"r_xnm722",	"r_xnm723",	"r_xnm724",	"r_xnm725",	"r_xnm726",	"r_xnm727",	"r_xnm728",	"r_xnm729",	"r_xnm730",	"r_xnm731",	"r_xnm732",	"r_xnm733",	"r_xnm734",	"r_xnm735",	"r_xnm736",	"r_xnm737",	"r_xnm738",	"r_xnm739",	"r_xnm740",	"r_xnm741",	"r_xnm742",	"r_xnm743",	"r_xnm744",	"r_xnm745",	"r_xnm746",	"r_xnm747",	"r_xnm748",	"r_xnm749",	"r_xnm750",	"r_xnm751",	"r_xnm752",	"r_xnm753",	"r_xnm754",	"r_xnm755",	"r_xnm756",	"r_xnm757",	"r_xnm758",	"r_xnm759",	"r_xnm760",	"r_xnm761",	"r_xnm762",	"r_xnm763",	"r_xnm764",	"r_xnm765",	"r_xnm766",	"r_xnm767",	"r_xnm768",	"r_xnm769",	"r_xnm770",	"r_xnm771",	"r_xnm772",	"r_xnm773",	"r_xnm774",	"r_xnm775",	"r_xnm776",	"r_xnm777",	"r_xnm778",	"r_xnm779",	"r_xnm780",	"r_xnm781",	"r_xnm782",	"r_xnm783",	"r_xnm784",	"r_xnm785",	"r_xnm786",	"r_xnm787",	"r_xnm788",	"r_xnm789",	"r_xnm790",	"r_xnm791",	"r_xnm792",	"r_xnm793",	"r_xnm794",	"r_xnm795",	"r_xnm796",	"r_xnm797",	"r_xnm798",	"r_xnm799",	"r_xnm800",	"r_xnm801",	"r_xnm802",	"r_xnm803",	"r_xnm804",	"r_xnm805",	"r_xnm806",	"r_xnm807",	"r_xnm808",	"r_xnm809",	"r_xnm810",	"r_xnm811",	"r_xnm812",	"r_xnm813",	"r_xnm814",	"r_xnm815",	"r_xnm816",	"r_xnm817",	"r_xnm818",	"r_xnm819",	"r_xnm820",	"r_xnm821",	"r_xnm822",	"r_xnm823",	"r_xnm824",	"r_xnm825",	"r_xnm826",	"r_xnm827",	"r_xnm828",	"r_xnm829",	"r_xnm830",	"r_xnm831",	"r_xnm832",	"r_xnm833",	"r_xnm834",	"r_xnm835",	"r_xnm836",	"r_xnm837",	"r_xnm838",	"r_xnm839",	"r_xnm840",	"r_xnm841",	"r_xnm842",	"r_xnm843",	"r_xnm844",	"r_xnm845",	"r_xnm846",	"r_xnm847",	"r_xnm848",	"r_xnm850",	"r_xnm851",	"r_xnm853",	"r_xnm854",	"r_xnm855",	"r_xnm856",	"r_xnm857",	"r_xnm858",	"r_xnm859",	"r_xnm860",	"r_xnm861",	"r_xnm862",	"r_xnm863",	"r_xnm865",	"r_xnm866",	"r_xnm867",	"r_xnm868",	"r_xnm869",	"r_xnm870",	"r_xnm871",	"r_xnm872",	"r_xnm873",	"r_xnm874",	"r_xnm876",	"r_xnm877",	"r_xnm878",	"r_xnm879",	"r_xnm880",	"r_xnm881",	"r_xnm882",	"r_xnm883",	"r_xnm884",	"r_xnm885",	"r_xnm887",	"r_xnm888",	"r_xnm889",	"r_xnm890",	"r_xnm891",	"r_xnm892",	"r_xnm893",	"r_xnm894",	"r_xnm895",	"r_xnm896",	"r_xnm897",	"r_xnm899",	"r_xnm900",	"r_xnm901",	"r_xnm902",	"r_xnm903",	"r_xnm904",	"r_xnm905",	"r_xnm906",	"r_xnm907",	"r_xnm908",	"r_xnm909",	"r_xnm911",	"r_xnm912",	"r_xnm913",	"r_xnm914",	"r_xnm915",	"r_xnm916",	"r_xnm917",	"r_xnm918",	"r_xnm919",	"r_xnm920",	"r_xnm921",	"r_xnm923",	"r_xnm924",	"r_xnm925",	"r_xnm926",	"r_xnm927",	"r_xnm928",	"r_xnm929",	"r_xnm930",	"r_xnm931",	"r_xnm932",	"r_xnm933",	"r_xnm935",	"r_xnm936",	"r_xnm937",	"r_xnm938",	"r_xnm939",	"r_xnm940",	"r_xnm941",	"r_xnm942",	"r_xnm943",	"r_xnm944",	"r_xnm945",	"r_xnm947",	"r_xnm948",	"r_xnm949",	"r_xnm950",	"r_xnm951",	"r_xnm952",	"r_xnm953",	"r_xnm954",	"r_xnm955",	"r_xnm956",	"r_xnm957",	"r_xnm959",	"r_xnm960",	"r_xnm961",	"r_xnm962",	"r_xnm963",	"r_xnm964",	"r_xnm965",	"r_xnm966",	"r_xnm967",	"r_xnm968",	"r_xnm969",	"r_xnm971",	"r_xnm972",	"r_xnm973",	"r_xnm974",	"r_xnm975",	"r_xnm976",	"r_xnm977",	"r_xnm978",	"r_xnm979",	"r_xnm980",	"r_xnm982",	"r_xnm983",	"r_xnm984",	"r_xnm985",	"r_xnm986",	"r_xnm987",	"r_xnm988",	"r_xnm989",	"r_xnm990",	"r_xnm991",	"r_xnm992",	"r_xnm994",	"r_xnm995",	"r_xnm996",	"r_xnm997",	"r_xnm998",	"r_xnm999",	"r_xnm1000",	"r_xnm1001",	"r_xnm1002",	"r_xnm1003",	"r_xnm1004",	"r_xnm1006",	"r_xnm1007",	"r_xnm1008",	"r_xnm1009",	"r_xnm1010",	"r_xnm1011",	"r_xnm1012",	"r_xnm1013",	"r_xnm1014",	"r_xnm1015",	"r_xnm1016",	"r_xnm1018",	"r_xnm1019",	"r_xnm1020",	"r_xnm1021",	"r_xnm1022",	"r_xnm1023",	"r_xnm1024",	"r_xnm1025",	"r_xnm1026",	"r_xnm1027",	"r_xnm1028",	"r_xnm1030",	"r_xnm1031",	"r_xnm1032",	"r_xnm1033",	"r_xnm1034",	"r_xnm1035",	"r_xnm1036",	"r_xnm1037",	"r_xnm1038",	"r_xnm1039",	"r_xnm1040",	"r_xnm1042",	"r_xnm1043",	"r_xnm1044",	"r_xnm1045",	"r_xnm1046",	"r_xnm1047",	"r_xnm1048",	"r_xnm1049",	"r_xnm1050",	"r_xnm1051",	"r_xnm1052",	"r_xnm1054",	"r_xnm1055",	"r_xnm1056",	"r_xnm1057",	"r_xnm1058",	"r_xnm1059",	"r_xnm1060",	"r_xnm1061",	"r_xnm1062",	"r_xnm1063",	"r_xnm1064",	"r_xnm1066",	"r_xnm1067",	"r_xnm1068",	"r_xnm1069",	"r_xnm1070",	"r_xnm1071",	"r_xnm1072",	"r_xnm1073",	"r_xnm1074",	"r_xnm1075",	"r_xnm1076",	"r_xnm1078",	"r_xnm1079",	"r_xnm1080",	"r_xnm1081",	"r_xnm1082",	"r_xnm1083",	"r_xnm1084",	"r_xnm1085",	"r_xnm1086",	"r_xnm1087",	"r_xnm1088",	"r_xnm1090",	"r_xnm1091",	"r_xnm1092",	"r_xnm1093",	"r_xnm1094",	"r_xnm1095",	"r_xnm1096",	"r_xnm1097",	"r_xnm1098",	"r_xnm1099",	"r_xnm1100",	"r_xnm1102",	"r_xnm1103",	"r_xnm1104",	"r_xnm1105",	"r_xnm1106",	"r_xnm1107",	"r_xnm1108",	"r_xnm1109",	"r_xnm1110",	"r_xnm1111",	"r_xnm1112",	"r_xnm1114",	"r_xnm1115",	"r_xnm1116",	"r_xnm1117",	"r_xnm1118",	"r_xnm1119",	"r_xnm1120",	"r_xnm1121",	"r_xnm1122",	"r_xnm1123",	"r_xnm1125",	"r_xnm1126",	"r_xnm1127",	"r_xnm1128",	"r_xnm1129",	"r_xnm1130",	"r_xnm1131",	"r_xnm1132",	"r_xnm1133",	"r_xnm1134",	"r_xnm1135",	"r_xnm1137",	"r_xnm1138",	"r_xnm1139",	"r_xnm1140",	"r_xnm1141",	"r_xnm1142",	"r_xnm1143",	"r_xnm1144",	"r_xnm1145",	"r_xnm1146",	"r_xnm1147",	"r_xnm1149",	"r_xnm1150",	"r_xnm1151",	"r_xnm1152",	"r_xnm1153",	"r_xnm1154",	"r_xnm1155",	"r_xnm1156",	"r_xnm1157",	"r_xnm1158",	"r_xnm1159",	"r_xnm1161",	"r_xnm1162",	"r_xnm1163",	"r_xnm1164",	"r_xnm1165",	"r_xnm1166",	"r_xnm1167",	"r_xnm1168",	"r_xnm1169",	"r_xnm1170",	"r_xnm1171",	"r_xnm1173",	"r_xnm1174",	"r_xnm1175",	"r_xnm1176",	"r_xnm1177",	"r_xnm1178",	"r_xnm1179",	"r_xnm1180",	"r_xnm1181",	"r_xnm1182",	"r_xnm1183",	"r_xnm1185",	"r_xnm1186",	"r_xnm1187",	"r_xnm1188",	"r_xnm1189",	"r_xnm1190",	"r_xnm1191",	"r_xnm1192",	"r_xnm1193",	"r_xnm1194",	"r_xnm1195",	"r_xnm1197",	"r_xnm1198",	"r_xnm1199",	"r_xnm1200",	"r_xnm1201",	"r_xnm1202",	"r_xnm1203",	"r_xnm1204",	"r_xnm1205",	"r_xnm1206",	"r_xnm1207",	"r_xnm1209",	"r_xnm1210",	"r_xnm1211",	"r_xnm1212",	"r_xnm1213",	"r_xnm1214",	"r_xnm1215",	"r_xnm1216",	"r_xnm1217",	"r_xnm1218",	"r_xnm1219",	"r_xnm1221",	"r_xnm1222",	"r_xnm1223",	"r_xnm1224",	"r_xnm1225",	"r_xnm1226",	"r_xnm1227",	"r_xnm1228",	"r_xnm1229",	"r_xnm1230",	"r_xnm1231",	"r_xnm1233",	"r_xnm1234",	"r_xnm1235",	"r_xnm1236",	"r_xnm1237",	"r_xnm1238",	"r_xnm1239",	"r_xnm1240",	"r_xnm1241",	"r_xnm1242",	"r_xnm1243",	"r_xnm1245",	"r_xnm1246",	"r_xnm1247",	"r_xnm1248",	"r_xnm1249",	"r_xnm1250",	"r_xnm1251",	"r_xnm1252",	"r_xnm1253",	"r_xnm1254",	"r_xnm1255",	"r_xnm1257",	"r_xnm1258",	"r_xnm1259",	"r_xnm1260",	"r_xnm1261",	"r_xnm1262",	"r_xnm1263",	"r_xnm1264",	"r_xnm1265",	"r_xnm1266",	"r_xnm1268",	"r_xnm1269",	"r_xnm1270",	"r_xnm1271",	"r_xnm1272",	"r_xnm1273",	"r_xnm1274",	"r_xnm1275",	"r_xnm1276",	"r_xnm1277",	"r_xnm1278",	"r_xnm1280",	"r_xnm1281",	"r_xnm1282",	"r_xnm1283",	"r_xnm1284",	"r_xnm1285",	"r_xnm1286",	"r_xnm1287",	"r_xnm1288",	"r_xnm1289",	"r_xnm1290",	"r_xnm1292",	"r_xnm1293",	"r_xnm1294",	"r_xnm1295",	"r_xnm1296",	"r_xnm1297",	"r_xnm1298",	"r_xnm1299",	"r_xnm1300",	"r_xnm1301",	"r_xnm1302",	"r_xnm1304",	"r_xnm1305",	"r_xnm1306",	"r_xnm1307",	"r_xnm1308",	"r_xnm1309",	"r_xnm1310",	"r_xnm1311",	"r_xnm1312",	"r_xnm1313",	"r_xnm1314",	"r_xnm1316",	"r_xnm1317",	"r_xnm1318",	"r_xnm1319",	"r_xnm1320",	"r_xnm1321",	"r_xnm1322",	"r_xnm1323",	"r_xnm1324",	"r_xnm1325",	"r_xnm1327",	"r_xnm1328",	"r_xnm1329",	"r_xnm1330",	"r_xnm1331",	"r_xnm1332",	"r_xnm1333",	"r_xnm1334",	"r_xnm1335",	"r_xnm1336",	"r_xnm1337",	"r_xnm1339",	"r_xnm1340",	"r_xnm1341",	"r_xnm1342",	"r_xnm1343",	"r_xnm1344",	"r_xnm1345",	"r_xnm1346",	"r_xnm1347",	"r_xnm1348",	"r_xnm1349",	"r_xnm1351",	"r_xnm1352",	"r_xnm1353",	"r_xnm1354",	"r_xnm1355",	"r_xnm1356",	"r_xnm1357",	"r_xnm1358",	"r_xnm1359",	"r_xnm1360",	"r_xnm1361",	"r_xnm1363",	"r_xnm1364",	"r_xnm1365",	"r_xnm1366",	"r_xnm1367",	"r_xnm1368",	"r_xnm1369",	"r_xnm1370",	"r_xnm1371",	"r_xnm1372",	"r_xnm1373",	"r_xnm1375",	"r_xnm1376",	"r_xnm1377",	"r_xnm1378",	"r_xnm1379",	"r_xnm1380",	"r_xnm1381",	"r_xnm1382",	"r_xnm1383",	"r_xnm1384",	"r_xnm1385",	"r_xnm1387",	"r_xnm1388",	"r_xnm1389",	"r_xnm1390",	"r_xnm1391",	"r_xnm1392",	"r_xnm1393",	"r_xnm1394",	"r_xnm1395",	"r_xnm1396",	"r_xnm1397",	"r_xnm1399",	"r_xnm1400",	"r_xnm1401",	"r_xnm1402",	"r_xnm1403",	"r_xnm1404",	"r_xnm1405",	"r_xnm1406",	"r_xnm1407",	"r_xnm1408",	"r_xnm1409",	"r_xnm1411",	"r_xnm1412",	"r_xnm1413",	"r_xnm1414",	"r_xnm1415",	"r_xnm1416",	"r_xnm1417",	"r_xnm1418",	"r_xnm1419",	"r_xnm1420",	"r_xnm1422",	"r_xnm1423",	"r_xnm1424",	"r_xnm1425",	"r_xnm1426",	"r_xnm1427",	"r_xnm1428",	"r_xnm1429",	"r_xnm1430",	"r_xnm1431",	"r_xnm1432",	"r_xnm1434",	"r_xnm1435",	"r_xnm1436",	"r_xnm1437",	"r_xnm1438",	"r_xnm1439",	"r_xnm1440",	"r_xnm1441",	"r_xnm1442",	"r_xnm1443",	"r_xnm1444",	"r_xnm1446",	"r_xnm1447",	"r_xnm1448",	"r_xnm1449",	"r_xnm1450",	"r_xnm1451",	"r_xnm1452",	"r_xnm1453",	"r_xnm1454",	"r_xnm1455",	"r_xnm1456",	"r_xnm1458",	"r_xnm1459",	"r_xnm1460",	"r_xnm1461",	"r_xnm1462",	"r_xnm1463",	"r_xnm1464",	"r_xnm1465",	"r_xnm1466",	"r_xnm1467",	"r_xnm1468",	"r_xnm1470",	"r_xnm1471",	"r_xnm1472",	"r_xnm1473",	"r_xnm1474",	"r_xnm1475",	"r_xnm1476",	"r_xnm1477",	"r_xnm1478",	"r_xnm1479",	"r_xnm1480",	"r_xnm1482",	"r_xnm1483",	"r_xnm1484",	"r_xnm1485",	"r_xnm1486",	"r_xnm1487",	"r_xnm1488",	"r_xnm1489",	"r_xnm1490",	"r_xnm1491",	"r_xnm1492",	"r_xnm1494",	"r_xnm1495",	"r_xnm1496",	"r_xnm1497",	"r_xnm1498",	"r_xnm1499",	"r_xnm1500",	"r_xnm1501",	"r_xnm1502",	"r_xnm1503",	"r_xnm1504",	"r_xnm1506",	"r_xnm1507",	"r_xnm1508",	"r_xnm1509",	"r_xnm1510",	"r_xnm1511",	"r_xnm1512",	"r_xnm1513",	"r_xnm1514",	"r_xnm1515",	"r_xnm1516",	"r_xnm1518",	"r_xnm1519",	"r_xnm1520",	"r_xnm1521",	"r_xnm1522",	"r_xnm1523",	"r_xnm1524",	"r_xnm1525",	"r_xnm1526",	"r_xnm1527",	"r_xnm1528",	"r_xnm1530",	"r_xnm1531",	"r_xnm1532",	"r_xnm1533",	"r_xnm1534",	"r_xnm1535",	"r_xnm1536",	"r_xnm1537",	"r_xnm1538",	"r_xnm1539",	"r_xnm1540",	"r_xnm1542",	"r_xnm1543",	"r_xnm1544",	"r_xnm1545",	"r_xnm1546",	"r_xnm1547",	"r_xnm1548",	"r_xnm1549",	"r_xnm1550",	"r_xnm1551",	"r_xnm1552",	"r_xnm1554",	"r_xnm1555",	"r_xnm1556",	"r_xnm1557",	"r_xnm1558",	"r_xnm1559",	"r_xnm1560",	"r_xnm1561",	"r_xnm1562",	"r_xnm1563",	"r_xnm1565",	"r_xnm1566",	"r_xnm1567",	"r_xnm1568",	"r_xnm1569",	"r_xnm1570",	"r_xnm1571",	"r_xnm1572",	"r_xnm1573",	"r_xnm1574",	"r_xnm1575",	"r_xnm1577",	"r_xnm1578",	"r_xnm1579",	"r_xnm1580",	"r_xnm1581",	"r_xnm1582",	"r_xnm1583",	"r_xnm1584",	"r_xnm1585",	"r_xnm1586",	"r_xnm1587",	"r_xnm1589",	"r_xnm1590",	"r_xnm1591",	"r_xnm1592",	"r_xnm1593",	"r_xnm1594",	"r_xnm1595",	"r_xnm1596",	"r_xnm1597",	"r_xnm1598",	"r_xnm1599",	"r_xnm1601",	"r_xnm1602",	"r_xnm1603",	"r_xnm1604",	"r_xnm1605",	"r_xnm1606",	"r_xnm1607",	"r_xnm1608",	"r_xnm1609",	"r_xnm1610",	"r_xnm1611",	"r_xnm1613",	"r_xnm1614",	"r_xnm1615",	"r_xnm1616",	"r_xnm1617",	"r_xnm1618",	"r_xnm1619",	"r_xnm1620",	"r_xnm1621",	"r_xnm1623",	"r_xnm1625",	"r_xnm1626",	"r_xnm1627",	"r_xnm1628",	"r_xnm1629",	"r_xnm1630",	"r_xnm1631",	"r_xnm1632",	"r_xnm1633",	"r_xnm1634",	"r_xnm1635",	"r_xnm1636",	"r_xnm1637",	"r_xnm1638",	"r_xnm1639",	"r_xnm1640",	"r_xnm1641",	"r_xnm1642",	"r_xnm1643",	"r_xnm1644",	"r_xnm1645",	"r_xnm1646",	"r_xnm1647",	"r_xnm1648",	"r_xnm1649",	"r_xnm1650",	"r_xnm1651",	"r_xnm1652",	"r_xnm1653",	"r_xnm1654",	"r_xnm1655",	"r_xnm1656",	"r_xnm1657",	"r_xnm1658",	"r_xnm1659",	"r_xnm1660",	"r_xnm1661",	"r_xnm1663",	"r_xnm1664",	"r_xnm1665",	"r_xnm1666",	"r_xnm1667",	"r_xnm1668",	"r_xnm1669",	"r_xnm1670",	"r_xnm1671",	"r_xnm1672",	"r_xnm1673",	"r_xnm1674",	"r_xnm1675",	"r_xnm1676",	"r_xnm1677",	"r_xnm1678",	"r_xnm1679",	"r_xnm1680",	"r_xnm1681",	"r_xnm1683",	"r_xnm1684",	"r_xnm1685",	"r_xnm1686",	"r_xnm1687",	"r_xnm1688",	"r_xnm1689",	"r_xnm1690",	"r_xnm1691",	"r_xnm1692",	"r_xnm1693",	"r_xnm1694",	"r_xnm1695",	"r_xnm1696",	"r_xnm1697",	"r_xnm1698",	"r_xnm1699",	"r_xnm1700",	"r_xnm1701",	"r_xnm1702",	"r_xnm1703",	"r_xnm1704",	"r_xnm1705",	"r_xnm1706",	"r_xnm1707",	"r_xnm1710",	"r_xnm1712",	"r_xnm1713",	"r_xnm1714",	"r_xnm1715",	"r_xnm1716",	"r_xnm1717",	"r_xnm1718",	"r_xnm1719",	"r_xnm1720",	"r_xnm1721",	"r_xnm1722",	"r_xnm1723",	"r_xnm1724",	"r_xnm1726",	"r_xnm1727",	"r_xnm1728",	"r_xnm1729",	"r_xnm1730",	"r_xnm1731",	"r_xnm1732",	"r_xnm1733",	"r_xnm1735",	"r_xnm1736",	"r_xnm1737",	"r_xnm1738",	"r_xnm1739",	"r_xnm1740",	"r_xnm1741",	"r_xnm1742",	"r_xnm1744",	"r_xnm1745",	"r_xnm1746",	"r_xnm1747",	"r_xnm1748",	"r_xnm1749",	"r_xnm1750",	"r_xnm1751",	"r_xnm1753",	"r_xnm1754",	"r_xnm1755",	"r_xnm1756",	"r_xnm1757",	"r_xnm1758",	"r_xnm1759",	"r_xnm1760",	"r_xnm1762",	"r_xnm1763",	"r_xnm1764",	"r_xnm1765",	"r_xnm1766",	"r_xnm1767",	"r_xnm1768",	"r_xnm1769",	"r_xnm1771",	"r_xnm1772",	"r_xnm1773",	"r_xnm1774",	"r_xnm1775",	"r_xnm1776",	"r_xnm1777",	"r_xnm1778",	"r_xnm1780",	"r_xnm1781",	"r_xnm1782",	"r_xnm1783",	"r_xnm1784",	"r_xnm1785",	"r_xnm1786",	"r_xnm1787",	"r_xnm1788",	"r_xnm1789",	"r_xnm1790",	"r_xnm1791",	"r_xnm1792",	"r_xnm1793",	"r_xnm1794",	"r_xnm1795",	"r_xnm1796",	"r_xnm1797",	"r_xnm1798",	"r_xnm1799",	"r_xnm1800",	"r_xnm1801",	"r_xnm1802",	"r_xnm1803",	"r_xnm1804",	"r_xnm1805",	"r_xnm1806",	"r_xnm1807",	"r_xnm1808",	"r_xnm1809",	"r_xnm1810",	"r_xnm1811",	"r_xnm1812",	"r_xnm1813",	"r_xnm1814",	"r_xnm1815",	"r_xnm1816",	"r_xnm1817",	"r_xnm1818",	"r_xnm1819",	"r_xnm1820",	"r_xnm1821",	"r_xnm1822",	"r_xnm1823",	"r_xnm1824",	"r_xnm1825",	"r_xnm1826",	"r_xnm1827",	"r_xnm1828",	"r_xnm1829",	"r_xnm1830",	"r_xnm1831",	"r_xnm1832",	"r_xnm1833",	"r_xnm1834",	"r_xnm1835",	"r_xnm1836",	"r_xnm1837",	"r_xnm1838",	"r_xnm1839",	"r_xnm1840",	"r_xnm1841",	"r_xnm1842",	"r_xnm1843",	"r_xnm1844",	"r_xnm1845",	"r_xnm1846",	"r_xnm1847",	"r_xnm1848",	"r_xnm1849",	"r_xnm1850",	"r_xnm1851",	"r_xnm1852",	"r_xnm1853",	"r_xnm1854",	"r_xnm1856",	"r_xnm1857",	"r_xnm1858",	"r_xnm1859",	"r_xnm1860",	"r_xnm1861",	"r_xnm1862",	"r_xnm1863",	"r_xnm1865",	"r_xnm1866",	"r_xnm1867",	"r_xnm1868",	"r_xnm1869",	"r_xnm1870",	"r_xnm1871",	"r_xnm1873",	"r_xnm1874",	"r_xnm1875",	"r_xnm1876",	"r_xnm1877",	"r_xnm1878",	"r_xnm1880",	"r_xnm1881",	"r_xnm1882",	"r_xnm1883",	"r_xnm1884",	"r_xnm1886",	"r_xnm1887",	"r_xnm1888",	"r_xnm1890",	"r_xnm1891",	"r_xnm1892",	"r_xnm1894",	"r_xnm1895",	"r_xnm1897",	"r_xnm1899",	"r_xnm1900",	"r_xnm1901",	"r_xnm1902",	"r_xnm1904",	"r_xnm1905",	"r_xnm1906",	"r_xnm1907",	"r_xnm1908",	"r_xnm1909",	"r_xnm1910",	"r_xnm1911",	"r_xnm1912",	"r_xnm1914",	"r_xnm1915",	"r_xnm1916",	"r_xnm1917",	"r_xnm1918",	"r_xnm1919",	"r_xnm1920",	"r_xnm1921",	"r_xnm1922",	"r_xnm1924",	"r_xnm1925",	"r_xnm1926",	"r_xnm1927",	"r_xnm1928",	"r_xnm1929",	"r_xnm1930",	"r_xnm1931",	"r_xnm1932",	"r_xnm1934",	"r_xnm1935",	"r_xnm1936",	"r_xnm1937",	"r_xnm1938",	"r_xnm1939",	"r_xnm1940",	"r_xnm1941",	"r_xnm1942",	"r_xnm1944",	"r_xnm1945",	"r_xnm1946",	"r_xnm1947",	"r_xnm1948",	"r_xnm1949",	"r_xnm1950",	"r_xnm1951",	"r_xnm1952",	"r_xnm1954",	"r_xnm1955",	"r_xnm1956",	"r_xnm1957",	"r_xnm1958",	"r_xnm1959",	"r_xnm1960",	"r_xnm1961",	"r_xnm1962",	"r_xnm1964",	"r_xnm1965",	"r_xnm1966",	"r_xnm1967",	"r_xnm1968",	"r_xnm1969",	"r_xnm1970",	"r_xnm1971",	"r_xnm1972",	"r_xnm1974",	"r_xnm1975",	"r_xnm1976",	"r_xnm1977",	"r_xnm1978",	"r_xnm1979",	"r_xnm1980",	"r_xnm1981",	"r_xnm1982",	"r_xnm1984",	"r_xnm1985",	"r_xnm1986",	"r_xnm1987",	"r_xnm1988",	"r_xnm1989",	"r_xnm1990",	"r_xnm1991",	"r_xnm1992",	"r_xnm1993",	"r_xnm1994",	"r_xnm1995",	"r_xnm1996",	"r_xnm1997",	"r_xnm1998",	"r_xnm1999",	"r_xnm2000",	"r_xnm2001",	"r_xnm2003",	"r_xnm2004",	"r_xnm2005",	"r_xnm2006",	"r_xnm2007",	"r_xnm2009",	"r_xnm2010",	"r_xnm2011",	"r_xnm2012",	"r_xnm2013",	"r_xnm2014",	"r_xnm2015",	"r_xnm2016",	"r_xnm2017",	"r_xnm2018",	"r_xnm2019",	"r_xnm2020",	"r_xnm2021",	"r_xnm2022",	"r_xnm2023",	"r_xnm2024",	"r_xnm2025",	"r_xnm2026",	"r_xnm2027",	"r_xnm2028",	"r_xnm2029",	"r_xnm2030",	"r_xnm2031",	"r_xnm2032",	"r_xnm2033",	"r_xnm2034",	"r_xnm2035",	"r_xnm2036",	"r_xnm2037",	"r_xnm2038",	"r_xnm2039",	"r_xnm2040",	"r_xnm2041",	"r_xnm2042",	"r_xnm2043",	"r_xnm2044",	"r_xnm2045",	"r_xnm2046",	"r_xnm2047",	"r_xnm2048",	"r_xnm2049",	"r_xnm2050",	"r_xnm2051",	"r_xnm2052",	"r_xnm2053",	"r_xnm2054",	"r_xnm2055",	"r_xnm2056",	"r_xnm2057",	"r_xnm2058",	"r_xnm2059",	"r_xnm2060",	"r_xnm2061",	"r_xnm2062",	"r_xnm2063",	"r_xnm2064",	"r_xnm2065",	"r_xnm2066",	"r_xnm2067",	"r_xnm2068",	"r_xnm2069",	"r_xnm2070",	"r_xnm2071",	"r_xnm2072",	"r_xnm2073",	"r_xnm2074",	"r_xnm2076",	"r_xnm2077",	"r_xnm2078",	"r_xnm2079",	"r_xnm2080",	"r_xnm2081",	"r_xnm2082",	"r_xnm2083",	"r_xnm2084",	"r_xnm2085",	"r_xnm2086",	"r_xnm2088",	"r_xnm2089",	"r_xnm2090",	"r_xnm2091",	"r_xnm2092",	"r_xnm2093",	"r_xnm2094",	"r_xnm2095",	"r_xnm2096",	"r_xnm2097",	"r_xnm2098",	"r_xnm2099",	"r_xnm2100",	"r_xnm2101",	"r_xnm2102",	"r_xnm2103",	"r_xnm2104",	"r_xnm2105",	"r_xnm2106",	"r_xnm2107",	"r_xnm2108",	"r_xnm2109",	"r_xnm2110",	"r_xnm2111",	"r_xnm2112",	"r_xnm2113",	"r_xnm2114",	"r_xnm2115",	"r_xnm2116",	"r_xnm2117",	"r_xnm2118",	"r_xnm2119",	"r_xnm2120",	"r_xnm2121",	"r_xnm2122",	"r_xnm2123",	"r_xnm2124",	"r_xnm2125",	"r_xnm2126",	"r_xnm2127",	"r_xnm2128",	"r_xnm2129",	"r_xnm2130",	"r_xnm2131",	"r_xnm2132",	"r_xnm2133",	"r_xnm2134",	"r_xnm2135",	"r_xnm2136",	"r_xnm2137",	"r_xnm2138",	"r_xnm2139",	"r_xnm2141",	"r_xnm2142",	"r_xnm2143",	"r_xnm2144",	"r_xnm2145",	"r_xnm2146",	"r_xnm2147",	"r_xnm2148",	"r_xnm2149",	"r_xnm2150",	"r_xnm2151",	"r_xnm2152",	"r_xnm2153",	"r_xnm2154",	"r_xnm2155",	"r_xnm2156",	"r_xnm2157",	"r_xnm2158",	"r_xnm2159",	"r_xnm2160",	"r_xnm2161",	"r_xnm2162",	"r_xnm2163",	"r_xnm2164",	"r_xnm2165",	"r_xnm2166",	"r_xnm2167",	"r_xnm2168",	"r_xnm2169",	"r_xnm2170",	"r_xnm2171",	"r_xnm2172",	"r_xnm2173",	"r_xnm2174",	"r_xnm2175",	"r_xnm2176",	"r_xnm2177",	"r_xnm2178",	"r_xnm2179",	"r_xnm2180",	"r_xnm2181",	"r_xnm2182",	"r_xnm2183",	"r_xnm2184",	"r_xnm2185",	"r_xnm2186",	"r_xnm2187",	"r_xnm2188",	"r_xnm2189",	"r_xnm2190",	"r_xnm2191",	"r_xnm2192",	"r_xnm2193",	"r_xnm2194",	"r_xnm2195",	"r_xnm2196",	"r_xnm2201",	"r_xnm2202",	"r_xnm2203",	"r_xnm2204",	"r_xnm2205",	"r_xnm2206",	"r_xnm2211",	"r_xnm2212",	"r_xnm2213",	"r_xnm2214",	"r_xnm2215",	"r_xnm2216",	"r_xnm2218",	"r_xnm2219",	"r_xnm2220",	"r_xnm2221",	"r_xnm2222",	"r_xnm2223",	"r_xnm2224",	"r_xnm2225",	"r_xnm2226",	"r_xnm2227",	"r_xnm2228",	"r_xnm2229",	"r_xnm2230",	"r_xnm2231",	"r_xnm2232",	"r_xnm2233",	"r_xnm2234",	"r_xnm2235",	"r_xnm2236",	"r_xnm2237",	"r_xnm2238",	"r_xnm2239",	"r_xnm2241",	"r_xnm2243",	"r_xnm2244",	"r_xnm2245",	"r_xnm2246",	"r_xnm2247",	"r_xnm2248",	"r_xnm2249",	"r_xnm2250",	"r_xnm2251",	"r_xnm2252",	"r_xnm2253",	"r_xnm2254",	"r_xnm2255",	"r_xnm2256",	"r_xnm2257",	"r_xnm2258",	"r_xnm2259",	"r_xnm2260",	"r_xnm2261",	"r_xnm2262",	"r_xnm2263",	"r_xnm2264",	"r_xnm2265",	"r_xnm2266",	"r_xnm2267",	"r_xnm2268",	"r_xnm2269",	"r_xnm2270",	"r_xnm2271",	"r_xnm2272",	"r_xnm2273",	"r_xnm2274",	"r_xnm2275",	"r_xnm2276",	"r_xnm2277",	"r_xnm2278",	"r_xnm2279",	"r_xnm2280",	"r_xnm2281",	"r_xnm2282",	"r_xnm2283",	"r_xnm2284",	"r_xnm2285",	"r_xnm2286",	"r_xnm2287",	"r_xnm2288",	"r_xnm2289",	"r_xnm2290",	"r_xnm2291",	"r_xnm2292",	"r_xnm2293",	"r_xnm2294",	"r_xnm2295",	"r_xnm2296",	"r_xnm2297",	"r_xnm2298",	"r_xnm2299",	"r_xnm2300",	"r_xnm2301",	"r_xnm2302",	"r_xnm2303",	"r_xnm2304",	"r_xnm2305",	"r_xnm2306",	"r_xnm2307",	"r_xnm2308",	"r_xnm2309",	"r_xnm2310",	"r_xnm2311",	"r_xnm2312",	"r_xnm2313",	"r_xnm2314",	"r_xnm2315",	"r_xnm2316",	"r_xnm2317",	"r_xnm2318",	"r_xnm2319",	"r_xnm2320",	"r_xnm2321",	"r_xnm2322",	"r_xnm2323",	"r_xnm2324",	"r_xnm2325",	"r_xnm2326",	"r_xnm2327",	"r_xnm2328",	"r_xnm2329",	"r_xnm2330",	"r_xnm2331",	"r_xnm2332",	"r_xnm2333",	"r_xnm2334",	"r_xnm2335",	"r_xnm2336",	"r_xnm2337",	"r_xnm2338",	"r_xnm2339",	"r_xnm2340",	"r_xnm2341",	"r_xnm2342",	"r_xnm2343",	"r_xnm2344",	"r_xnm2345",	"r_xnm2346",	"r_xnm2347",	"r_xnm2348",	"r_xnm2349",	"r_xnm2350",	"r_xnm2351",	"r_xnm2352",	"r_xnm2353",	"r_xnm2354",	"r_xnm2355",	"r_xnm2356",	"r_xnm2357",	"r_xnm2358",	"r_xnm2359",	"r_xnm2360",	"r_xnm2361",	"r_xnm2362",	"r_xnm2363",	"r_xnm2364",	"r_xnm2365",	"r_xnm2366",	"r_xnm2367",	"r_xnm2368",	"r_xnm2369",	"r_xnm2370",	"r_xnm2371",	"r_xnm2372",	"r_xnm2373",	"r_xnm2374",	"r_xnm2375",	"r_xnm2376",	"r_xnm2377",	"r_xnm2378",	"r_xnm2379",	"r_xnm2380",	"r_xnm2381",	"r_xnm2382",	"r_xnm2383",	"r_xnm2384",	"r_xnm2385",	"r_xnm2386",	"r_xnm2387",	"r_xnm2388",	"r_xnm2389",	"r_xnm2390",	"r_xnm2391",	"r_xnm2392",	"r_xnm2393",	"r_xnm2394",	"r_xnm2395",	"r_xnm2396",	"r_xnm2397",	"r_xnm2398",	"r_xnm2399",	"r_xnm2400",	"r_xnm2401",	"r_xnm2402",	"r_xnm2403",	"r_xnm2404",	"r_xnm2405",	"r_xnm2406",	"r_xnm2407",	"r_xnm2408",	"r_xnm2409",	"r_xnm2410",	"r_xnm2411",	"r_xnm2412",	"r_xnm2413",	"r_xnm2414",	"r_xnm2415",	"r_xnm2416",	"r_xnm2417",	"r_xnm2418",	"r_xnm2419",	"r_xnm2420",	"r_xnm2421",	"r_xnm2422",	"r_xnm2423",	"r_xnm2424",	"r_xnm2425",	"r_xnm2426",	"r_xnm2427",	"r_xnm2428",	"r_xnm2429",	"r_xnm2430",	"r_xnm2431",	"r_xnm2432",	"r_xnm2433",	"r_xnm2434",	"r_xnm2435",	"r_xnm2436",	"r_xnm2437",	"r_xnm2438",	"r_xnm2439",	"r_xnm2440",	"r_xnm2441",	"r_xnm2442",	"r_xnm2443",	"r_xnm2444",	"r_xnm2445",	"r_xnm2446",	"r_xnm2447",	"r_xnm2448",	"r_xnm2449",	"r_xnm2450",	"r_xnm2451",	"r_xnm2452",	"r_xnm2453",	"r_xnm2454",	"r_xnm2455",	"r_xnm2456",	"r_xnm2457",	"r_xnm2458",	"r_xnm2459",	"r_xnm2460",	"r_xnm2461",	"r_xnm2462",	"r_xnm2463",	"r_xnm2464",	"r_xnm2465",	"r_xnm2466",	"r_xnm2467",	"r_xnm2468",	"r_xnm2469",	"r_xnm2470",	"r_xnm2471",	"r_xnm2472",	"r_xnm2473",	"r_xnm2474",	"r_xnm2475",	"r_xnm2476",	"r_xnm2477",	"r_xnm2478",	"r_xnm2479",	"r_xnm2480",	"r_xnm2481",	"r_xnm2482",	"r_xnm2483",	"r_xnm2484",	"r_xnm2485",	"r_xnm2486",	"r_xnm2487",	"r_xnm2488",	"r_xnm2489",	"r_xnm2490",	"r_xnm2491",	"r_xnm2492",	"r_xnm2493",	"r_xnm2494",	"r_xnm2495",	"r_xnm2496",	"r_xnm2497",	"r_xnm2498",	"r_xnm2499",	"r_xnm2500",	"r_xnm2501",	"r_xnm2502",	"r_xnm2503",	"r_xnm2504",	"r_xnm2505",	"r_xnm2506",	"r_xnm2507",	"r_xnm2508",	"r_xnm2509",	"r_xnm2510",	"r_xnm2511",	"r_xnm2512",	"r_xnm2513",	"r_xnm2514",	"r_xnm2515",	"r_xnm2516",	"r_xnm2517",	"r_xnm2518",	"r_xnm2519",	"r_xnm2520",	"r_xnm2521",	"r_xnm2522",	"r_xnm2523",	"r_xnm2524",	"r_xnm2525",	"r_xnm2526",	"r_xnm2527",	"r_xnm2528",	"r_xnm2529",	"r_xnm2530",	"r_xnm2531",	"r_xnm2532",	"r_xnm2535",	"r_xnm2536",	"r_xnm2537",	"r_xnm2538",	"r_xnm2539",	"r_xnm2540",	"r_xnm2541",	"r_xnm2542",	"r_xnm2543",	"r_xnm2544",	"r_xnm2545",	"r_xnm2546",	"r_xnm2547",	"r_xnm2548",	"r_xnm2549",	"r_xnm2550",	"r_xnm2551",	"r_xnm2552",	"r_xnm2553",	"r_xnm2554",	"r_xnm2555",	"r_xnm2556",	"r_xnm2557",	"r_xnm2558",	"r_xnm2559",	"r_xnm2560",	"r_xnm2561",	"r_xnm2562",	"r_xnm2563",	"r_xnm2564",	"r_xnm2565",	"r_xnm2566",	"r_xnm2567",	"r_xnm2568",	"r_xnm2569",	"r_xnm2570",	"r_xnm2571",	"r_xnm2572",	"r_xnm2573",	"r_xnm2574",	"r_xnm2575",	"r_xnm2576",	"r_xnm2577",	"r_xnm2578",	"r_xnm2579",	"r_xnm2580",	"r_xnm2581",	"r_xnm2582",	"r_xnm2583",	"r_xnm2584",	"r_xnm2585",	"r_xnm2586",	"r_xnm2587",	"r_xnm2588",	"r_xnm2589",	"r_xnm2590",	"r_xnm2591",	"r_xnm2592",	"r_xnm2593",	"r_xnm2594",	"r_xnm2595",	"r_xnm2596",	"r_xnm2597",	"r_xnm2598",	"r_xnm2599",	"r_xnm2600",	"r_xnm2601",	"r_xnm2602",	"r_xnm2603",	"r_xnm2604",	"r_xnm2605",	"r_xnm2606",	"r_xnm2607",	"r_xnm2608",	"r_xnm2609",	"r_xnm2610",	"r_xnm2611",	"r_xnm2612",	"r_xnm2613",	"r_xnm2614",	"r_xnm2615",	"r_xnm2616",	"r_xnm2617",	"r_xnm2618",	"r_xnm2619",	"r_xnm2620",	"r_xnm2621",	"r_xnm2622",	"r_xnm2623",	"r_xnm2624",	"r_xnm2625",	"r_xnm2626",	"r_xnm2627",	"r_xnm2628",	"r_xnm2629",	"r_xnm2630",	"r_xnm2631",	"r_xnm2632",	"r_xnm2633",	"r_xnm2634",	"r_xnm2635",	"r_xnm2636",	"r_xnm2637",	"r_xnm2638",	"r_xnm2639",	"r_xnm2640",	"r_xnm2641",	"r_xnm2642",	"r_xnm2643",	"r_xnm2644",	"r_xnm2645",	"r_xnm2646",	"r_xnm2647",	"r_xnm2648",	"r_xnm2649",	"r_xnm2650",	"r_xnm2651",	"r_xnm2652",	"r_xnm2653",	"r_xnm2654",	"r_xnm2655",	"r_xnm2656",	"r_xnm2657",	"r_xnm2658",	"r_xnm2659",	"r_xnm2660",	"r_xnm2661",	"r_xnm2662",	"r_xnm2663",	"r_xnm2664",	"r_xnm2665",	"r_xnm2666",	"r_xnm2667",	"r_xnm2668",	"r_xnm2669",	"r_xnm2670",	"r_xnm2671",	"r_xnm2672",	"r_xnm2673",	"r_xnm2674",	"r_xnm2675",	"r_xnm2676",	"r_xnm2677",	"r_xnm2678",	"r_xnm2679",	"r_xnm2680",	"r_xnm2681",	"r_xnm2682",	"r_xnm2683",	"r_xnm2684",	"r_xnm2685",	"r_xnm2686",	"r_xnm2687",	"r_xnm2688",	"r_xnm2689",	"r_xnm2690",	"r_xnm2691",	"r_xnm2692",	"r_xnm2693",	"r_xnm2694",	"r_xnm2695",	"r_xnm2696",	"r_xnm2697",	"r_xnm2698",	"r_xnm2699",	"r_xnm2700",	"r_xnm2701",	"r_xnm2702",	"r_xnm2703",	"r_xnm2704",	"r_xnm2705",	"r_xnm2706",	"r_xnm2707",	"r_xnm2708",	"r_xnm2709",	"r_xnm2710",	"r_xnm2711",	"r_xnm2712",	"r_xnm2713",	"r_xnm2714",	"r_xnm2715",	"r_xnm2716",	"r_xnm2717",	"r_xnm2718",	"r_xnm2719",	"r_xnm2720",	"r_xnm2721",	"r_xnm2722",	"r_xnm2723",	"r_xnm2724",	"r_xnm2725",	"r_xnm2726",	"r_xnm2727",	"r_xnm2728",	"r_xnm2729",	"r_xnm2730",	"r_xnm2731",	"r_xnm2732",	"r_xnm2733",	"r_xnm2734",	"r_xnm2735",	"r_xnm2736",	"r_xnm2737",	"r_xnm2738",	"r_xnm2739",	"r_xnm2740",	"r_xnm2741",	"r_xnm2742",	"r_xnm2743",	"r_xnm2744",	"r_xnm2745",	"r_xnm2746",	"r_xnm2747",	"r_xnm2748",	"r_xnm2749",	"r_xnm2750",	"r_xnm2751",	"r_xnm2752",	"r_xnm2753",	"r_xnm2754",	"r_xnm2755",	"r_xnm2756",	"r_xnm2757",	"r_xnm2758",	"r_xnm2759",	"r_xnm2760",	"r_xnm2761",	"r_xnm2762",	"r_xnm2763",	"r_xnm2764",	"r_xnm2765",	"r_xnm2766",	"r_xnm2767",	"r_xnm2768",	"r_xnm2769",	"r_xnm2770",	"r_xnm2771",	"r_xnm2772",	"r_xnm2773",	"r_xnm2774",	"r_xnm2775",	"r_xnm2776",	"r_xnm2777",	"r_xnm2778",	"r_xnm2779",	"r_xnm2780",	"r_xnm2781",	"r_xnm2782",	"r_xnm2783",	"r_xnm2784",	"r_xnm2785",	"r_xnm2786",	"r_xnm2787",	"r_xnm2788",	"r_xnm2789",	"r_xnm2790",	"r_xnm2791",	"r_xnm2792",	"r_xnm2793",	"r_xnm2794",	"r_xnm2795",	"r_xnm2796",	"r_xnm2797",	"r_xnm2798",	"r_xnm2799",	"r_xnm2800",	"r_xnm2801",	"r_xnm2802",	"r_xnm2803",	"r_xnm2804",	"r_xnm2805",	"r_xnm2806",	"r_xnm2807",	"r_xnm2808",	"r_xnm2809",	"r_xnm2810",	"r_xnm2811",	"r_xnm2812",	"r_xnm2813",	"r_xnm2814",	"r_xnm2815",	"r_xnm2816",	"r_xnm2817",	"r_xnm2818",	"r_xnm2819",	"r_xnm2820",	"r_xnm2821",	"r_xnm2822",	"r_xnm2823",	"r_xnm2824",	"r_xnm2825",	"r_xnm2826",	"r_xnm2827",	"r_xnm2828",	"r_xnm2829",	"r_xnm2830",	"r_xnm2831",	"r_xnm2832",	"r_xnm2833",	"r_xnm2834",	"r_xnm2835",	"r_xnm2836",	"r_xnm2837",	"r_xnm2838",	"r_xnm2839",	"r_xnm2840",	"r_xnm2841",	"r_xnm2842",	"r_xnm2843",	"r_xnm2844",	"r_xnm2845",	"r_xnm2846",	"r_xnm2847",	"r_xnm2848",	"r_xnm2849",	"r_xnm2850",	"r_xnm2851",	"r_xnm2852",	"r_xnm2853",	"r_xnm2854",	"r_xnm2855",	"r_xnm2856",	"r_xnm2857",	"r_xnm2858",	"r_xnm2859",	"r_xnm2860",	"r_xnm2861",	"r_xnm2862",	"r_xnm2863",	"r_xnm2864",	"r_xnm2865",	"r_xnm2866",	"r_xnm2867",	"r_xnm2868",	"r_xnm2869",	"r_xnm2870",	"r_xnm2871",	"r_xnm2872",	"r_xnm2873",	"r_xnm2874",	"r_xnm2875",	"r_xnm2876",	"r_xnm2877",	"r_xnm2878",	"r_xnm2879",	"r_xnm2880",	"r_xnm2881",	"r_xnm2882",	"r_xnm2883",	"r_xnm2884",	"r_xnm2885",	"r_xnm2886",	"r_xnm2887",	"r_xnm2888",	"r_xnm2889",	"r_xnm2890",	"r_xnm2891",	"r_xnm2892",	"r_xnm2893",	"r_xnm2894",	"r_xnm2895",	"r_xnm2896",	"r_xnm2897",	"r_xnm2898",	"r_xnm2899",	"r_xnm2900",	"r_xnm2901",	"r_xnm2902",	"r_xnm2903",	"r_xnm2904",	"r_xnm2905",	"r_xnm2906",	"r_xnm2907",	"r_xnm2908",	"r_xnm2909",	"r_xnm2910",	"r_xnm2911",	"r_xnm2912",	"r_xnm2913",	"r_xnm2914",	"r_xnm2915",	"r_xnm2916",	"r_xnm2917",	"r_xnm2918",	"r_xnm2919",	"r_xnm2920",	"r_xnm2921",	"r_xnm2922",	"r_xnm2923",	"r_xnm2924",	"r_xnm2925",	"r_xnm2926",	"r_xnm2927",	"r_xnm2928",	"r_xnm2929",	"r_xnm2930",	"r_xnm2931",	"r_xnm2932",	"r_xnm2933",	"r_xnm2934",	"r_xnm2935",	"r_xnm2936",	"r_xnm2937",	"r_xnm2938",	"r_xnm2939",	"r_xnm2940",	"r_xnm2941",	"r_xnm2942",	"r_xnm2943",	"r_xnm2944",	"r_xnm2945",	"r_xnm2946",	"r_xnm2947",	"r_xnm2948",	"r_xnm2949",	"r_xnm2950",	"r_xnm2951",	"r_xnm2952",	"r_xnm2953",	"r_xnm2954",	"r_xnm2955",	"r_xnm2956",	"r_xnm2957",	"r_xnm2958",	"r_xnm2959",	"r_xnm2960",	"r_xnm2961",	"r_xnm2962",	"r_xnm2963",	"r_xnm2964",	"r_xnm2965",	"r_xnm2966",	"r_xnm2967",	"r_xnm2968",	"r_xnm2969",	"r_xnm2970",	"r_xnm2971",	"r_xnm2972",	"r_xnm2973",	"r_xnm2974",	"r_xnm2975",	"r_xnm2976",	"r_xnm2977",	"r_xnm2978",	"r_xnm2979",	"r_xnm2980",	"r_xnm2981",	"r_xnm2982",	"r_xnm2983",	"r_xnm2984",	"r_xnm2985",	"r_xnm2986",	"r_xnm2987",	"r_xnm2988",	"r_xnm2989",	"r_xnm2990",	"r_xnm2991",	"r_xnm2992",	"r_xnm2993",	"r_xnm2994",	"r_xnm2995",	"r_xnm2996",	"r_xnm2997",	"r_xnm2998",	"r_xnm2999",	"r_xnm3000",	"r_xnm3001",	"r_xnm3002",	"r_xnm3003",	"r_xnm3004",	"r_xnm3005",	"r_xnm3006",	"r_xnm3007",	"r_xnm3008",	"r_xnm3009",	"r_xnm3010",	"r_xnm3011",	"r_xnm3012",	"r_xnm3013",	"r_xnm3014",	"r_xnm3015",	"r_xnm3016",	"r_xnm3017",	"r_xnm3018",	"r_xnm3019",	"r_xnm3020",	"r_xnm3021",	"r_xnm3022",	"r_xnm3023",	"r_xnm3024",	"r_xnm3025",	"r_xnm3026",	"r_xnm3027",	"r_xnm3028",	"r_xnm3029",	"r_xnm3030",	"r_xnm3031",	"r_xnm3032",	"r_xnm3033",	"r_xnm3034",	"r_xnm3035",	"r_xnm3036",	"r_xnm3037",	"r_xnm3038",	"r_xnm3039",	"r_xnm3040",	"r_xnm3041",	"r_xnm3042",	"r_xnm3043",	"r_xnm3044",	"r_xnm3045",	"r_xnm3046",	"r_xnm3047",	"r_xnm3048",	"r_xnm3049",	"r_xnm3050",	"r_xnm3051",	"r_xnm3052",	"r_xnm3053",	"r_xnm3054",	"r_xnm3055",	"r_xnm3056",	"r_xnm3057",	"r_xnm3058",	"r_xnm3059",	"r_xnm3060",	"r_xnm3061",	"r_xnm3062",	"r_xnm3063",	"r_xnm3064",	"r_xnm3065",	"r_xnm3066",	"r_xnm3067",	"r_xnm3068",	"r_xnm3069",	"r_xnm3070",	"r_xnm3071",	"r_xnm3072",	"r_xnm3073",	"r_xnm3074",	"r_xnm3075",	"r_xnm3076",	"r_xnm3077",	"r_xnm3078",	"r_xnm3079",	"r_xnm3080",	"r_xnm3081",	"r_xnm3082",	"r_xnm3083",	"r_xnm3084",	"r_xnm3085",	"r_xnm3086",	"r_xnm3087",	"r_xnm3088",	"r_xnm3089",	"r_xnm3090",	"r_xnm3091",	"r_xnm3092",	"r_xnm3093",	"r_xnm3094",	"r_xnm3095",	"r_xnm3096",	"r_xnm3097",	"r_xnm3098",	"r_xnm3099",	"r_xnm3100",	"r_xnm3101",	"r_xnm3102",	"r_xnm3103",	"r_xnm3104",	"r_xnm3105",	"r_xnm3106",	"r_xnm3107",	"r_xnm3108",	"r_xnm3109",	"r_xnm3110",	"r_xnm3111",	"r_xnm3112",	"r_xnm3113",	"r_xnm3114",	"r_xnm3115",	"r_xnm3116",	"r_xnm3117",	"r_xnm3118",	"r_xnm3119",	"r_xnm3120",	"r_xnm3121",	"r_xnm3122",	"r_xnm3123",	"r_xnm3124",	"r_xnm3125",	"r_xnm3126",	"r_xnm3127",	"r_xnm3128",	"r_xnm3129",	"r_xnm3130",	"r_xnm3131",	"r_xnm3132",	"r_xnm3133",	"r_xnm3134",	"r_xnm3135",	"r_xnm3136",	"r_xnm3137",	"r_xnm3138",	"r_xnm3139",	"r_xnm3140",	"r_xnm3141",	"r_xnm3142",	"r_xnm3143",	"r_xnm3144",	"r_xnm3145",	"r_xnm3146",	"r_xnm3147",	"r_xnm3148",	"r_xnm3149",	"r_xnm3150",	"r_xnm3151",	"r_xnm3152",	"r_xnm3153",	"r_xnm3154",	"r_xnm3155",	"r_xnm3156",	"r_xnm3157",	"r_xnm3158",	"r_xnm3159",	"r_xnm3160",	"r_xnm3161",	"r_xnm3162",	"r_xnm3163",	"r_xnm3164",	"r_xnm3165",	"r_xnm3166",	"r_xnm3167",	"r_xnm3168",	"r_xnm3169",	"r_xnm3170",	"r_xnm3171",	"r_xnm3172",	"r_xnm3173",	"r_xnm3174",	"r_xnm3175",	"r_xnm3176",	"r_xnm3177",	"r_xnm3178",	"r_xnm3179",	"r_xnm3180",	"r_xnm3181",	"r_xnm3182",	"r_xnm3183",	"r_xnm3184",	"r_xnm3185",	"r_xnm3186",	"r_xnm3187",	"r_xnm3188",	"r_xnm3189",	"r_xnm3190",	"r_xnm3191",	"r_xnm3192",	"r_xnm3193",	"r_xnm3194",	"r_xnm3195",	"r_xnm3196",	"r_xnm3197",	"r_xnm3198",	"r_xnm3199",	"r_xnm3200",	"r_xnm3201",	"r_xnm3202",	"r_xnm3203",	"r_xnm3204",	"r_xnm3205",	"r_xnm3206",	"r_xnm3207",	"r_xnm3208",	"r_xnm3209",	"r_xnm3210",	"r_xnm3211",	"r_xnm3212",	"r_xnm3213",	"r_xnm3214",	"r_xnm3215",	"r_xnm3216",	"r_xnm3217",	"r_xnm3218",	"r_xnm3219",	"r_xnm3220",	"r_xnm3221",	"r_xnm3222",	"r_xnm3223",	"r_xnm3224",	"r_xnm3225",	"r_xnm3226",	"r_xnm3227",	"r_xnm3228",	"r_xnm3229",	"r_xnm3230",	"r_xnm3231",	"r_xnm3232",	"r_xnm3233",	"r_xnm3234",	"r_xnm3235",	"r_xnm3236",	"r_xnm3237",	"r_xnm3238",	"r_xnm3239",	"r_xnm3240",	"r_xnm3241",	"r_xnm3242",	"r_xnm3243",	"r_xnm3244",	"r_xnm3245",	"r_xnm3246",	"r_xnm3247",	"r_xnm3248",	"r_xnm3249",	"r_xnm3250",	"r_xnm3251",	"r_xnm3252",	"r_xnm3253",	"r_xnm3254",	"r_xnm3255",	"r_xnm3256",	"r_xnm3257",	"r_xnm3258",	"r_xnm3259",	"r_xnm3260",	"r_xnm3261",	"r_xnm3262",	"r_xnm3263",	"r_xnm3264",	"r_xnm3265",	"r_xnm3266",	"r_xnm3267",	"r_xnm3268",	"r_xnm3269",	"r_xnm3270",	"r_xnm3271",	"r_xnm3272",	"r_xnm3273",	"r_xnm3274",	"r_xnm3275",	"r_xnm3276",	"r_xnm3277",	"r_xnm3278",	"r_xnm3279",	"r_xnm3280",	"r_xnm3281",	"r_xnm3282",	"r_xnm3283",	"r_xnm3284",	"r_xnm3285",	"r_xnm3286",	"r_xnm3287",	"r_xnm3288",	"r_xnm3289",	"r_xnm3290",	"r_xnm3291",	"r_xnm3292",	"r_xnm3293",	"r_xnm3294",	"r_xnm3295",	"r_xnm3296",	"r_xnm3297",	"r_xnm3298",	"r_xnm3299",	"r_xnm3300",	"r_xnm3301",	"r_xnm3302",	"r_xnm3303",	"r_xnm3304",	"r_xnm3305",	"r_xnm3306",	"r_xnm3307",	"r_xnm3308",	"r_xnm3309",	"r_xnm3310",	"r_xnm3311",	"r_xnm3312",	"r_xnm3313",	"r_xnm3314",	"r_xnm3315",	"r_xnm3316",	"r_xnm3317",	"r_xnm3318",	"r_xnm3319",	"r_xnm3320",	"r_xnm3321",	"r_xnm3322",	"r_xnm3323",	"r_xnm3324",	"r_xnm3325",	"r_xnm3326",	"r_xnm3327",	"r_xnm3328",	"r_xnm3329",	"r_xnm3330",	"r_xnm3331",	"r_xnm3332",	"r_xnm3333",	"r_xnm3334",	"r_xnm3335",	"r_xnm3336",	"r_xnm3337",	"r_xnm3338",	"r_xnm3339",	"r_xnm3340",	"r_xnm3341",	"r_xnm3342",	"r_xnm3343",	"r_xnm3344",	"r_xnm3345",	"r_xnm3346",	"r_xnm3347",	"r_xnm3348",	"r_xnm3349",	"r_xnm3350",	"r_xnm3351",	"r_xnm3352",	"r_xnm3353",	"r_xnm3354",	"r_xnm3355",	"r_xnm3356",	"r_xnm3357",	"r_xnm3358",	"r_xnm3359",	"r_xnm3360",	"r_xnm3361",	"r_xnm3362",	"r_xnm3363",	"r_xnm3364",	"r_xnm3365",	"r_xnm3366",	"r_xnm3367",	"r_xnm3368",	"r_xnm3369",	"r_xnm3370",	"r_xnm3371",	"r_xnm3372",	"r_xnm3373",	"r_xnm3374",	"r_xnm3375",	"r_xnm3376",	"r_xnm3377",	"r_xnm3378",	"r_xnm3379",	"r_xnm3380",	"r_xnm3381",	"r_xnm3382",	"r_xnm3383",	"r_xnm3384",	"r_xnm3385",	"r_xnm3386",	"r_xnm3387",	"r_xnm3388",	"r_xnm3389",	"r_xnm3390",	"r_xnm3391",	"r_xnm3392",	"r_xnm3393",	"r_xnm3394",	"r_xnm3395",	"r_xnm3396",	"r_xnm3397",	"r_xnm3398",	"r_xnm3399",	"r_xnm3400",	"r_xnm3401",	"r_xnm3402",	"r_xnm3403",	"r_xnm3404",	"r_xnm3405",	"r_xnm3406",	"r_xnm3407",	"r_xnm3408",	"r_xnm3409",	"r_xnm3410",	"r_xnm3411",	"r_xnm3412",	"r_xnm3413",	"r_xnm3414",	"r_xnm3415",	"r_xnm3416",	"r_xnm3417",	"r_xnm3418",	"r_xnm3419",	"r_xnm3420",	"r_xnm3421",	"r_xnm3422",	"r_xnm3423",	"r_xnm3424",	"r_xnm3425",	"r_xnm3426",	"r_xnm3427",	"r_xnm3428",	"r_xnm3429",	"r_xnm3430",	"r_xnm3431",	"r_xnm3432",	"r_xnm3433",	"r_xnm3434",	"r_xnm3435",	"r_xnm3436",	"r_xnm3437",	"r_xnm3438",	"r_xnm3439",	"r_xnm3440",	"r_xnm3441",	"r_xnm3442",	"r_xnm3443",	"r_xnm3444",	"r_xnm3445",	"r_xnm3446",	"r_xnm3447",	"r_xnm3448",	"r_xnm3449",	"r_xnm3450",	"r_xnm3451",	"r_xnm3452",	"r_xnm3453",	"r_xnm3454",	"r_xnm3455",	"r_xnm3456",	"r_xnm3457",	"r_xnm3458",	"r_xnm3459",	"r_xnm3460",	"r_xnm3461",	"r_xnm3462",	"r_xnm3463",	"r_xnm3464",	"r_xnm3465",	"r_xnm3466",	"r_xnm3467",	"r_xnm3468",	"r_xnm3469",	"r_xnm3470",	"r_xnm3471",	"r_xnm3472",	"r_xnm3473",	"r_xnm3474",	"r_xnm3475",	"r_xnm3476",	"r_xnm3477",	"r_xnm3478",	"r_xnm3479",	"r_xnm3480",	"r_xnm3481",	"r_xnm3482",	"r_xnm3483",	"r_xnm3484",	"r_xnm3485",	"r_xnm3486",	"r_xnm3487",	"r_xnm3488",	"r_xnm3489",	"r_xnm3490",	"r_xnm3491",	"r_xnm3492",	"r_xnm3493",	"r_xnm3494",	"r_xnm3495",	"r_xnm3496",	"r_xnm3497",	"r_xnm3498",	"r_xnm3499",	"r_xnm3500",	"r_xnm3501",	"r_xnm3502",	"r_xnm3503",	"r_xnm3504",	"r_xnm3505",	"r_xnm3506",	"r_xnm3507",	"r_xnm3508",	"r_xnm3509",	"r_xnm3510",	"r_xnm3511",	"r_xnm3512",	"r_xnm3513",	"r_xnm3514",	"r_xnm3515",	"r_xnm3516",	"r_xnm3517",	"r_xnm3518",	"r_xnm3519",	"r_xnm3520",	"r_xnm3521",	"r_xnm3522",	"r_xnm3523",	"r_xnm3524",	"r_xnm3525",	"r_xnm3526",	"r_xnm3527",	"r_xnm3528",	"r_xnm3529",	"r_xnm3530",	"r_xnm3531",	"r_xnm3532",	"r_xnm3533",	"r_xnm3534",	"r_xnm3535",	"r_xnm3536",	"r_xnm3537",	"r_xnm3538",	"r_xnm3539",	"r_xnm3540",	"r_xnm3541",	"r_xnm3542",	"r_xnm3543",	"r_xnm3544",	"r_xnm3545",	"r_xnm3546",	"r_xnm3547",	"r_xnm3548",	"r_xnm3549",	"r_xnm3550",	"r_xnm3551",	"r_xnm3552",	"r_xnm3553",	"r_xnm3554",	"r_xnm3555",	"r_xnm3556",	"r_xnm3557",	"r_xnm3558",	"r_xnm3559",	"r_xnm3560",	"r_xnm3561",	"r_xnm3562",	"r_xnm3563",	"r_xnm3564",	"r_xnm3565",	"r_xnm3566",	"r_xnm3567",	"r_xnm3568",	"r_xnm3569",	"r_xnm3570",	"r_xnm3571",	"r_xnm3572",	"r_xnm3573",	"r_xnm3574",	"r_xnm3575",	"r_xnm3576",	"r_xnm3577",	"r_xnm3578",	"r_xnm3579",	"r_xnm3580",	"r_xnm3581",	"r_xnm3582",	"r_xnm3583",	"r_xnm3584",	"r_xnm3585",	"r_xnm3586",	"r_xnm3587",	"r_xnm3588",	"r_xnm3589",	"r_xnm3590",	"r_xnm3591",	"r_xnm3592",	"r_xnm3593",	"r_xnm3594",	"r_xnm3595",	"r_xnm3596",	"r_xnm3597",	"r_xnm3598",	"r_xnm3599",	"r_xnm3600",	"r_xnm3601",	"r_xnm3602",	"r_xnm3603",	"r_xnm3604",	"r_xnm3605",	"r_xnm3606",	"r_xnm3607",	"r_xnm3608",	"r_xnm3609",	"r_xnm3610",	"r_xnm3611",	"r_xnm3612",	"r_xnm3613",	"r_xnm3614",	"r_xnm3615",	"r_xnm3616",	"r_xnm3617",	"r_xnm3618",	"r_xnm3619",	"r_xnm3620",	"r_xnm3621",	"r_xnm3622",	"r_xnm3623",	"r_xnm3624",	"r_xnm3625",	"r_xnm3626",	"r_xnm3627",	"r_xnm3628",	"r_xnm3629",	"r_xnm3630",	"r_xnm3631",	"r_xnm3632",	"r_xnm3633",	"r_xnm3634",	"r_xnm3635",	"r_xnm3636",	"r_xnm3637",	"r_xnm3638",	"r_xnm3639",	"r_xnm3640",	"r_xnm3641",	"r_xnm3642",	"r_xnm3643",	"r_xnm3644",	"r_xnm3645",	"r_xnm3646",	"r_xnm3647",	"r_xnm3648",	"r_xnm3649",	"r_xnm3650",	"r_xnm3651",	"r_xnm3652",	"r_xnm3653",	"r_xnm3654",	"r_xnm3655",	"r_xnm3656",	"r_xnm3657",	"r_xnm3658",	"r_xnm3659",	"r_xnm3660",	"r_xnm3661",	"r_xnm3662",	"r_xnm3663",	"r_xnm3664",	"r_xnm3665",	"r_xnm3666",	"r_xnm3667",	"r_xnm3668",	"r_xnm3669",	"r_xnm3670",	"r_xnm3671",	"r_xnm3672",	"r_xnm3673",	"r_xnm3674",	"r_xnm3675",	"r_xnm3676",	"r_xnm3677",	"r_xnm3678",	"r_xnm3679",	"r_xnm3680",	"r_xnm3681",	"r_xnm3682",	"r_xnm3683",	"r_xnm3684",	"r_xnm3685",	"r_xnm3686",	"r_xnm3687",	"r_xnm3688",	"r_xnm3689",	"r_xnm3690",	"r_xnm3691",	"r_xnm3692",	"r_xnm3693",	"r_xnm3694",	"r_xnm3695",	"r_xnm3696",	"r_xnm3697",	"r_xnm3698",	"r_xnm3699",	"r_xnm3700",	"r_xnm3701",	"r_xnm3702",	"r_xnm3703",	"r_xnm3704",	"r_xnm3705",	"r_xnm3706",	"r_xnm3707",	"r_xnm3708",	"r_xnm3709",	"r_xnm3710",	"r_xnm3711",	"r_xnm3712",	"r_xnm3713",	"r_xnm3714",	"r_xnm3715",	"r_xnm3716",	"r_xnm3717",	"r_xnm3718",	"r_xnm3719",	"r_xnm3720",	"r_xnm3721",	"r_xnm3722",	"r_xnm3723",	"r_xnm3724",	"r_xnm3725",	"r_xnm3726",	"r_xnm3727",	"r_xnm3728",	"r_xnm3729",	"r_xnm3730",	"r_xnm3731",	"r_xnm3732",	"r_xnm3733",	"r_xnm3734",	"r_xnm3735",	"r_xnm3736",	"r_xnm3737",	"r_xnm3738",	"r_xnm3739",	"r_xnm3740",	"r_xnm3741",	"r_xnm3742",	"r_xnm3743",	"r_xnm3744",	"r_xnm3745",	"r_xnm3746",	"r_xnm3747",	"r_xnm3748",	"r_xnm3749",	"r_xnm3750",	"r_xnm3751",	"r_xnm3752",	"r_xnm3753",	"r_xnm3754",	"r_xnm3755",	"r_xnm3756",	"r_xnm3757",	"r_xnm3758",	"r_xnm3759",	"r_xnm3760",	"r_xnm3761",	"r_xnm3762",	"r_xnm3763",	"r_xnm3764",	"r_xnm3765",	"r_xnm3766",	"r_xnm3767",	"r_xnm3768",	"r_xnm3769",	"r_xnm3770",	"r_xnm3771",	"r_xnm3772",	"r_xnm3773",	"r_xnm3774",	"r_xnm3775",	"r_xnm3776",	"r_xnm3777",	"r_xnm3778",	"r_xnm3779",	"r_xnm3780",	"r_xnm3781",	"r_xnm3782",	"r_xnm3783",	"r_xnm3784",	"r_xnm3785",	"r_xnm3786",	"r_xnm3787",	"r_xnm3788",	"r_xnm3789",	"r_xnm3790",	"r_xnm3791",	"r_xnm3792",	"r_xnm3793",	"r_xnm3794",	"r_xnm3795",	"r_xnm3796",	"r_xnm3797",	"r_xnm3798",	"r_xnm3799",	"r_xnm3800",	"r_xnm3801",	"r_xnm3802",	"r_xnm3803",	"r_xnm3804",	"r_xnm3805",	"r_xnm3806",	"r_xnm3807",	"r_xnm3808",	"r_xnm3809",	"r_xnm3810",	"r_xnm3811",	"r_xnm3812",	"r_xnm3813",	"r_xnm3814",	"r_xnm3815",	"r_xnm3816",	"r_xnm3817",	"r_xnm3818",	"r_xnm3819",	"r_xnm3820",	"r_xnm3821",	"r_xnm3822",	"r_xnm3823",	"r_xnm3824",	"r_xnm3825",	"r_xnm3826",	"r_xnm3827",	"r_xnm3828",	"r_xnm3829",	"r_xnm3830",	"r_xnm3831",	"r_xnm3832",	"r_xnm3833",	"r_xnm3834",	"r_xnm3835",	"r_xnm3836",	"r_xnm3837",	"r_xnm3838",	"r_xnm3839",	"r_xnm3840",	"r_xnm3841",	"r_xnm3842",	"r_xnm3843",	"r_xnm3844",	"r_xnm3845",	"r_xnm3846",	"r_xnm3847",	"r_xnm3848",	"r_xnm3849",	"r_xnm3850",	"r_xnm3851",	"r_xnm3852",	"r_xnm3853",	"r_xnm3854",	"r_xnm3855",	"r_xnm3856",	"r_xnm3857",	"r_xnm3858",	"r_xnm3859",	"r_xnm3860",	"r_xnm3861",	"r_xnm3862",	"r_xnm3863",	"r_xnm3864",	"r_xnm3865",	"r_xnm3866",	"r_xnm3867",	"r_xnm3868",	"r_xnm3869",	"r_xnm3870",	"r_xnm3871",	"r_xnm3872",	"r_xnm3873",	"r_xnm3874",	"r_xnm3875",	"r_xnm3876",	"r_xnm3877",	"r_xnm3878",	"r_xnm3879",	"r_xnm3880",	"r_xnm3881",	"r_xnm3882",	"r_xnm3883",	"r_xnm3884",	"r_xnm3885",	"r_xnm3886",	"r_xnm3887",	"r_xnm3888",	"r_xnm3889",	"r_xnm3890",	"r_xnm3891",	"r_xnm3892",	"r_xnm3893",	"r_xnm3894",	"r_xnm3895",	"r_xnm3896",	"r_xnm3897",	"r_xnm3898",	"r_xnm3899",	"r_xnm3900",	"r_xnm3901",	"r_xnm3902",	"r_xnm3903",	"r_xnm3904",	"r_xnm3905",	"r_xnm3906",	"r_xnm3907",	"r_xnm3908",	"r_xnm3909",	"r_xnm3910",	"r_xnm3911",	"r_xnm3912",	"r_xnm3913",	"r_xnm3914",	"r_xnm3915",	"r_xnm3916",	"r_xnm3917",	"r_xnm3918",	"r_xnm3919",	"r_xnm3920",	"r_xnm3921",	"r_xnm3922",	"r_xnm3923",	"r_xnm3924",	"r_xnm3925",	"r_xnm3926",	"r_xnm3927",	"r_xnm3928",	"r_xnm3929",	"r_xnm3930",	"r_xnm3931",	"r_xnm3932",	"r_xnm3933",	"r_xnm3934",	"r_xnm3935",	"r_xnm3936",	"r_xnm3937",	"r_xnm3938",	"r_xnm3939",	"r_xnm3940",	"r_xnm3941",	"r_xnm3942",	"r_xnm3943",	"r_xnm3944",	"r_xnm3945",	"r_xnm3946",	"r_xnm3947",	"r_xnm3948",	"r_xnm3949",	"r_xnm3950",	"r_xnm3951",	"r_xnm3952",	"r_xnm3953",	"r_xnm3954",	"r_xnm3955",	"r_xnm3956",	"r_xnm3957",	"r_xnm3958",	"r_xnm3959",	"r_xnm3960",	"r_xnm3961",	"r_xnm3962",	"r_xnm3963",	"r_xnm3964",	"r_xnm3965",	"r_xnm3966",	"r_xnm3967",	"r_xnm3968",	"r_xnm3969",	"r_xnm3970",	"r_xnm3971",	"r_xnm3972",	"r_xnm3973",	"r_xnm3974",	"r_xnm3975",	"r_xnm3976",	"r_xnm3977",	"r_xnm3978",	"r_xnm3979",	"r_xnm3980",	"r_xnm3981",	"r_xnm3982",	"r_xnm3983",	"r_xnm3984",	"r_xnm3985",	"r_xnm3986",	"r_xnm3987",	"r_xnm3988",	"r_xnm3989",	"r_xnm3990",	"r_xnm3991",	"r_xnm3992",	"r_xnm3993",	"r_xnm3994",	"r_xnm3995",	"r_xnm3996",	"r_xnm3997",	"r_xnm3998",	"r_xnm3999",	"r_xnm4000",	"r_xnm4001",	"r_xnm4002",	"r_xnm4003",	"r_xnm4004",	"r_xnm4005",	"r_xnm4006",	"r_xnm4007",	"r_xnm4008",	"r_xnm4009",	"r_xnm4010",	"r_xnm4011",	"r_xnm4012",	"r_xnm4013",	"r_xnm4014",	"r_xnm4015",	"r_xnm4016",	"r_xnm4017",	"r_xnm4018",	"r_xnm4019",	"r_xnm4020",	"r_xnm4021",	"r_xnm4022",	"r_xnm4023",	"r_xnm4024",	"r_xnm4025",	"r_xnm4026",	"r_xnm4027",	"r_xnm4028",	"r_xnm4029",	"r_xnm4030",	"r_xnm4031",	"r_xnm4032",	"r_xnm4033",	"r_xnm4034",	"r_xnm4035",	"r_xnm4036",	"r_xnm4037",	"r_xnm4038",	"r_xnm4039",	"r_xnm4040",	"r_xnm4041",	"r_xnm4042",	"r_xnm4043",	"r_xnm4044",	"r_xnm4045",	"r_xnm4046",	"r_xnm4047",	"r_xnm4048",	"r_xnm4049",	"r_xnm4050",	"r_xnm4051",	"r_xnm4052",	"r_xnm4053",	"r_xnm4054",	"r_xnm4055",	"r_xnm4056",	"r_xnm4057",	"r_xnm4058",	"r_xnm4059",	"r_xnm4060",	"r_xnm4061",	"r_xnm4062",	"r_xnm4063",	"r_xnm4064",	"r_xnm4065",	"r_xnm4066",	"r_xnm4067",	"r_xnm4068",	"r_xnm4069",	"r_xnm4070",	"r_xnm4071",	"r_xnm4072",	"r_xnm4073",	"r_xnm4074",	"r_xnm4075",	"r_xnm4076",	"r_xnm4077",	"r_xnm4078",	"r_xnm4079",	"r_xnm4080",	"r_xnm4081",	"r_xnm4082",	"r_xnm4083",	"r_xnm4084",	"r_xnm4085",	"r_xnm4086",	"r_xnm4087",	"r_xnm4088",	"r_xnm4089",	"r_xnm4090",	"r_xnm4091",	"r_xnm4092",	"r_xnm4093",	"r_xnm4094",	"r_xnm4095",	"r_xnm4096",	"r_xnm4097",	"r_xnm4098",	"r_xnm4099",	"r_xnm4100",	"r_xnm4101",	"r_xnm4102",	"r_xnm4103",	"r_xnm4104",	"r_xnm4105",	"r_xnm4106",	"r_xnm4107",	"r_xnm4108",	"r_xnm4109",	"r_xnm4110",	"r_xnm4111",	"r_xnm4112",	"r_xnm4113",	"r_xnm4114",	"r_xnm4115",	"r_xnm4116",	"r_xnm4117",	"r_xnm4118",	"r_xnm4119",	"r_xnm4120",	"r_xnm4121",	"r_xnm4122",	"r_xnm4123",	"r_xnm4124",	"r_xnm4125",	"r_xnm4126",	"r_xnm4127",	"r_xnm4128",	"r_xnm4129",	"r_xnm4130",	"r_xnm4131",	"r_xnm4132",	"r_xnm4133",	"r_xnm4134",	"r_xnm4135",	"r_xnm4136",	"r_xnm4137",	"r_xnm4138",	"r_xnm4139",	"r_xnm4140",	"r_xnm4141",	"r_xnm4142",	"r_xnm4143",	"r_xnm4144",	"r_xnm4145",	"r_xnm4146",	"r_xnm4147",	"r_xnm4148",	"r_xnm4149",	"r_xnm4150",	"r_xnm4151",	"r_xnm4152",	"r_xnm4153",	"r_xnm4154",	"r_xnm4155",	"r_xnm4156",	"r_xnm4157",	"r_xnm4158",	"r_xnm4159",	"r_xnm4160",	"r_xnm4161",	"r_xnm4162",	"r_xnm4163",	"r_xnm4164",	"r_xnm4165",	"r_xnm4166",	"r_xnm4167",	"r_xnm4168",	"r_xnm4169",	"r_xnm4170",	"r_xnm4171",	"r_xnm4172",	"r_xnm4173",	"r_xnm4174",	"r_xnm4175",	"r_xnm4176",	"r_xnm4177",	"r_xnm4178",	"r_xnm4179",	"r_xnm4180",	"r_xnm4181",	"r_xnm4182",	"r_xnm4183",	"r_xnm4184",	"r_xnm4185",	"r_xnm4186",	"r_xnm4187",	"r_xnm4188",	"r_xnm4189",	"r_xnm4190",	"r_xnm4191",	"r_xnm4192",	"r_xnm4193",	"r_xnm4194",	"r_xnm4195",	"r_xnm4196",	"r_xnm4197",	"r_xnm4198",	"r_xnm4199",	"r_xnm4200",	"r_xnm4201",	"r_xnm4202",	"r_xnm4203",	"r_xnm4204",	"r_xnm4205",	"r_xnm4206",	"r_xnm4207",	"r_xnm4208",	"r_xnm4209",	"r_xnm4210",	"r_xnm4211",	"r_xnm4212",	"r_xnm4213",	"r_xnm4214",	"r_xnm4215",	"r_xnm4216",	"r_xnm4217",	"r_xnm4218",	"r_xnm4219",	"r_xnm4220",	"r_xnm4221",	"r_xnm4222",	"r_xnm4223",	"r_xnm4224",	"r_xnm4225",	"r_xnm4226",	"r_xnm4227",	"r_xnm4228",	"r_xnm4229",	"r_xnm4230",	"r_xnm4231",	"r_xnm4232",	"r_xnm4233",	"r_xnm4234",	"r_xnm4235",	"r_xnm4236",	"r_xnm4237",	"r_xnm4238",	"r_xnm4239",	"r_xnm4240",	"r_xnm4241",	"r_xnm4242",	"r_xnm4243",	"r_xnm4244",	"r_xnm4245",	"r_xnm4246",	"r_xnm4247",	"r_xnm4248",	"r_xnm4249",	"r_xnm4250",	"r_xnm4251",	"r_xnm4252",	"r_xnm4253",	"r_xnm4254",	"r_xnm4255",	"r_xnm4256",	"r_xnm4257",	"r_xnm4258",	"r_xnm4259",	"r_xnm4260",	"r_xnm4261",	"r_xnm4262",	"r_xnm4263",	"r_xnm4264",	"r_xnm4265",	"r_xnm4266",	"r_xnm4267",	"r_xnm4268",	"r_xnm4269",	"r_xnm4270",	"r_xnm4271",	"r_xnm4272",	"r_xnm4273",	"r_xnm4274",	"r_xnm4275",	"r_xnm4276",	"r_xnm4277",	"r_xnm4278",	"r_xnm4279",	"r_xnm4280",	"r_xnm4281",	"r_xnm4282",	"r_xnm4283",	"r_xnm4284",	"r_xnm4285",	"r_xnm4286",	"r_xnm4287",	"r_xnm4288",	"r_xnm4289",	"r_xnm4290",	"r_xnm4291",	"r_xnm4292",	"r_xnm4293",	"r_xnm4294",	"r_xnm4295",	"r_xnm4296",	"r_xnm4297",	"r_xnm4298",	"r_xnm4299",	"r_xnm4300",	"r_xnm4301",	"r_xnm4302",	"r_xnm4303",	"r_xnm4304",	"r_xnm4305",	"r_xnm4306",	"r_xnm4307",	"r_xnm4308",	"r_xnm4309",	"r_xnm4310",	"r_xnm4311",	"r_xnm4312",	"r_xnm4313",	"r_xnm4314",	"r_xnm4315",	"r_xnm4316",	"r_xnm4317",	"r_xnm4318",	"r_xnm4319",	"r_xnm4320",	"r_xnm4321",	"r_xnm4322",	"r_xnm4323",	"r_xnm4324",	"r_xnm4325",	"r_xnm4326",	"r_xnm4327",	"r_xnm4328",	"r_xnm4329",	"r_xnm4330",	"r_xnm4331",	"r_xnm4332",	"r_xnm4333",	"r_xnm4334",	"r_xnm4335",	"r_xnm4336",	"r_xnm4337",	"r_xnm4338",	"r_xnm4339",	"r_xnm4340",	"r_xnm4341",	"r_xnm4342",	"r_xnm4343",	"r_xnm4344",	"r_xnm4345",	"r_xnm4346",	"r_xnm4347",	"r_xnm4348",	"r_xnm4349",	"r_xnm4350",	"r_xnm4351",	"r_xnm4352",	"r_xnm4353",	"r_xnm4354",	"r_xnm4355",	"r_xnm4356",	"r_xnm4357",	"r_xnm4358",	"r_xnm4359",	"r_xnm4360",	"r_xnm4361",	"r_xnm4362",	"r_xnm4363",	"r_xnm4364",	"r_xnm4365",	"r_xnm4366",	"r_xnm4367",	"r_xnm4368",	"r_xnm4369",	"r_xnm4370",	"r_xnm4371",	"r_xnm4372",	"r_xnm4373",	"r_xnm4374",	"r_xnm4375",	"r_xnm4376",	"r_xnm4377",	"r_xnm4378",	"r_xnm4379",	"r_xnm4380",	"r_xnm4381",	"r_xnm4382",	"r_xnm4383",	"r_xnm4384",	"r_xnm4385",	"r_xnm4386",	"r_xnm4387",	"r_xnm4388",	"r_xnm4389",	"r_xnm4390",	"r_xnm4391",	"r_xnm4392",	"r_xnm4393",	"r_xnm4394",	"r_xnm4395",	"r_xnm4396",	"r_xnm4397",	"r_xnm4398",	"r_xnm4399",	"r_xnm4400",	"r_xnm4401",	"r_xnm4402",	"r_xnm4403",	"r_xnm4404",	"r_xnm4405",	"r_xnm4406",	"r_xnm4407",	"r_xnm4408",	"r_xnm4409",	"r_xnm4410",	"r_xnm4411",	"r_xnm4412",	"r_xnm4413",	"r_xnm4414",	"r_xnm4415",	"r_xnm4416",	"r_xnm4417",	"r_xnm4418",	"r_xnm4419",	"r_xnm4420",	"r_xnm4421",	"r_xnm4422",	"r_xnm4423",	"r_xnm4424",	"r_xnm4425",	"r_xnm4426",	"r_xnm4427",	"r_xnm4428",	"r_xnm4429",	"r_xnm4430",	"r_xnm4431",	"r_xnm4432",	"r_xnm4433",	"r_xnm4434",	"r_xnm4435",	"r_xnm4436",	"r_xnm4437",	"r_xnm4438",	"r_xnm4439",	"r_xnm4440",	"r_xnm4441",	"r_xnm4442",	"r_xnm4443",	"r_xnm4444",	"r_xnm4445",	"r_xnm4446",	"r_xnm4447",	"r_xnm4448",	"r_xnm4449",	"r_xnm4450",	"r_xnm4451",	"r_xnm4452",	"r_xnm4453",	"r_xnm4454",	"r_xnm4455",	"r_xnm4456",	"r_xnm4457",	"r_xnm4458",	"r_xnm4459",	"r_xnm4460",	"r_xnm4461",	"r_xnm4462",	"r_xnm4463",	"r_xnm4464",	"r_xnm4465",	"r_xnm4466",	"r_xnm4467",	"r_xnm4468",	"r_xnm4469",	"r_xnm4470",	"r_xnm4471",	"r_xnm4472",	"r_xnm4473",	"r_xnm4474",	"r_xnm4475",	"r_xnm4476",	"r_xnm4477",	"r_xnm4478",	"r_xnm4479",	"r_xnm4480",	"r_xnm4481",	"r_xnm4482",	"r_xnm4483",	"r_xnm4484",	"r_xnm4485",	"r_xnm4486",	"r_xnm4487",	"r_xnm4488",	"r_xnm4489",	"r_xnm4490",	"r_xnm4491",	"r_xnm4492",	"r_xnm4493",	"r_xnm4494",	"r_xnm4495",	"r_xnm4496",	"r_xnm4497",	"r_xnm4498",	"r_xnm4499",	"r_xnm4500",	"r_xnm4501",	"r_xnm4502",	"r_xnm4503",	"r_xnm4504",	"r_xnm4505",	"r_xnm4506",	"r_xnm4507",	"r_xnm4508",	"r_xnm4509",	"r_xnm4510",	"r_xnm4511",	"r_xnm4512",	"r_xnm4513",	"r_xnm4514",	"r_xnm4515",	"r_xnm4516",	"r_xnm4517",	"r_xnm4518",	"r_xnm4519",	"r_xnm4520",	"r_xnm4521",	"r_xnm4522",	"r_xnm4523",	"r_xnm4524",	"r_xnm4525",	"r_xnm4526",	"r_xnm4527",	"r_xnm4528",	"r_xnm4529",	"r_xnm4530",	"r_xnm4531",	"r_xnm4532",	"r_xnm4533",	"r_xnm4534",	"r_xnm4535",	"r_xnm4536",	"r_xnm4537",	"r_xnm4538",	"r_xnm4539",	"r_xnm4540",	"r_xnm4541",	"r_xnm4542",	"r_xnm4543",	"r_xnm4544",	"r_xnm4545",	"r_xnm4546",	"r_xnm4547",	"r_xnm4548",	"r_xnm4549",	"r_xnm4550",	"r_xnm4551",	"r_xnm4552",	"r_xnm4553",	"r_xnm4554",	"r_xnm4555",	"r_xnm4556",	"r_xnm4557",	"r_xnm4558",	"r_xnm4559",	"r_xnm4560",	"r_xnm4561",	"r_xnm4562",	"r_xnm4563",	"r_xnm4564",	"r_xnm4565",	"r_xnm4566",	"r_xnm4567",	"r_xnm4568",	"r_xnm4569",	"r_xnm4570",	"r_xnm4571",	"r_xnm4572",	"r_xnm4573",	"r_xnm4574",	"r_xnm4575",	"r_xnm4576",	"r_xnm4577",	"r_xnm4578",	"r_xnm4579",	"r_xnm4580",	"r_xnm4581",	"r_xnm4582",	"r_xnm4583",	"r_xnm4584",	"r_xnm4585",	"r_xnm4586",	"r_xnm4587",	"r_xnm4588",	"r_xnm4589",	"r_xnm4590",	"r_xnm4591",	"r_xnm4592",	"r_xnm4593",	"r_xnm4594",	"r_xnm4595",	"r_xnm4596",	"r_xnm4597",	"r_xnm4598",	"r_xnm4599",	"r_xnm4600",	"r_xnm4601",	"r_xnm4602",	"r_xnm4603",	"r_xnm4604",	"r_xnm4605",	"r_xnm4606",	"r_xnm4607",	"r_xnm4608",	"r_xnm4609",	"r_xnm4610",	"r_xnm4611",	"r_xnm4612",	"r_xnm4613",	"r_xnm4614",	"r_xnm4615",	"r_xnm4616",	"r_xnm4617",	"r_xnm4618",	"r_xnm4619",	"r_xnm4620",	"r_xnm4621",	"r_xnm4622",	"r_xnm4623",	"r_xnm4624",	"r_xnm4625",	"r_xnm4626",	"r_xnm4627",	"r_xnm4628",	"r_xnm4629",	"r_xnm4630",	"r_xnm4631",	"r_xnm4632",	"r_xnm4633",	"r_xnm4634",	"r_xnm4635",	"r_xnm4636",	"r_xnm4637",	"r_xnm4638",	"r_xnm4639",	"r_xnm4640",	"r_xnm4641",	"r_xnm4642",	"r_xnm4643",	"r_xnm4644",	"r_xnm4645",	"r_xnm4646",	"r_xnm4647",	"r_xnm4648",	"r_xnm4649",	"r_xnm4650",	"r_xnm4651",	"r_xnm4652",	"r_xnm4653",	"r_xnm4654",	"r_xnm4655",	"r_xnm4656",	"r_xnm4657",	"r_xnm4658",	"r_xnm4659",	"r_xnm4660",	"r_xnm4661",	"r_xnm4662",	"r_xnm4663",	"r_xnm4664",	"r_xnm4665",	"r_xnm4666",	"r_xnm4667",	"r_xnm4668",	"r_xnm4669",	"r_xnm4670",	"r_xnm4671",	"r_xnm4672",	"r_xnm4673",	"r_xnm4674",	"r_xnm4675",	"r_xnm4676",	"r_xnm4677",	"r_xnm4678",	"r_xnm4679",	"r_xnm4680",	"r_xnm4681",	"r_xnm4682",	"r_xnm4683",	"r_xnm4684",	"r_xnm4685",	"r_xnm4686",	"r_xnm4687",	"r_xnm4688",	"r_xnm4689",	"r_xnm4690",	"r_xnm4691",	"r_xnm4692",	"r_xnm4693",	"r_xnm4694",	"r_xnm4695",	"r_xnm4696",	"r_xnm4697",	"r_xnm4698",	"r_xnm4699",	"r_xnm4700",	"r_xnm4701",	"r_xnm4702",	"r_xnm4703",	"r_xnm4704",	"r_xnm4705",	"r_xnm4706",	"r_xnm4707",	"r_xnm4708",	"r_xnm4709",	"r_xnm4710",	"r_xnm4711",	"r_xnm4712",	"r_xnm4713",	"r_xnm4714",	"r_xnm4715",	"r_xnm4716",	"r_xnm4717",	"r_xnm4718",	"r_xnm4719",	"r_xnm4720",	"r_xnm4721",	"r_xnm4722",	"r_xnm4723",	"r_xnm4724",	"r_xnm4725",	"r_xnm4726",	"r_xnm4727",	"r_xnm4728",	"r_xnm4729",	"r_xnm4730",	"r_xnm4731",	"r_xnm4732",	"r_xnm4733",	"r_xnm4734",	"r_xnm4735",	"r_xnm4736",	"r_xnm4737",	"r_xnm4738",	"r_xnm4739",	"r_xnm4740",	"r_xnm4741",	"r_xnm4742",	"r_xnm4743",	"r_xnm4744",	"r_xnm4745",	"r_xnm4746",	"r_xnm4747",	"r_xnm4748",	"r_xnm4749",	"r_xnm4750",	"r_xnm4751",	"r_xnm4752",	"r_xnm4753",	"r_xnm4754",	"r_xnm4755",	"r_xnm4756",	"r_xnm4757",	"r_xnm4758",	"r_xnm4759",	"r_xnm4760",	"r_xnm4761",	"r_xnm4762",	"r_xnm4763",	"r_xnm4764",	"r_xnm4765",	"r_xnm4766",	"r_xnm4767",	"r_xnm4768",	"r_xnm4769",	"r_xnm4770",	"r_xnm4771",	"r_xnm4772",	"r_xnm4773",	"r_xnm4774",	"r_xnm4775",	"r_xnm4776",	"r_xnm4777",	"r_xnm4778",	"r_xnm4779",	"r_xnm4780",	"r_xnm4781",	"r_xnm4782",	"r_xnm4783",	"r_xnm4784",	"r_xnm4785",	"r_xnm4786",	"r_xnm4787",	"r_xnm4788",	"r_xnm4789",	"r_xnm4790",	"r_xnm4791",	"r_xnm4792",	"r_xnm4793",	"r_xnm4794",	"r_xnm4795",	"r_xnm4796",	"r_xnm4797",	"r_xnm4798",	"r_xnm4799",	"r_xnm4800",	"r_xnm4801",	"r_xnm4802",	"r_xnm4803",	"r_xnm4804",	"r_xnm4805",	"r_xnm4806",	"r_xnm4807",	"r_xnm4808",	"r_xnm4809",	"r_xnm4810",	"r_xnm4811",	"r_xnm4812",	"r_xnm4813",	"r_xnm4814",	"r_xnm4815",	"r_xnm4816",	"r_xnm4817",	"r_xnm4818",	"r_xnm4819",	"r_xnm4820",	"r_xnm4821",	"r_xnm4822",	"r_xnm4823",	"r_xnm4824",	"r_xnm4825",	"r_xnm4826",	"r_xnm4827",	"r_xnm4828",	"r_xnm4829",	"r_xnm4830",	"r_xnm4831",	"r_xnm4832",	"r_xnm4833",	"r_xnm4834",	"r_xnm4835",	"r_xnm4836",	"r_xnm4837",	"r_xnm4838",	"r_xnm4839",	"r_xnm4840",	"r_xnm4841",	"r_xnm4842",	"r_xnm4843",	"r_xnm4844",	"r_xnm4845",	"r_xnm4846",	"r_xnm4847",	"r_xnm4848",	"r_xnm4849",	"r_xnm4850",	"r_xnm4851",	"r_xnm4852",	"r_xnm4853",	"r_xnm4854",	"r_xnm4855",	"r_xnm4856",	"r_xnm4857",	"r_xnm4858",	"r_xnm4859",	"r_xnm4860",	"r_xnm4861",	"r_xnm4862",	"r_xnm4863",	"r_xnm4864",	"r_xnm4865",	"r_xnm4866",	"r_xnm4867",	"r_xnm4868",	"r_xnm4869",	"r_xnm4870",	"r_xnm4871",	"r_xnm4872",	"r_xnm4873",	"r_xnm4874",	"r_xnm4875",	"r_xnm4876",	"r_xnm4877",	"r_xnm4878",	"r_xnm4879",	"r_xnm4880",	"r_xnm4881",	"r_xnm4882",	"r_xnm4883",	"r_xnm4884",	"r_xnm4885",	"r_xnm4886",	"r_xnm4887",	"r_xnm4888",	"r_xnm4889",	"r_xnm4890",	"r_xnm4891",	"r_xnm4892",	"r_xnm4893",	"r_xnm4894",	"r_xnm4895",	"r_xnm4896",	"r_xnm4897",	"r_xnm4898",	"r_xnm4899",	"r_xnm4900",	"r_xnm4901",	"r_xnm4902",	"r_xnm4903",	"r_xnm4904",	"r_xnm4905",	"r_xnm4906",	"r_xnm4907",	"r_xnm4908",	"r_xnm4909",	"r_xnm4910",	"r_xnm4911",	"r_xnm4912",	"r_xnm4913",	"r_xnm4914",	"r_xnm4915",	"r_xnm4916",	"r_xnm4917",	"r_xnm4918",	"r_xnm4919",	"r_xnm4920",	"r_xnm4921",	"r_xnm4922",	"r_xnm4923",	"r_xnm4924",	"r_xnm4925",	"r_xnm4926",	"r_xnm4927",	"r_xnm4928",	"r_xnm4929",	"r_xnm4930",	"r_xnm4931",	"r_xnm4932",	"r_xnm4933",	"r_xnm4934",	"r_xnm4935",	"r_xnm4936",	"r_xnm4937",	"r_xnm4938",	"r_xnm4939",	"r_xnm4940",	"r_xnm4941",	"r_xnm4942",	"r_xnm4943",	"r_xnm4944",	"r_xnm4945",	"r_xnm4946",	"r_xnm4947",	"r_xnm4948",	"r_xnm4949",	"r_xnm4950",	"r_xnm4951",	"r_xnm4952",	"r_xnm4953",	"r_xnm4954",	"r_xnm4955",	"r_xnm4956",	"r_xnm4957",	"r_xnm4958",	"r_xnm4959",	"r_xnm4960",	"r_xnm4961",	"r_xnm4962",	"r_xnm4963",	"r_xnm4964",	"r_xnm4965",	"r_xnm4966",	"r_xnm4967",	"r_xnm4968",	"r_xnm4969",	"r_xnm4970",	"r_xnm4971",	"r_xnm4972",	"r_xnm4973",	"r_xnm4974",	"r_xnm4975",	"r_xnm4976",	"r_xnm4977",	"r_xnm4978",	"r_xnm4979",	"r_xnm4980",	"r_xnm4981",	"r_xnm4982",	"r_xnm4983",	"r_xnm4984",	"r_xnm4985",	"r_xnm4986",	"r_xnm4987",	"r_xnm4988",	"r_xnm4989",	"r_xnm4990",	"r_xnm4991",	"r_xnm4992",	"r_xnm4993",	"r_xnm4994",	"r_xnm4995",	"r_xnm4996",	"r_xnm4997",	"r_xnm4998",	"r_xnm4999",	"r_xnm5000",	"r_xnm5001",	"r_xnm5002",	"r_xnm5003",	"r_xnm5004",	"r_xnm5005",	"r_xnm5006",	"r_xnm5007",	"r_xnm5008",	"r_xnm5009",	"r_xnm5010",	"r_xnm5011",	"r_xnm5012",	"r_xnm5013",	"r_xnm5014",	"r_xnm5015",	"r_xnm5016",	"r_xnm5017",	"r_xnm5018",	"r_xnm5019",	"r_xnm5020",	"r_xnm5021",	"r_xnm5022",	"r_xnm5023",	"r_xnm5024",	"r_xnm5025",	"r_xnm5026",	"r_xnm5027",	"r_xnm5028",	"r_xnm5029",	"r_xnm5030",	"r_xnm5031",	"r_xnm5032",	"r_xnm5033",	"r_xnm5034",	"r_xnm5035",	"r_xnm5036",	"r_xnm5037",	"r_xnm5038",	"r_xnm5039",	"r_xnm5040",	"r_xnm5041",	"r_xnm5042",	"r_xnm5043",	"r_xnm5044",	"r_xnm5045",	"r_xnm5046",	"r_xnm5047",	"r_xnm5048",	"r_xnm5049",	"r_xnm5050",	"r_xnm5051",	"r_xnm5052",	"r_xnm5053",	"r_xnm5054",	"r_xnm5055",	"r_xnm5056",	"r_xnm5057",	"r_xnm5058",	"r_xnm5059",	"r_xnm5060",	"r_xnm5061",	"r_xnm5062",	"r_xnm5063",	"r_xnm5064",	"r_xnm5065",	"r_xnm5066",	"r_xnm5067",	"r_xnm5068",	"r_xnm5069",	"r_xnm5070",	"r_xnm5071",	"r_xnm5072",	"r_xnm5073",	"r_xnm5074",	"r_xnm5075",	"r_xnm5076",	"r_xnm5077",	"r_xnm5078",	"r_xnm5079",	"r_xnm5080",	"r_xnm5081",	"r_xnm5082",	"r_xnm5083",	"r_xnm5084",	"r_xnm5085",	"r_xnm5086",	"r_xnm5087",	"r_xnm5088",	"r_xnm5089",	"r_xnm5090",	"r_xnm5091",	"r_xnm5092",	"r_xnm5093",	"r_xnm5094",	"r_xnm5095",	"r_xnm5096",	"r_xnm5097",	"r_xnm5098",	"r_xnm5099",	"r_xnm5100",	"r_xnm5101",	"r_xnm5102",	"r_xnm5103",	"r_xnm5104",	"r_xnm5105",	"r_xnm5106",	"r_xnm5107",	"r_xnm5108",	"r_xnm5109",	"r_xnm5110",	"r_xnm5111",	"r_xnm5112",	"r_xnm5113",	"r_xnm5114",	"r_xnm5115",	"r_xnm5116",	"r_xnm5117",	"r_xnm5118",	"r_xnm5119",	"r_xnm5120",	"r_xnm5121",	"r_xnm5122",	"r_xnm5123",	"r_xnm5124",	"r_xnm5125",	"r_xnm5126",	"r_xnm5127",	"r_xnm5128",	"r_xnm5129",	"r_xnm5130",	"r_xnm5131",	"r_xnm5132",	"r_xnm5133",	"r_xnm5134",	"r_xnm5135",	"r_xnm5136",	"r_xnm5137",	"r_xnm5138",	"r_xnm5139",	"r_xnm5140",	"r_xnm5141",	"r_xnm5142",	"r_xnm5143",	"r_xnm5144",	"r_xnm5145",	"r_xnm5146",	"r_xnm5147",	"r_xnm5148",	"r_xnm5149",	"r_xnm5150",	"r_xnm5151",	"r_xnm5152",	"r_xnm5153",	"r_xnm5154",	"r_xnm5155",	"r_xnm5156",	"r_xnm5157",	"r_xnm5158",	"r_xnm5159",	"r_xnm5160",	"r_xnm5161",	"r_xnm5162",	"r_xnm5163",	"r_xnm5164",	"r_xnm5165",	"r_xnm5166",	"r_xnm5167",	"r_xnm5168",	"r_xnm5169",	"r_xnm5170",	"r_xnm5171",	"r_xnm5172",	"r_xnm5173",	"r_xnm5174",	"r_xnm5175",	"r_xnm5176",	"r_xnm5177",	"r_xnm5178",	"r_xnm5179",	"r_xnm5180",	"r_xnm5181",	"r_xnm5182",	"r_xnm5183",	"r_xnm5184",	"r_xnm5185",	"r_xnm5186",	"r_xnm5187",	"r_xnm5188",	"r_xnm5189",	"r_xnm5190",	"r_xnm5191",	"r_xnm5192",	"r_xnm5193",	"r_xnm5194",	"r_xnm5195",	"r_xnm5196",	"r_xnm5197",	"r_xnm5198",	"r_xnm5199",	"r_xnm5200",	"r_xnm5201",	"r_xnm5202",	"r_xnm5203",	"r_xnm5204",	"r_xnm5205",	"r_xnm5206",	"r_xnm5207",	"r_xnm5208",	"r_xnm5209",	"r_xnm5210",	"r_xnm5211",	"r_xnm5212",	"r_xnm5213",	"r_xnm5214",	"r_xnm5215",	"r_xnm5216",	"r_xnm5217",	"r_xnm5218",	"r_xnm5219",	"r_xnm5220",	"r_xnm5221",	"r_xnm5222",	"r_xnm5223",	"r_xnm5224",	"r_xnm5225",	"r_xnm5226",	"r_xnm5227",	"r_xnm5228",	"r_xnm5229",	"r_xnm5230",	"r_xnm5231",	"r_xnm5232",	"r_xnm5233",	"r_xnm5234",	"r_xnm5235",	"r_xnm5236",	"r_xnm5237",	"r_xnm5238",	"r_xnm5239",	"r_xnm5240",	"r_xnm5241",	"r_xnm5242",	"r_xnm5243",	"r_xnm5244",	"r_xnm5245",	"r_xnm5246",	"r_xnm5247",	"r_xnm5248",	"r_xnm5249",	"r_xnm5250",	"r_xnm5251",	"r_xnm5252",	"r_xnm5253",	"r_xnm5254",	"r_xnm5255",	"r_xnm5256",	"r_xnm5257",	"r_xnm5258",	"r_xnm5259",	"r_xnm5260",	"r_xnm5261",	"r_xnm5262",	"r_xnm5263",	"r_xnm5264",	"r_xnm5265",	"r_xnm5266",	"r_xnm5267",	"r_xnm5268",	"r_xnm5269",	"r_xnm5270",	"r_xnm5271",	"r_xnm5272",	"r_xnm5273",	"r_xnm5274",	"r_xnm5275",	"r_xnm5276",	"r_xnm5277",	"r_xnm5278",	"r_xnm5279",	"r_xnm5280",	"r_xnm5281",	"r_xnm5282",	"r_xnm5283",	"r_xnm5284",	"r_xnm5285",	"r_xnm5286",	"r_xnm5287",	"r_xnm5288",	"r_xnm5289",	"r_xnm5290",	"r_xnm5291",	"r_xnm5292",	"r_xnm5293",	"r_xnm5294",	"r_xnm5295",	"r_xnm5296",	"r_xnm5297",	"r_xnm5298",	"r_xnm5299",	"r_xnm5300",	"r_xnm5301",	"r_xnm5302",	"r_xnm5303",	"r_xnm5304",	"r_xnm5305",	"r_xnm5306",	"r_xnm5307",	"r_xnm5308",	"r_xnm5309",	"r_xnm5310",	"r_xnm5311",	"r_xnm5312",	"r_xnm5313",	"r_xnm5314",	"r_xnm5315",	"r_xnm5316",	"r_xnm5317",	"r_xnm5318",	"r_xnm5319",	"r_xnm5320",	"r_xnm5321",	"r_xnm5322",	"r_xnm5323",	"r_xnm5324",	"r_xnm5325",	"r_xnm5326",	"r_xnm5327",	"r_xnm5328",	"r_xnm5329",	"r_xnm5330",	"r_xnm5331",	"r_xnm5332",	"r_xnm5333",	"r_xnm5334",	"r_xnm5335",	"r_xnm5336",	"r_xnm5337",	"r_xnm5338",	"r_xnm5339",	"r_xnm5340",	"r_xnm5341",	"r_xnm5342",	"r_xnm5343",	"r_xnm5344",	"r_xnm5345",	"r_xnm5346",	"r_xnm5347",	"r_xnm5348",	"r_xnm5349",	"r_xnm5350",	"r_xnm5351",	"r_xnm5352",	"r_xnm5353",	"r_xnm5354",	"r_xnm5355",	"r_xnm5356",	"r_xnm5357",	"r_xnm5358",	"r_xnm5359",	"r_xnm5360",	"r_xnm5361",	"r_xnm5362",	"r_xnm5363",	"r_xnm5364",	"r_xnm5365",	"r_xnm5366",	"r_xnm5367",	"r_xnm5368",	"r_xnm5369",	"r_xnm5370",	"r_xnm5371",	"r_xnm5372",	"r_xnm5373",	"r_xnm5374",	"r_xnm5375",	"r_xnm5376",	"r_xnm5377",	"r_xnm5378",	"r_xnm5379",	"r_xnm5380",	"r_xnm5381",	"r_xnm5382",	"r_xnm5383",	"r_xnm5384",	"r_xnm5385",	"r_xnm5386",	"r_xnm5387",	"r_xnm5388",	"r_xnm5389",	"r_xnm5390",	"r_xnm5391",	"r_xnm5392",	"r_xnm5393",	"r_xnm5394",	"r_xnm5395",	"r_xnm5396",	"r_xnm5397",	"r_xnm5398",	"r_xnm5399",	"r_xnm5400",	"r_xnm5401",	"r_xnm5402",	"r_xnm5403",	"r_xnm5404",	"r_xnm5405",	"r_xnm5406",	"r_xnm5407",	"r_xnm5408",	"r_xnm5409",	"r_xnm5410",	"r_xnm5411",	"r_xnm5412",	"r_xnm5413",	"r_xnm5414",	"r_xnm5415",	"r_xnm5416",	"r_xnm5417",	"r_xnm5418",	"r_xnm5419",	"r_xnm5420",	"r_xnm5421",	"r_xnm5422",	"r_xnm5423",	"r_xnm5424",	"r_xnm5425",	"r_xnm5426",	"r_xnm5427",	"r_xnm5428",	"r_xnm5429",	"r_xnm5430",	"r_xnm5431",	"r_xnm5432",	"r_xnm5433",	"r_xnm5434",	"r_xnm5435",	"r_xnm5436",	"r_xnm5437",	"r_xnm5438",	"r_xnm5439",	"r_xnm5440",	"r_xnm5441",	"r_xnm5442",	"r_xnm5443",	"r_xnm5444",	"r_xnm5445",	"r_xnm5446",	"r_xnm5447",	"r_xnm5448",	"r_xnm5449",	"r_xnm5450",	"r_xnm5451",	"r_xnm5452",	"r_xnm5453",	"r_xnm5454",	"r_xnm5455",	"r_xnm5456",	"r_xnm5457",	"r_xnm5458",	"r_xnm5459",	"r_xnm5460",	"r_xnm5461",	"r_xnm5462",	"r_xnm5463",	"r_xnm5464",	"r_xnm5465",	"r_xnm5466",	"r_xnm5467",	"r_xnm5468",	"r_xnm5469",	"r_xnm5470",	"r_xnm5471",	"r_xnm5472",	"r_xnm5473",	"r_xnm5474",	"r_xnm5475",	"r_xnm5476",	"r_xnm5477",	"r_xnm5478",	"r_xnm5479",	"r_xnm5480",	"r_xnm5481",	"r_xnm5482",	"r_xnm5483",	"r_xnm5484",	"r_xnm5485",	"r_xnm5486",	"r_xnm5487",	"r_xnm5488",	"r_xnm5489",	"r_xnm5490",	"r_xnm5491",	"r_xnm5492",	"r_xnm5493",	"r_xnm5494",	"r_xnm5495",	"r_xnm5496",	"r_xnm5497",	"r_xnm5498",	"r_xnm5499",	"r_xnm5500",	"r_xnm5501",	"r_xnm5502",	"r_xnm5503",	"r_xnm5504",	"r_xnm5505",	"r_xnm5506",	"r_xnm5507",	"r_xnm5508",	"r_xnm5509",	"r_xnm5510",	"r_xnm5511",	"r_xnm5512",	"r_xnm5513",	"r_xnm5514",	"r_xnm5515",	"r_xnm5516",	"r_xnm5517",	"r_xnm5518",	"r_xnm5519",	"r_xnm5520",	"r_xnm5521",	"r_xnm5522",	"r_xnm5523",	"r_xnm5524",	"r_xnm5525",	"r_xnm5526",	"r_xnm5527",	"r_xnm5528",	"r_xnm5529",	"r_xnm5530",	"r_xnm5531",	"r_xnm5532",	"r_xnm5533",	"r_xnm5534",	"r_xnm5535",	"r_xnm5536",	"r_xnm5537",	"r_xnm5538",	"r_xnm5539",	"r_xnm5540",	"r_xnm5541",	"r_xnm5542",	"r_xnm5543",	"r_xnm5544",	"r_xnm5545",	"r_xnm5546",	"r_xnm5547",	"r_xnm5548",	"r_xnm5549",	"r_xnm5550",	"r_xnm5551",	"r_xnm5552",	"r_xnm5553",	"r_xnm5554",	"r_xnm5555",	"r_xnm5556",	"r_xnm5557",	"r_xnm5558",	"r_xnm5559",	"r_xnm5560",	"r_xnm5561",	"r_xnm5562",	"r_xnm5563",	"r_xnm5564",	"r_xnm5565",	"r_xnm5566",	"r_xnm5567",	"r_xnm5568",	"r_xnm5569",	"r_xnm5570",	"r_xnm5571",	"r_xnm5572",	"r_xnm5573",	"r_xnm5574",	"r_xnm5575",	"r_xnm5576",	"r_xnm5577",	"r_xnm5578",	"r_xnm5579",	"r_xnm5580",	"r_xnm5581",	"r_xnm5582",	"r_xnm5583",	"r_xnm5584",	"r_xnm5585",	"r_xnm5586",	"r_xnm5587",	"r_xnm5588",	"r_xnm5589",	"r_xnm5590",	"r_xnm5591",	"r_xnm5592",	"r_xnm5593",	"r_xnm5594",	"r_xnm5595",	"r_xnm5596",	"r_xnm5597",	"r_xnm5598",	"r_xnm5599",	"r_xnm5600",	"r_xnm5601",	"r_xnm5602",	"r_xnm5603",	"r_xnm5604",	"r_xnm5605",	"r_xnm5606",	"r_xnm5607",	"r_xnm5608",	"r_xnm5609",	"r_xnm5610",	"r_xnm5611",	"r_xnm5612",	"r_xnm5613",	"r_xnm5614",	"r_xnm5615",	"r_xnm5616",	"r_xnm5617",	"r_xnm5618",	"r_xnm5619",	"r_xnm5620",	"r_xnm5621",	"r_xnm5622",	"r_xnm5623",	"r_xnm5624",	"r_xnm5625",	"r_xnm5626",	"r_xnm5627",	"r_xnm5628",	"r_xnm5629",	"r_xnm5630",	"r_xnm5631",	"r_xnm5632",	"r_xnm5633",	"r_xnm5634",	"r_xnm5635",	"r_xnm5636",	"r_xnm5637",	"r_xnm5638",	"r_xnm5639",	"r_xnm5640",	"r_xnm5641",	"r_xnm5642",	"r_xnm5643",	"r_xnm5644",	"r_xnm5645",	"r_xnm5646",	"r_xnm5647",	"r_xnm5648",	"r_xnm5649",	"r_xnm5650",	"r_xnm5651",	"r_xnm5652",	"r_xnm5653",	"r_xnm5654",	"r_xnm5655",	"r_xnm5656",	"r_xnm5657",	"r_xnm5658",	"r_xnm5659",	"r_xnm5660",	"r_xnm5661",	"r_xnm5662",	"r_xnm5663",	"r_xnm5664",	"r_xnm5665",	"r_xnm5666",	"r_xnm5667",	"r_xnm5668",	"r_xnm5669",	"r_xnm5670",	"r_xnm5671",	"r_xnm5672",	"r_xnm5673",	"r_xnm5674",	"r_xnm5675",	"r_xnm5676",	"r_xnm5677",	"r_xnm5678",	"r_xnm5679",	"r_xnm5680",	"r_xnm5681",	"r_xnm5682",	"r_xnm5683",	"r_xnm5684",	"r_xnm5685",	"r_xnm5686",	"r_xnm5687",	"r_xnm5688",	"r_xnm5689",	"r_xnm5690",	"r_xnm5691",	"r_xnm5692",	"r_xnm5693",	"r_xnm5694",	"r_xnm5695",	"r_xnm5696",	"r_xnm5697",	"r_xnm5698",	"r_xnm5699",	"r_xnm5700",	"r_xnm5701",	"r_xnm5702",	"r_xnm5703",	"r_xnm5704",	"r_xnm5705",	"r_xnm5706",	"r_xnm5707",	"r_xnm5708",	"r_xnm5709",	"r_xnm5710",	"r_xnm5711",	"r_xnm5712",	"r_xnm5713",	"r_xnm5714",	"r_xnm5715",	"r_xnm5716",	"r_xnm5717",	"r_xnm5718",	"r_xnm5719",	"r_xnm5720",	"r_xnm5721",	"r_xnm5722",	"r_xnm5723",	"r_xnm5724",	"r_xnm5725",	"r_xnm5726",	"r_xnm5727",	"r_xnm5728",	"r_xnm5729",	"r_xnm5730",	"r_xnm5731",	"r_xnm5732",	"r_xnm5733",	"r_xnm5734",	"r_xnm5735",	"r_xnm5736",	"r_xnm5737",	"r_xnm5738",	"r_xnm5739",	"r_xnm5740",	"r_xnm5741",	"r_xnm5742",	"r_xnm5743",	"r_xnm5744",	"r_xnm5745",	"r_xnm5746",	"r_xnm5747",	"r_xnm5748",	"r_xnm5749",	"r_xnm5750",	"r_xnm5751",	"r_xnm5752",	"r_xnm5753",	"r_xnm5754",	"r_xnm5755",	"r_xnm5756",	"r_xnm5757",	"r_xnm5758",	"r_xnm5759",	"r_xnm5760",	"r_xnm5761",	"r_xnm5762",	"r_xnm5763",	"r_xnm5764",	"r_xnm5765",	"r_xnm5766",	"r_xnm5767",	"r_xnm5768",	"r_xnm5769",	"r_xnm5770",	"r_xnm5771",	"r_xnm5772",	"r_xnm5773",	"r_xnm5774",	"r_xnm5775",	"r_xnm5776",	"r_xnm5777",	"r_xnm5778",	"r_xnm5779",	"r_xnm5780",	"r_xnm5781",	"r_xnm5782",	"r_xnm5783",	"r_xnm5784",	"r_xnm5785",	"r_xnm5786",	"r_xnm5787",	"r_xnm5788",	"r_xnm5789",	"r_xnm5790",	"r_xnm5791",	"r_xnm5792",	"r_xnm5793",	"r_xnm5794",	"r_xnm5795",	"r_xnm5796",	"r_xnm5797",	"r_xnm5798",	"r_xnm5799",	"r_xnm5800",	"r_xnm5801",	"r_xnm5802",	"r_xnm5803",	"r_xnm5804",	"r_xnm5805",	"r_xnm5806",	"r_xnm5807",	"r_xnm5808",	"r_xnm5809",	"r_xnm5810",	"r_xnm5811",	"r_xnm5812",	"r_xnm5813",	"r_xnm5814",	"r_xnm5815",	"r_xnm5816",	"r_xnm5817",	"r_xnm5818",	"r_xnm5819",	"r_xnm5820",	"r_xnm5821",	"r_xnm5822",	"r_xnm5823",	"r_xnm5824",	"r_xnm5825",	"r_xnm5826",	"r_xnm5827",	"r_xnm5828",	"r_xnm5829",	"r_xnm5830",	"r_xnm5831",	"r_xnm5832",	"r_xnm5833",	"r_xnm5834",	"r_xnm5835",	"r_xnm5836",	"r_xnm5837",	"r_xnm5838",	"r_xnm5839",	"r_xnm5840",	"r_xnm5841",	"r_xnm5842",	"r_xnm5843",	"r_xnm5844",	"r_xnm5845",	"r_xnm5846",	"r_xnm5847",	"r_xnm5848",	"r_xnm5849",	"r_xnm5850",	"r_xnm5851",	"r_xnm5852",	"r_xnm5853",	"r_xnm5854",	"r_xnm5855",	"r_xnm5856",	"r_xnm5857",	"r_xnm5858",	"r_xnm5859",	"r_xnm5860",	"r_xnm5861",	"r_xnm5862",	"r_xnm5863",	"r_xnm5864",	"r_xnm5865",	"r_xnm5866",	"r_xnm5867",	"r_xnm5868",	"r_xnm5869",	"r_xnm5870",	"r_xnm5871",	"r_xnm5872",	"r_xnm5873",	"r_xnm5874",	"r_xnm5875",	"r_xnm5876",	"r_xnm5877",	"r_xnm5878",	"r_xnm5879",	"r_xnm5880",	"r_xnm5881",	"r_xnm5882",	"r_xnm5883",	"r_xnm5884",	"r_xnm5885",	"r_xnm5886",	"r_xnm5887",	"r_xnm5888",	"r_xnm5889",	"r_xnm5890",	"r_xnm5891",	"r_xnm5892",	"r_xnm5893",	"r_xnm5894",	"r_xnm5895",	"r_xnm5896",	"r_xnm5897",	"r_xnm5898",	"r_xnm5899",	"r_xnm5900",	"r_xnm5901",	"r_xnm5902",	"r_xnm5903",	"r_xnm5904",	"r_xnm5905",	"r_xnm5906",	"r_xnm5907",	"r_xnm5908",	"r_xnm5909",	"r_xnm5910",	"r_xnm5911",	"r_xnm5912",	"r_xnm5913",	"r_xnm5914",	"r_xnm5915",	"r_xnm5916",	"r_xnm5917",	"r_xnm5918",	"r_xnm5919",	"r_xnm5920",	"r_xnm5921",	"r_xnm5922",	"r_xnm5923",	"r_xnm5924",	"r_xnm5925",	"r_xnm5926",	"r_xnm5927",	"r_xnm5928",	"r_xnm5929",	"r_xnm5930",	"r_xnm5931",	"r_xnm5932",	"r_xnm5933",	"r_xnm5934",	"r_xnm5935",	"r_xnm5936",	"r_xnm5937",	"r_xnm5938",	"r_xnm5939",	"r_xnm5940",	"r_xnm5941",	"r_xnm5942",	"r_xnm5943",	"r_xnm5944",	"r_xnm5945",	"r_xnm5946",	"r_xnm5947",	"r_xnm5948",	"r_xnm5949",	"r_xnm5950",	"r_xnm5951",	"r_xnm5952",	"r_xnm5953",	"r_xnm5954",	"r_xnm5955",	"r_xnm5956",	"r_xnm5957",	"r_xnm5958",	"r_xnm5959",	"r_xnm5960",	"r_xnm5961",	"r_xnm5962",	"r_xnm5963",	"r_xnm5964",	"r_xnm5965",	"r_xnm5966",	"r_xnm5967",	"r_xnm5968",	"r_xnm5969",	"r_xnm5970",	"r_xnm5971",	"r_xnm5972",	"r_xnm5973",	"r_xnm5974",	"r_xnm5975",	"r_xnm5976",	"r_xnm5977",	"r_xnm5978",	"r_xnm5979",	"r_xnm5980",	"r_xnm5981",	"r_xnm5982",	"r_xnm5983",	"r_xnm5984",	"r_xnm5985",	"r_xnm5986",	"r_xnm5987",	"r_xnm5988",	"r_xnm5989",	"r_xnm5990",	"r_xnm5991",	"r_xnm5992",	"r_xnm5993",	"r_xnm5994",	"r_xnm5995",	"r_xnm5996",	"r_xnm5997",	"r_xnm5998",	"r_xnm5999",	"r_xnm6000",	"r_xnm6001",	"r_xnm6002",	"r_xnm6003",	"r_xnm6004",	"r_xnm6005",	"r_xnm6006",	"r_xnm6007",	"r_xnm6008",	"r_xnm6009",	"r_xnm6010",	"r_xnm6011",	"r_xnm6012",	"r_xnm6013",	"r_xnm6014",	"r_xnm6015",	"r_xnm6016",	"r_xnm6017",	"r_xnm6018",	"r_xnm6019",	"r_xnm6020",	"r_xnm6021",	"r_xnm6022",	"r_xnm6023",	"r_xnm6024",	"r_xnm6025",	"r_xnm6026",	"r_xnm6027",	"r_xnm6028",	"r_xnm6029",	"r_xnm6030",	"r_xnm6031",	"r_xnm6032",	"r_xnm6033",	"r_xnm6034",	"r_xnm6035",	"r_xnm6036",	"r_xnm6037",	"r_xnm6038",	"r_xnm6039",	"r_xnm6040",	"r_xnm6041",	"r_xnm6042",	"r_xnm6043",	"r_xnm6044",	"r_xnm6045",	"r_xnm6046",	"r_xnm6047",	"r_xnm6048",	"r_xnm6049",	"r_xnm6050",	"r_xnm6051",	"r_xnm6052",	"r_xnm6053",	"r_xnm6054",	"r_xnm6055",	"r_xnm6056",	"r_xnm6057",	"r_xnm6058",	"r_xnm6059",	"r_xnm6060",	"r_xnm6061",	"r_xnm6062",	"r_xnm6063",	"r_xnm6064",	"r_xnm6065",	"r_xnm6066",	"r_xnm6067",	"r_xnm6068",	"r_xnm6069",	"r_xnm6070",	"r_xnm6071",	"r_xnm6072",	"r_xnm6073",	"r_xnm6074",	"r_xnm6075",	"r_xnm6076",	"r_xnm6077",	"r_xnm6078",	"r_xnm6079",	"r_xnm6080",	"r_xnm6081",	"r_xnm6082",	"r_xnm6083",	"r_xnm6084",	"r_xnm6085",	"r_xnm6086",	"r_xnm6087",	"r_xnm6088",	"r_xnm6089",	"r_xnm6090",	"r_xnm6091",	"r_xnm6092",	"r_xnm6093",	"r_xnm6094",	"r_xnm6095",	"r_xnm6096",	"r_xnm6097",	"r_xnm6098",	"r_xnm6099",	"r_xnm6100",	"r_xnm6101",	"r_xnm6102",	"r_xnm6103",	"r_xnm6104",	"r_xnm6105",	"r_xnm6106",	"r_xnm6107",	"r_xnm6108",	"r_xnm6109",	"r_xnm6110",	"r_xnm6111",	"r_xnm6112",	"r_xnm6113",	"r_xnm6114",	"r_xnm6115",	"r_xnm6116",	"r_xnm6117",	"r_xnm6118",	"r_xnm6119",	"r_xnm6120",	"r_xnm6121",	"r_xnm6122",	"r_xnm6123",	"r_xnm6124",	"r_xnm6125",	"r_xnm6126",	"r_xnm6127",	"r_xnm6128",	"r_xnm6129",	"r_xnm6130",	"r_xnm6131",	"r_xnm6132",	"r_xnm6133",	"r_xnm6134",	"r_xnm6135",	"r_xnm6136",	"r_xnm6137",	"r_xnm6138",	"r_xnm6139",	"r_xnm6140",	"r_xnm6141",	"r_xnm6142",	"r_xnm6143",	"r_xnm6144",	"r_xnm6145",	"r_xnm6146",	"r_xnm6147",	"r_xnm6148",	"r_xnm6149",	"r_xnm6150",	"r_xnm6151",	"r_xnm6152",	"r_xnm6153",	"r_xnm6154",	"r_xnm6155",	"r_xnm6156",	"r_xnm6157",	"r_xnm6158",	"r_xnm6159",	"r_xnm6160",	"r_xnm6161",	"r_xnm6162",	"r_xnm6163",	"r_xnm6164",	"r_xnm6165",	"r_xnm6166",	"r_xnm6167",	"r_xnm6168",	"r_xnm6169",	"r_xnm6170",	"r_xnm6171",	"r_xnm6172",	"r_xnm6173",	"r_xnm6174",	"r_xnm6175",	"r_xnm6176",	"r_xnm6177",	"r_xnm6178",	"r_xnm6179",	"r_xnm6180",	"r_xnm6181",	"r_xnm6182",	"r_xnm6183",	"r_xnm6184",	"r_xnm6185",	"r_xnm6186",	"r_xnm6187",	"r_xnm6188",	"r_xnm6189",	"r_xnm6190",	"r_xnm6191",	"r_xnm6192",	"r_xnm6193",	"r_xnm6194",	"r_xnm6195",	"r_xnm6196",	"r_xnm6197",	"r_xnm6198",	"r_xnm6199",	"r_xnm6200",	"r_xnm6201",	"r_xnm6202",	"r_xnm6203",	"r_xnm6204",	"r_xnm6205",	"r_xnm6206",	"r_xnm6207",	"r_xnm6208",	"r_xnm6209",	"r_xnm6210",	"r_xnm6211",	"r_xnm6212",	"r_xnm6213",	"r_xnm6214",	"r_xnm6215",	"r_xnm6216",	"r_xnm6217",	"r_xnm6218",	"r_xnm6219",	"r_xnm6220",	"r_xnm6221",	"r_xnm6222",	"r_xnm6223",	"r_xnm6224",	"r_xnm6225",	"r_xnm6226",	"r_xnm6227",	"r_xnm6228",	"r_xnm6229",	"r_xnm6230",	"r_xnm6231",	"r_xnm6232",	"r_xnm6233",	"r_xnm6234",	"r_xnm6235",	"r_xnm6236",	"r_xnm6237",	"r_xnm6238",	"r_xnm6239",	"r_xnm6240",	"r_xnm6241",	"r_xnm6242",	"r_xnm6243",	"r_xnm6244",	"r_xnm6245",	"r_xnm6246",	"r_xnm6247",	"r_xnm6248",	"r_xnm6249",	"r_xnm6250",	"r_xnm6251",	"r_xnm6252",	"r_xnm6253",	"r_xnm6254",	"r_xnm6255",	"r_xnm6256",	"r_xnm6257",	"r_xnm6258",	"r_xnm6259",	"r_xnm6260",	"r_xnm6261",	"r_xnm6262",	"r_xnm6263",	"r_xnm6264",	"r_xnm6265",	"r_xnm6266",	"r_xnm6267",	"r_xnm6268",	"r_xnm6269",	"r_xnm6270",	"r_xnm6271",	"r_xnm6272",	"r_xnm6273",	"r_xnm6274",	"r_xnm6275",	"r_xnm6276",	"r_xnm6277",	"r_xnm6278",	"r_xnm6279",	"r_xnm6280",	"r_xnm6281",	"r_xnm6282",	"r_xnm6283",	"r_xnm6284",	"r_xnm6285",	"r_xnm6286",	"r_xnm6287",	"r_xnm6288",	"r_xnm6289",	"r_xnm6290",	"r_xnm6291",	"r_xnm6292",	"r_xnm6293",	"r_xnm6294",	"r_xnm6295",	"r_xnm6296",	"r_xnm6297",	"r_xnm6298",	"r_xnm6299",	"r_xnm6300",	"r_xnm6301",	"r_xnm6302",	"r_xnm6303",	"r_xnm6304",	"r_xnm6305",	"r_xnm6306",	"r_xnm6307",	"r_xnm6308",	"r_xnm6309",	"r_xnm6310",	"r_xnm6311",	"r_xnm6312",	"r_xnm6313",	"r_xnm6314",	"r_xnm6315",	"r_xnm6316",	"r_xnm6317",	"r_xnm6318",	"r_xnm6319",	"r_xnm6320",	"r_xnm6321",	"r_xnm6322",	"r_xnm6323",	"r_xnm6324",	"r_xnm6325",	"r_xnm6326",	"r_xnm6327",	"r_xnm6328",	"r_xnm6329",	"r_xnm6330",	"r_xnm6331",	"r_xnm6332",	"r_xnm6333",	"r_xnm6334",	"r_xnm6335",	"r_xnm6336",	"r_xnm6337",	"r_xnm6338",	"r_xnm6339",	"r_xnm6340",	"r_xnm6341",	"r_xnm6342",	"r_xnm6343",	"r_xnm6344",	"r_xnm6345",	"r_xnm6346",	"r_xnm6347",	"r_xnm6348",	"r_xnm6349",	"r_xnm6350",	"r_xnm6351",	"r_xnm6352",	"r_xnm6353",	"r_xnm6354",	"r_xnm6355",	"r_xnm6356",	"r_xnm6357",	"r_xnm6358",	"r_xnm6359",	"r_xnm6360",	"r_xnm6361",	"r_xnm6362",	"r_xnm6363",	"r_xnm6364",	"r_xnm6365",	"r_xnm6366",	"r_xnm6367",	"r_xnm6368",	"r_xnm6369",	"r_xnm6370",	"r_xnm6371",	"r_xnm6372",	"r_xnm6373",	"r_xnm6374",	"r_xnm6375",	"r_xnm6376",	"r_xnm6377",	"r_xnm6378",	"r_xnm6379",	"r_xnm6380",	"r_xnm6381",	"r_xnm6382",	"r_xnm6383",	"r_xnm6384",	"r_xnm6385",	"r_xnm6386",	"r_xnm6387",	"r_xnm6388",	"r_xnm6389",	"r_xnm6390",	"r_xnm6391",	"r_xnm6392",	"r_xnm6393",	"r_xnm6394",	"r_xnm6395",	"r_xnm6396",	"r_xnm6397",	"r_xnm6398",	"r_xnm6399",	"r_xnm6400",	"r_xnm6401",	"r_xnm6402",	"r_xnm6403",	"r_xnm6404",	"r_xnm6405",	"r_xnm6406",	"r_xnm6407",	"r_xnm6408",	"r_xnm6409",	"r_xnm6410",	"r_xnm6411",	"r_xnm6412",	"r_xnm6413",	"r_xnm6414",	"r_xnm6415",	"r_xnm6416",	"r_xnm6417",	"r_xnm6418",	"r_xnm6419",	"r_xnm6420",	"r_xnm6421",	"r_xnm6422",	"r_xnm6423",	"r_xnm6424",	"r_xnm6425",	"r_xnm6426",	"r_xnm6427",	"r_xnm6428",	"r_xnm6429",	"r_xnm6430",	"r_xnm6431",	"r_xnm6432",	"r_xnm6433",	"r_xnm6434",	"r_xnm6435",	"r_xnm6436",	"r_xnm6437",	"r_xnm6438",	"r_xnm6439",	"r_xnm6440",	"r_xnm6441",	"r_xnm6442",	"r_xnm6443",	"r_xnm6444",	"r_xnm6445",	"r_xnm6446",	"r_xnm6448",	"r_xnm6449",	"r_xnm6451",	"r_xnm6452",	"r_xnm6454",	"r_xnm6455",	"r_xnm6456",	"r_xnm6457",	"r_xnm6458",	"r_xnm6459",	"r_xnm6460",	"r_xnm6461",	"r_xnm6462",	"r_xnm6463",	"r_xnm6464",	"r_xnm6465",	"r_xnm6467",	"r_xnm6468",	"r_xnm6469",	"r_xnm6471",	"r_xnm6472",	"r_xnm6473",	"r_xnm6474",	"r_xnm6475",	"r_xnm6476",	"r_xnm6477",	"r_xnm6478",	"r_xnm6479",	"r_xnm6480",	"r_xnm6481",	"r_xnm6482",	"r_xnm6483",	"r_xnm6484",	"r_xnm6485",	"r_xnm6486",	"r_xnm6487",	"r_xnm6488",	"r_xnm6489",	"r_xnm6490",	"r_xnm6491",	"r_xnm6492",	"r_xnm6493",	"r_xnm6494",	"r_xnm6495",	"r_xnm6496",	"r_xnm6500",	"r_xnm6501",	"r_xnm6502",	"r_xnm6503",	"r_xnm6504",	"r_xnm6505",	"r_xnm6506",	"r_xnm6507",	"r_xnm6508",	"r_xnm6509",	"r_xnm6510",	"r_xnm6511",	"r_xnm6512",	"r_xnm6513",	"r_xnm6515",	"r_xnm6516",	"r_xnm6517",	"r_xnm6518",	"r_xnm6519",	"r_xnm6520",	"r_xnm6521",	"r_xnm6522",	"r_xnm6523",	"r_xnm6524",	"r_xnm6525",	"r_xnm6526",	"r_xnm6527",	"r_xnm6531",	"r_xnm6532",	"r_xnm6533",	"r_xnm6534",	"r_xnm6535",	"r_xnm6537",	"r_xnm6539",	"r_xnm6541",	"r_xnm6543",	"r_xnm6544",	"r_xnm6545",	"r_xnm6547",	"r_xnm6548",	"r_xnm6549",	"r_xnm6550",	"r_xnm6552",	"r_xnm6553",	"r_xnm6554",	"r_xnm6555",	"r_xnm6556",	"r_xnm6557",	"resp_score",	"response_band",	"vantage",	"vantage_band",
]
## PROC SORT equivalent
# Sort the DataFrame
df_sorted = df.sort_values(['LoanNumber', 'everdef', 'weight'])

## PROC TRANSPOSE equivalent (first one)
# Melt the DataFrame to transform numeric variables

num_attr = pd.melt(df_sorted, 
                  id_vars=['year', 'month', 'LoanNumber', 'everdef', 'weight'],
                  value_vars=cn_list,
                  var_name='x_nm',
                  value_name='x_value')

## PROC SQL equivalent for missing pattern exploration
missing_n_pattern = num_attr.groupby(['year', 'month', 'x_nm']).apply(
    lambda x: x['x_value'].isna().mean()
).reset_index(name='missPCT')

missing_n_pattern = missing_n_pattern.sort_values(['year', 'month', 'x_nm'])#export the tableinto EXCEL for checking 


## PROC TRANSPOSE equivalent (second one)
# Pivot the missing pattern data
missing_n_pattern2 = missing_n_pattern.pivot_table(
    index=['year', 'month'],
    columns='x_nm',
    values='missPCT'
).reset_index()

# Optional: Rename columns if needed
#missing_n_pattern2.columns.name = None  # Remove the column name for x_n
#extract kind='String' attributes and input them as below for char_list 
char_list=["Carrier",	"IncomeSource",	"Message",	"PaydayMethod",	"Ref",	"UNIT_TYPE",	"r_xnm1",	"r_xnm2",	"r_xnm9",	"r_xnm10",	"r_xnm21",	"r_xnm22",	"r_xnm23",	"r_xnm24",	"r_xnm25",	"r_xnm56",	"r_xnm88",	"r_xnm421",	"r_xnm849",	"r_xnm852",	"r_xnm864",	"r_xnm875",	"r_xnm886",	"r_xnm898",	"r_xnm910",	"r_xnm922",	"r_xnm934",	"r_xnm946",	"r_xnm958",	"r_xnm970",	"r_xnm981",	"r_xnm993",	"r_xnm1005",	"r_xnm1017",	"r_xnm1029",	"r_xnm1041",	"r_xnm1053",	"r_xnm1065",	"r_xnm1077",	"r_xnm1089",	"r_xnm1101",	"r_xnm1113",	"r_xnm1124",	"r_xnm1136",	"r_xnm1148",	"r_xnm1160",	"r_xnm1172",	"r_xnm1184",	"r_xnm1196",	"r_xnm1208",	"r_xnm1220",	"r_xnm1232",	"r_xnm1244",	"r_xnm1256",	"r_xnm1267",	"r_xnm1279",	"r_xnm1291",	"r_xnm1303",	"r_xnm1315",	"r_xnm1326",	"r_xnm1338",	"r_xnm1350",	"r_xnm1362",	"r_xnm1374",	"r_xnm1386",	"r_xnm1398",	"r_xnm1410",	"r_xnm1421",	"r_xnm1433",	"r_xnm1445",	"r_xnm1457",	"r_xnm1469",	"r_xnm1481",	"r_xnm1493",	"r_xnm1505",	"r_xnm1517",	"r_xnm1529",	"r_xnm1541",	"r_xnm1553",	"r_xnm1564",	"r_xnm1576",	"r_xnm1588",	"r_xnm1600",	"r_xnm1612",	"r_xnm1622",	"r_xnm1624",	"r_xnm1662",	"r_xnm1682",	"r_xnm1708",	"r_xnm1709",	"r_xnm1711",	"r_xnm1725",	"r_xnm1734",	"r_xnm1743",	"r_xnm1752",	"r_xnm1761",	"r_xnm1770",	"r_xnm1779",	"r_xnm1855",	"r_xnm1864",	"r_xnm1872",	"r_xnm1879",	"r_xnm1885",	"r_xnm1889",	"r_xnm1893",	"r_xnm1896",	"r_xnm1898",	"r_xnm1903",	"r_xnm1913",	"r_xnm1923",	"r_xnm1933",	"r_xnm1943",	"r_xnm1953",	"r_xnm1963",	"r_xnm1973",	"r_xnm1983",	"r_xnm2002",	"r_xnm2008",	"r_xnm2075",	"r_xnm2087",	"r_xnm2140",	"r_xnm2197",	"r_xnm2198",	"r_xnm2199",	"r_xnm2200",	"r_xnm2207",	"r_xnm2208",	"r_xnm2209",	"r_xnm2210",	"r_xnm2217",	"r_xnm2240",	"r_xnm2242",	"r_xnm2533",	"r_xnm2534",	"r_xnm6447",	"r_xnm6450",	"r_xnm6453",	"r_xnm6466",	"r_xnm6470",	"r_xnm6497",	"r_xnm6498",	"r_xnm6499",	"r_xnm6514",	"r_xnm6528",	"r_xnm6529",	"r_xnm6530",	"r_xnm6536",	"r_xnm6538",	"r_xnm6540",	"r_xnm6542",	"r_xnm6546",	"r_xnm6551",	"r_xnm6558",
]



char_attr = pd.melt(df_sorted, 
                  id_vars=['year', 'month', 'LoanNumber', 'everdef', 'weight'],
                  value_vars=char_list,
                  var_name='x_nm',
                  value_name='x_value')
#loading lookup table from bureau doc**
RSVP_RISK_NEW_LU=pd.read_csv(r'D:\HA_study_plan_tasks\banking_financial_scorecard_dev\model_1\lu\RSVP_RISK_NEW_LU.csv')

query_tempdate = """
SELECT DISTINCT a.x_nm, b.f1 AS description
FROM char_attr a
LEFT JOIN RSVP_RISK_NEW_LU b ON a.x_nm = b._x_nm
WHERE x_value LIKE '%-%-%' 
"""
#excluding date type attributes
tempdate = pysqldf(query_tempdate)

query_missing = """
SELECT year, month, x_nm,
       SUM(CASE WHEN x_value IS NULL OR x_value = '' THEN 1 ELSE 0 END) * 1.0 / COUNT(*) AS missPCT
FROM char_attr
WHERE x_nm NOT IN (SELECT x_nm FROM tempdate)
GROUP BY year, month, x_nm
ORDER BY year, month, x_nm
"""
missing_c_pattern = pysqldf(query_missing)


## PROC SQL equivalent for missing pattern calculation
# missing_n_pattern = char_attr.groupby(['year', 'month', 'x_nm']).apply(
#     lambda x: x['x_value'].isna().mean()
# ).reset_index(name='missPCT')



## PROC TRANSPOSE equivalent (second one)
# Pivot the missing pattern data
missing_c_pattern = missing_c_pattern.pivot_table(
    index=['year', 'month'],
    columns='x_nm',
    values='missPCT'
).reset_index()


n_m_r_list=["ALL0100",	"ALL0200",	"ALL0216",	"ALL0317",	"ALL0336",	"ALL0400",	"ALL0416",	"ALL0436",	"ALL0446",	"ALL0448",	"ALL0700",	"ALL1380",	"ALL2000",	"ALL2001",	"ALL2002",	"ALL2005",	"ALL2106",	"ALL2136",	"ALL2176",	"ALL2326",	"ALL2327",	"ALL2357",	"ALL2380",	"ALL2387",	"ALL2427",	"ALL2428",	"ALL2480",	"ALL2707",	"ALL2870",	"ALL2907",	"ALL2937",	"ALL2967",	"ALL2990",	"ALL3517",	"ALL4080",	"ALL5047",	"ALL5070",	"ALL5743",	"ALL8026",	"ALL8164",	"ALL8167",	"ALL8325",	"ALX5839",	"anual_income",	"AUA0416",	"AUA1300",	"AUA2358",	"AUA5820",	"AUA6160",	"BCA0416",	"BCA5030",	"BCA5070",	"BCA5740",	"BCA6220",	"BCC0436",	"BCC0446",	"BCC1360",	"BCC3342",	"BCC3345",	"BCC3423",	"BCC3512",	"BCC5020",	"BCC5320",	"BCC5421",	"BCC5620",	"BCC5627",	"BCC7117",	"BCC8120",	"BCX3421",	"BCX3422",	"BRC0416",	"BRC1300",	"BRC5320",	"BRC5830",	"BRC7140",	"COL3210",	"COL5060",	"COL8165",	"FIP0300",	"FIP2358",	"International",	"IQT9415",	"IQT9420",	"IQT9425",	"IQT9426",	"MTA6160",	"MTF2358",	"MTF8166",	"MTX5839",	"PIL5020",	"r_xnm100",	"r_xnm1000",	"r_xnm1001",	"r_xnm1002",	"r_xnm1003",	"r_xnm1004",	"r_xnm1006",	"r_xnm1007",	"r_xnm1008",	"r_xnm1009",	"r_xnm101",	"r_xnm1010",	"r_xnm1011",	"r_xnm1012",	"r_xnm1013",	"r_xnm1014",	"r_xnm1015",	"r_xnm1016",	"r_xnm1018",	"r_xnm1019",	"r_xnm102",	"r_xnm1020",	"r_xnm1021",	"r_xnm1022",	"r_xnm1023",	"r_xnm1024",	"r_xnm1025",	"r_xnm1026",	"r_xnm1027",	"r_xnm1028",	"r_xnm103",	"r_xnm1030",	"r_xnm1031",	"r_xnm1032",	"r_xnm1033",	"r_xnm1034",	"r_xnm1035",	"r_xnm1036",	"r_xnm1037",	"r_xnm1038",	"r_xnm1039",	"r_xnm104",	"r_xnm1040",	"r_xnm1042",	"r_xnm1043",	"r_xnm1044",	"r_xnm1045",	"r_xnm1046",	"r_xnm1047",	"r_xnm1048",	"r_xnm1049",	"r_xnm105",	"r_xnm1050",	"r_xnm1051",	"r_xnm1052",	"r_xnm1054",	"r_xnm1055",	"r_xnm1056",	"r_xnm1057",	"r_xnm1058",	"r_xnm1059",	"r_xnm106",	"r_xnm1060",	"r_xnm1061",	"r_xnm1062",	"r_xnm1063",	"r_xnm1064",	"r_xnm1066",	"r_xnm1067",	"r_xnm1068",	"r_xnm1069",	"r_xnm107",	"r_xnm1070",	"r_xnm1071",	"r_xnm1072",	"r_xnm1073",	"r_xnm1074",	"r_xnm1075",	"r_xnm1076",	"r_xnm1078",	"r_xnm1079",	"r_xnm108",	"r_xnm1080",	"r_xnm1081",	"r_xnm1082",	"r_xnm1083",	"r_xnm1084",	"r_xnm1085",	"r_xnm1086",	"r_xnm1087",	"r_xnm1088",	"r_xnm109",	"r_xnm1090",	"r_xnm1091",	"r_xnm1092",	"r_xnm1093",	"r_xnm1094",	"r_xnm1095",	"r_xnm1096",	"r_xnm1097",	"r_xnm1098",	"r_xnm1099",	"r_xnm110",	"r_xnm1100",	"r_xnm1102",	"r_xnm1103",	"r_xnm1104",	"r_xnm1105",	"r_xnm1106",	"r_xnm1107",	"r_xnm1108",	"r_xnm1109",	"r_xnm111",	"r_xnm1110",	"r_xnm1111",	"r_xnm1112",	"r_xnm1114",	"r_xnm1115",	"r_xnm1116",	"r_xnm1117",	"r_xnm1118",	"r_xnm1119",	"r_xnm112",	"r_xnm1120",	"r_xnm1121",	"r_xnm1122",	"r_xnm1123",	"r_xnm1125",	"r_xnm1126",	"r_xnm1127",	"r_xnm1128",	"r_xnm1129",	"r_xnm113",	"r_xnm1130",	"r_xnm1131",	"r_xnm1132",	"r_xnm1133",	"r_xnm1134",	"r_xnm1135",	"r_xnm1137",	"r_xnm1138",	"r_xnm1139",	"r_xnm114",	"r_xnm1140",	"r_xnm1141",	"r_xnm1142",	"r_xnm1143",	"r_xnm1144",	"r_xnm1145",	"r_xnm1146",	"r_xnm1147",	"r_xnm1149",	"r_xnm115",	"r_xnm1150",	"r_xnm1151",	"r_xnm1152",	"r_xnm1153",	"r_xnm1154",	"r_xnm1155",	"r_xnm1156",	"r_xnm1157",	"r_xnm1158",	"r_xnm1159",	"r_xnm116",	"r_xnm1161",	"r_xnm1162",	"r_xnm1163",	"r_xnm1164",	"r_xnm1165",	"r_xnm1166",	"r_xnm1167",	"r_xnm1168",	"r_xnm1169",	"r_xnm117",	"r_xnm1170",	"r_xnm1171",	"r_xnm1173",	"r_xnm1174",	"r_xnm1175",	"r_xnm1176",	"r_xnm1177",	"r_xnm1178",	"r_xnm1179",	"r_xnm118",	"r_xnm1180",	"r_xnm1181",	"r_xnm1182",	"r_xnm1183",	"r_xnm1185",	"r_xnm1186",	"r_xnm1187",	"r_xnm1188",	"r_xnm1189",	"r_xnm119",	"r_xnm1190",	"r_xnm1191",	"r_xnm1192",	"r_xnm1193",	"r_xnm1194",	"r_xnm1195",	"r_xnm1197",	"r_xnm1198",	"r_xnm1199",	"r_xnm120",	"r_xnm1200",	"r_xnm1201",	"r_xnm1202",	"r_xnm1203",	"r_xnm1204",	"r_xnm1205",	"r_xnm1206",	"r_xnm1207",	"r_xnm1209",	"r_xnm121",	"r_xnm1210",	"r_xnm1211",	"r_xnm1212",	"r_xnm1213",	"r_xnm1214",	"r_xnm1215",	"r_xnm1216",	"r_xnm1217",	"r_xnm1218",	"r_xnm1219",	"r_xnm122",	"r_xnm1221",	"r_xnm1222",	"r_xnm1223",	"r_xnm1224",	"r_xnm1225",	"r_xnm1226",	"r_xnm1227",	"r_xnm1228",	"r_xnm1229",	"r_xnm123",	"r_xnm1230",	"r_xnm1231",	"r_xnm1233",	"r_xnm1234",	"r_xnm1235",	"r_xnm1236",	"r_xnm1237",	"r_xnm1238",	"r_xnm1239",	"r_xnm124",	"r_xnm1240",	"r_xnm1241",	"r_xnm1242",	"r_xnm1243",	"r_xnm1245",	"r_xnm1246",	"r_xnm1247",	"r_xnm1248",	"r_xnm1249",	"r_xnm125",	"r_xnm1250",	"r_xnm1251",	"r_xnm1252",	"r_xnm1253",	"r_xnm1254",	"r_xnm1255",	"r_xnm1257",	"r_xnm1258",	"r_xnm1259",	"r_xnm126",	"r_xnm1260",	"r_xnm1261",	"r_xnm1262",	"r_xnm1263",	"r_xnm1264",	"r_xnm1265",	"r_xnm1266",	"r_xnm1268",	"r_xnm1269",	"r_xnm127",	"r_xnm1270",	"r_xnm1271",	"r_xnm1272",	"r_xnm1273",	"r_xnm1274",	"r_xnm1275",	"r_xnm1276",	"r_xnm1277",	"r_xnm1278",	"r_xnm128",	"r_xnm1280",	"r_xnm1281",	"r_xnm1282",	"r_xnm1283",	"r_xnm1284",	"r_xnm1285",	"r_xnm1286",	"r_xnm1287",	"r_xnm1288",	"r_xnm1289",	"r_xnm129",	"r_xnm1290",	"r_xnm1292",	"r_xnm1293",	"r_xnm1294",	"r_xnm1295",	"r_xnm1296",	"r_xnm1297",	"r_xnm1298",	"r_xnm1299",	"r_xnm13",	"r_xnm130",	"r_xnm1300",	"r_xnm1301",	"r_xnm1302",	"r_xnm1304",	"r_xnm1305",	"r_xnm1306",	"r_xnm1307",	"r_xnm1308",	"r_xnm1309",	"r_xnm131",	"r_xnm1310",	"r_xnm1311",	"r_xnm1312",	"r_xnm1313",	"r_xnm1314",	"r_xnm1316",	"r_xnm1317",	"r_xnm1318",	"r_xnm1319",	"r_xnm132",	"r_xnm1320",	"r_xnm1321",	"r_xnm1322",	"r_xnm1323",	"r_xnm1324",	"r_xnm1325",	"r_xnm1327",	"r_xnm1328",	"r_xnm1329",	"r_xnm133",	"r_xnm1330",	"r_xnm1331",	"r_xnm1332",	"r_xnm1333",	"r_xnm1334",	"r_xnm1335",	"r_xnm1336",	"r_xnm1337",	"r_xnm1339",	"r_xnm134",	"r_xnm1340",	"r_xnm1341",	"r_xnm1342",	"r_xnm1343",	"r_xnm1344",	"r_xnm1345",	"r_xnm1346",	"r_xnm1347",	"r_xnm1348",	"r_xnm1349",	"r_xnm135",	"r_xnm1351",	"r_xnm1352",	"r_xnm1353",	"r_xnm1354",	"r_xnm1355",	"r_xnm1356",	"r_xnm1357",	"r_xnm1358",	"r_xnm1359",	"r_xnm136",	"r_xnm1360",	"r_xnm1361",	"r_xnm1363",	"r_xnm1364",	"r_xnm1365",	"r_xnm1366",	"r_xnm1367",	"r_xnm1368",	"r_xnm1369",	"r_xnm137",	"r_xnm1370",	"r_xnm1371",	"r_xnm1372",	"r_xnm1373",	"r_xnm1375",	"r_xnm1376",	"r_xnm1377",	"r_xnm1378",	"r_xnm1379",	"r_xnm138",	"r_xnm1380",	"r_xnm1381",	"r_xnm1382",	"r_xnm1383",	"r_xnm1384",	"r_xnm1385",	"r_xnm1387",	"r_xnm1388",	"r_xnm1389",	"r_xnm139",	"r_xnm1390",	"r_xnm1391",	"r_xnm1392",	"r_xnm1393",	"r_xnm1394",	"r_xnm1395",	"r_xnm1396",	"r_xnm1397",	"r_xnm1399",	"r_xnm14",	"r_xnm140",	"r_xnm1400",	"r_xnm1401",	"r_xnm1402",	"r_xnm1403",	"r_xnm1404",	"r_xnm1405",	"r_xnm1406",	"r_xnm1407",	"r_xnm1408",	"r_xnm1409",	"r_xnm141",	"r_xnm1411",	"r_xnm1412",	"r_xnm1413",	"r_xnm1414",	"r_xnm1415",	"r_xnm1416",	"r_xnm1417",	"r_xnm1418",	"r_xnm1419",	"r_xnm142",	"r_xnm1420",	"r_xnm1422",	"r_xnm1423",	"r_xnm1424",	"r_xnm1425",	"r_xnm1426",	"r_xnm1427",	"r_xnm1428",	"r_xnm1429",	"r_xnm143",	"r_xnm1430",	"r_xnm1431",	"r_xnm1432",	"r_xnm1434",	"r_xnm1435",	"r_xnm1436",	"r_xnm1437",	"r_xnm1438",	"r_xnm1439",	"r_xnm144",	"r_xnm1440",	"r_xnm1441",	"r_xnm1442",	"r_xnm1443",	"r_xnm1444",	"r_xnm1446",	"r_xnm1447",	"r_xnm1448",	"r_xnm1449",	"r_xnm145",	"r_xnm1450",	"r_xnm1451",	"r_xnm1452",	"r_xnm1453",	"r_xnm1454",	"r_xnm1455",	"r_xnm1456",	"r_xnm1458",	"r_xnm1459",	"r_xnm146",	"r_xnm1460",	"r_xnm1461",	"r_xnm1462",	"r_xnm1463",	"r_xnm1464",	"r_xnm1465",	"r_xnm1466",	"r_xnm1467",	"r_xnm1468",	"r_xnm147",	"r_xnm1470",	"r_xnm1471",	"r_xnm1472",	"r_xnm1473",	"r_xnm1474",	"r_xnm1475",	"r_xnm1476",	"r_xnm1477",	"r_xnm1478",	"r_xnm1479",	"r_xnm148",	"r_xnm1480",	"r_xnm1482",	"r_xnm1483",	"r_xnm1484",	"r_xnm1485",	"r_xnm1486",	"r_xnm1487",	"r_xnm1488",	"r_xnm1489",	"r_xnm149",	"r_xnm1490",	"r_xnm1491",	"r_xnm1492",	"r_xnm1494",	"r_xnm1495",	"r_xnm1496",	"r_xnm1497",	"r_xnm1498",	"r_xnm1499",	"r_xnm15",	"r_xnm150",	"r_xnm1500",	"r_xnm1501",	"r_xnm1502",	"r_xnm1503",	"r_xnm1504",	"r_xnm1506",	"r_xnm1507",	"r_xnm1508",	"r_xnm1509",	"r_xnm151",	"r_xnm1510",	"r_xnm1511",	"r_xnm1512",	"r_xnm1513",	"r_xnm1514",	"r_xnm1515",	"r_xnm1516",	"r_xnm1518",	"r_xnm1519",	"r_xnm152",	"r_xnm1520",	"r_xnm1521",	"r_xnm1522",	"r_xnm1523",	"r_xnm1524",	"r_xnm1525",	"r_xnm1526",	"r_xnm1527",	"r_xnm1528",	"r_xnm153",	"r_xnm1530",	"r_xnm1531",	"r_xnm1532",	"r_xnm1533",	"r_xnm1534",	"r_xnm1535",	"r_xnm1536",	"r_xnm1537",	"r_xnm1538",	"r_xnm1539",	"r_xnm154",	"r_xnm1540",	"r_xnm1542",	"r_xnm1543",	"r_xnm1544",	"r_xnm1545",	"r_xnm1546",	"r_xnm1547",	"r_xnm1548",	"r_xnm1549",	"r_xnm155",	"r_xnm1550",	"r_xnm1551",	"r_xnm1552",	"r_xnm1554",	"r_xnm1555",	"r_xnm1556",	"r_xnm1557",	"r_xnm1558",	"r_xnm1559",	"r_xnm156",	"r_xnm1560",	"r_xnm1561",	"r_xnm1562",	"r_xnm1563",	"r_xnm1565",	"r_xnm1566",	"r_xnm1567",	"r_xnm1568",	"r_xnm1569",	"r_xnm157",	"r_xnm1570",	"r_xnm1571",	"r_xnm1572",	"r_xnm1573",	"r_xnm1574",	"r_xnm1575",	"r_xnm1577",	"r_xnm1578",	"r_xnm1579",	"r_xnm158",	"r_xnm1580",	"r_xnm1581",	"r_xnm1582",	"r_xnm1583",	"r_xnm1584",	"r_xnm1585",	"r_xnm1586",	"r_xnm1587",	"r_xnm1589",	"r_xnm159",	"r_xnm1590",	"r_xnm1591",	"r_xnm1592",	"r_xnm1593",	"r_xnm1594",	"r_xnm1595",	"r_xnm1596",	"r_xnm1597",	"r_xnm1598",	"r_xnm1599",	"r_xnm160",	"r_xnm1601",	"r_xnm1602",	"r_xnm1603",	"r_xnm1604",	"r_xnm1605",	"r_xnm1606",	"r_xnm1607",	"r_xnm1608",	"r_xnm1609",	"r_xnm161",	"r_xnm1610",	"r_xnm1611",	"r_xnm1613",	"r_xnm1614",	"r_xnm1615",	"r_xnm1616",	"r_xnm1617",	"r_xnm1618",	"r_xnm1619",	"r_xnm162",	"r_xnm1620",	"r_xnm1621",	"r_xnm1623",	"r_xnm1625",	"r_xnm1626",	"r_xnm1627",	"r_xnm1628",	"r_xnm1629",	"r_xnm163",	"r_xnm1630",	"r_xnm1631",	"r_xnm1632",	"r_xnm1633",	"r_xnm1634",	"r_xnm1637",	"r_xnm1638",	"r_xnm1639",	"r_xnm164",	"r_xnm1640",	"r_xnm1641",	"r_xnm1642",	"r_xnm1643",	"r_xnm1646",	"r_xnm1647",	"r_xnm1648",	"r_xnm1649",	"r_xnm165",	"r_xnm1650",	"r_xnm1651",	"r_xnm1652",	"r_xnm1653",	"r_xnm166",	"r_xnm1661",	"r_xnm1669",	"r_xnm167",	"r_xnm1670",	"r_xnm1671",	"r_xnm1672",	"r_xnm1673",	"r_xnm1674",	"r_xnm1675",	"r_xnm1676",	"r_xnm1677",	"r_xnm1679",	"r_xnm168",	"r_xnm1680",	"r_xnm1681",	"r_xnm1683",	"r_xnm1710",	"r_xnm1717",	"r_xnm1719",	"r_xnm1721",	"r_xnm1723",	"r_xnm1726",	"r_xnm1728",	"r_xnm173",	"r_xnm1730",	"r_xnm1732",	"r_xnm1735",	"r_xnm1737",	"r_xnm1739",	"r_xnm174",	"r_xnm1741",	"r_xnm1744",	"r_xnm1746",	"r_xnm1748",	"r_xnm175",	"r_xnm1750",	"r_xnm1753",	"r_xnm1755",	"r_xnm1757",	"r_xnm1759",	"r_xnm176",	"r_xnm1762",	"r_xnm1764",	"r_xnm1766",	"r_xnm1768",	"r_xnm177",	"r_xnm1771",	"r_xnm1773",	"r_xnm1775",	"r_xnm1777",	"r_xnm178",	"r_xnm1780",	"r_xnm1785",	"r_xnm1816",	"r_xnm1824",	"r_xnm1827",	"r_xnm1828",	"r_xnm1829",	"r_xnm183",	"r_xnm1830",	"r_xnm1831",	"r_xnm1832",	"r_xnm1833",	"r_xnm1834",	"r_xnm1835",	"r_xnm1836",	"r_xnm1837",	"r_xnm1838",	"r_xnm1839",	"r_xnm184",	"r_xnm1840",	"r_xnm1841",	"r_xnm1842",	"r_xnm1843",	"r_xnm1844",	"r_xnm1845",	"r_xnm1846",	"r_xnm1847",	"r_xnm1848",	"r_xnm1849",	"r_xnm185",	"r_xnm1850",	"r_xnm1851",	"r_xnm1852",	"r_xnm1853",	"r_xnm1854",	"r_xnm1856",	"r_xnm1857",	"r_xnm1858",	"r_xnm1859",	"r_xnm186",	"r_xnm1860",	"r_xnm1861",	"r_xnm1862",	"r_xnm1863",	"r_xnm1865",	"r_xnm1866",	"r_xnm1867",	"r_xnm1868",	"r_xnm1869",	"r_xnm187",	"r_xnm1870",	"r_xnm1871",	"r_xnm1873",	"r_xnm1874",	"r_xnm1875",	"r_xnm1876",	"r_xnm1877",	"r_xnm1878",	"r_xnm188",	"r_xnm1880",	"r_xnm1881",	"r_xnm1882",	"r_xnm1883",	"r_xnm1884",	"r_xnm1886",	"r_xnm1887",	"r_xnm1888",	"r_xnm1890",	"r_xnm1891",	"r_xnm1892",	"r_xnm1894",	"r_xnm1895",	"r_xnm1897",	"r_xnm1901",	"r_xnm1902",	"r_xnm1904",	"r_xnm1905",	"r_xnm1906",	"r_xnm1907",	"r_xnm1908",	"r_xnm1909",	"r_xnm1910",	"r_xnm1911",	"r_xnm1912",	"r_xnm1914",	"r_xnm1915",	"r_xnm1916",	"r_xnm1917",	"r_xnm1918",	"r_xnm1919",	"r_xnm1920",	"r_xnm1921",	"r_xnm1922",	"r_xnm1924",	"r_xnm1925",	"r_xnm1926",	"r_xnm1927",	"r_xnm1928",	"r_xnm1929",	"r_xnm193",	"r_xnm1930",	"r_xnm1931",	"r_xnm1932",	"r_xnm1934",	"r_xnm1935",	"r_xnm1936",	"r_xnm1937",	"r_xnm1938",	"r_xnm1939",	"r_xnm194",	"r_xnm1940",	"r_xnm1941",	"r_xnm1942",	"r_xnm1944",	"r_xnm1945",	"r_xnm1946",	"r_xnm1947",	"r_xnm1948",	"r_xnm1949",	"r_xnm195",	"r_xnm1950",	"r_xnm1951",	"r_xnm1952",	"r_xnm1954",	"r_xnm1955",	"r_xnm1956",	"r_xnm1957",	"r_xnm1958",	"r_xnm1959",	"r_xnm196",	"r_xnm1960",	"r_xnm1961",	"r_xnm1962",	"r_xnm1964",	"r_xnm1965",	"r_xnm1966",	"r_xnm1967",	"r_xnm1968",	"r_xnm1969",	"r_xnm197",	"r_xnm1970",	"r_xnm1971",	"r_xnm1972",	"r_xnm1974",	"r_xnm1975",	"r_xnm1976",	"r_xnm1977",	"r_xnm1978",	"r_xnm1979",	"r_xnm198",	"r_xnm1980",	"r_xnm1981",	"r_xnm1982",	"r_xnm1984",	"r_xnm1985",	"r_xnm1986",	"r_xnm1987",	"r_xnm1988",	"r_xnm1989",	"r_xnm199",	"r_xnm1990",	"r_xnm1991",	"r_xnm1992",	"r_xnm1993",	"r_xnm200",	"r_xnm201",	"r_xnm202",	"r_xnm203",	"r_xnm204",	"r_xnm205",	"r_xnm206",	"r_xnm207",	"r_xnm208",	"r_xnm2089",	"r_xnm209",	"r_xnm2091",	"r_xnm2092",	"r_xnm2093",	"r_xnm2094",	"r_xnm2095",	"r_xnm2096",	"r_xnm2097",	"r_xnm210",	"r_xnm2100",	"r_xnm2102",	"r_xnm2103",	"r_xnm2104",	"r_xnm2105",	"r_xnm2106",	"r_xnm2107",	"r_xnm2108",	"r_xnm211",	"r_xnm2112",	"r_xnm2114",	"r_xnm2115",	"r_xnm2116",	"r_xnm2117",	"r_xnm2118",	"r_xnm2119",	"r_xnm212",	"r_xnm2120",	"r_xnm2121",	"r_xnm2123",	"r_xnm2124",	"r_xnm2125",	"r_xnm2126",	"r_xnm2127",	"r_xnm2128",	"r_xnm2129",	"r_xnm213",	"r_xnm2131",	"r_xnm2133",	"r_xnm2134",	"r_xnm2135",	"r_xnm2136",	"r_xnm2137",	"r_xnm2138",	"r_xnm2139",	"r_xnm214",	"r_xnm2144",	"r_xnm2149",	"r_xnm215",	"r_xnm216",	"r_xnm217",	"r_xnm2170",	"r_xnm218",	"r_xnm2192",	"r_xnm2193",	"r_xnm2194",	"r_xnm2201",	"r_xnm2218",	"r_xnm2224",	"r_xnm2225",	"r_xnm2226",	"r_xnm2227",	"r_xnm2228",	"r_xnm2229",	"r_xnm2230",	"r_xnm2231",	"r_xnm2232",	"r_xnm2233",	"r_xnm2234",	"r_xnm2235",	"r_xnm2236",	"r_xnm2237",	"r_xnm2238",	"r_xnm2239",	"r_xnm224",	"r_xnm225",	"r_xnm226",	"r_xnm228",	"r_xnm229",	"r_xnm230",	"r_xnm231",	"r_xnm232",	"r_xnm233",	"r_xnm234",	"r_xnm235",	"r_xnm236",	"r_xnm237",	"r_xnm238",	"r_xnm242",	"r_xnm243",	"r_xnm244",	"r_xnm245",	"r_xnm246",	"r_xnm247",	"r_xnm248",	"r_xnm249",	"r_xnm251",	"r_xnm252",	"r_xnm2521",	"r_xnm2522",	"r_xnm2523",	"r_xnm2524",	"r_xnm2525",	"r_xnm2530",	"r_xnm2531",	"r_xnm2532",	"r_xnm2535",	"r_xnm2536",	"r_xnm2537",	"r_xnm2538",	"r_xnm2539",	"r_xnm2540",	"r_xnm2544",	"r_xnm2545",	"r_xnm2547",	"r_xnm2559",	"r_xnm258",	"r_xnm259",	"r_xnm2595",	"r_xnm2596",	"r_xnm2598",	"r_xnm26",	"r_xnm2603",	"r_xnm2604",	"r_xnm2605",	"r_xnm2606",	"r_xnm2607",	"r_xnm2608",	"r_xnm2610",	"r_xnm2611",	"r_xnm262",	"r_xnm2620",	"r_xnm2623",	"r_xnm2625",	"r_xnm263",	"r_xnm2631",	"r_xnm2632",	"r_xnm2634",	"r_xnm2639",	"r_xnm264",	"r_xnm2640",	"r_xnm2641",	"r_xnm2642",	"r_xnm2643",	"r_xnm2644",	"r_xnm2646",	"r_xnm2647",	"r_xnm265",	"r_xnm2656",	"r_xnm2659",	"r_xnm266",	"r_xnm2661",	"r_xnm267",	"r_xnm268",	"r_xnm269",	"r_xnm27",	"r_xnm270",	"r_xnm271",	"r_xnm272",	"r_xnm276",	"r_xnm277",	"r_xnm278",	"r_xnm279",	"r_xnm28",	"r_xnm280",	"r_xnm281",	"r_xnm282",	"r_xnm283",	"r_xnm285",	"r_xnm286",	"r_xnm29",	"r_xnm292",	"r_xnm293",	"r_xnm296",	"r_xnm297",	"r_xnm298",	"r_xnm299",	"r_xnm30",	"r_xnm300",	"r_xnm301",	"r_xnm302",	"r_xnm303",	"r_xnm304",	"r_xnm305",	"r_xnm306",	"r_xnm31",	"r_xnm310",	"r_xnm311",	"r_xnm312",	"r_xnm313",	"r_xnm314",	"r_xnm315",	"r_xnm316",	"r_xnm317",	"r_xnm32",	"r_xnm320",	"r_xnm3207",	"r_xnm3208",	"r_xnm3210",	"r_xnm3214",	"r_xnm3215",	"r_xnm3216",	"r_xnm3217",	"r_xnm3218",	"r_xnm3219",	"r_xnm3220",	"r_xnm3222",	"r_xnm3223",	"r_xnm3232",	"r_xnm3235",	"r_xnm3237",	"r_xnm326",	"r_xnm327",	"r_xnm33",	"r_xnm330",	"r_xnm331",	"r_xnm332",	"r_xnm333",	"r_xnm334",	"r_xnm335",	"r_xnm336",	"r_xnm337",	"r_xnm338",	"r_xnm339",	"r_xnm34",	"r_xnm340",	"r_xnm344",	"r_xnm345",	"r_xnm346",	"r_xnm347",	"r_xnm348",	"r_xnm349",	"r_xnm35",	"r_xnm350",	"r_xnm351",	"r_xnm36",	"r_xnm3604",	"r_xnm3606",	"r_xnm3611",	"r_xnm3612",	"r_xnm3613",	"r_xnm3614",	"r_xnm3615",	"r_xnm3616",	"r_xnm3618",	"r_xnm3619",	"r_xnm3628",	"r_xnm3631",	"r_xnm3632",	"r_xnm3633",	"r_xnm37",	"r_xnm38",	"r_xnm39",	"r_xnm40",	"r_xnm4000",	"r_xnm4002",	"r_xnm4007",	"r_xnm4008",	"r_xnm4009",	"r_xnm4011",	"r_xnm4012",	"r_xnm4014",	"r_xnm4015",	"r_xnm4024",	"r_xnm4027",	"r_xnm41",	"r_xnm42",	"r_xnm428",	"r_xnm429",	"r_xnm43",	"r_xnm432",	"r_xnm433",	"r_xnm434",	"r_xnm435",	"r_xnm436",	"r_xnm437",	"r_xnm438",	"r_xnm439",	"r_xnm4396",	"r_xnm4398",	"r_xnm44",	"r_xnm440",	"r_xnm4403",	"r_xnm4404",	"r_xnm4405",	"r_xnm4407",	"r_xnm4408",	"r_xnm441",	"r_xnm4410",	"r_xnm4411",	"r_xnm442",	"r_xnm4420",	"r_xnm4423",	"r_xnm443",	"r_xnm444",	"r_xnm445",	"r_xnm446",	"r_xnm447",	"r_xnm448",	"r_xnm449",	"r_xnm45",	"r_xnm450",	"r_xnm451",	"r_xnm452",	"r_xnm453",	"r_xnm454",	"r_xnm455",	"r_xnm456",	"r_xnm457",	"r_xnm458",	"r_xnm459",	"r_xnm46",	"r_xnm460",	"r_xnm461",	"r_xnm462",	"r_xnm463",	"r_xnm464",	"r_xnm465",	"r_xnm466",	"r_xnm467",	"r_xnm468",	"r_xnm469",	"r_xnm47",	"r_xnm470",	"r_xnm471",	"r_xnm472",	"r_xnm473",	"r_xnm474",	"r_xnm477",	"r_xnm478",	"r_xnm4792",	"r_xnm4794",	"r_xnm4799",	"r_xnm48",	"r_xnm4800",	"r_xnm4801",	"r_xnm4803",	"r_xnm4804",	"r_xnm4806",	"r_xnm4807",	"r_xnm4816",	"r_xnm4819",	"r_xnm483",	"r_xnm488",	"r_xnm489",	"r_xnm49",	"r_xnm490",	"r_xnm491",	"r_xnm492",	"r_xnm493",	"r_xnm494",	"r_xnm495",	"r_xnm496",	"r_xnm497",	"r_xnm498",	"r_xnm499",	"r_xnm50",	"r_xnm500",	"r_xnm501",	"r_xnm502",	"r_xnm503",	"r_xnm504",	"r_xnm505",	"r_xnm506",	"r_xnm507",	"r_xnm508",	"r_xnm509",	"r_xnm51",	"r_xnm510",	"r_xnm511",	"r_xnm512",	"r_xnm513",	"r_xnm514",	"r_xnm5188",	"r_xnm5190",	"r_xnm5195",	"r_xnm5196",	"r_xnm5197",	"r_xnm5199",	"r_xnm52",	"r_xnm5200",	"r_xnm5202",	"r_xnm5203",	"r_xnm5212",	"r_xnm5215",	"r_xnm5216",	"r_xnm53",	"r_xnm538",	"r_xnm54",	"r_xnm542",	"r_xnm55",	"r_xnm5586",	"r_xnm5591",	"r_xnm5592",	"r_xnm5593",	"r_xnm5595",	"r_xnm5596",	"r_xnm5598",	"r_xnm5599",	"r_xnm5608",	"r_xnm5611",	"r_xnm58",	"r_xnm59",	"r_xnm5980",	"r_xnm5982",	"r_xnm5991",	"r_xnm5992",	"r_xnm5994",	"r_xnm5995",	"r_xnm60",	"r_xnm6004",	"r_xnm6007",	"r_xnm6008",	"r_xnm61",	"r_xnm613",	"r_xnm614",	"r_xnm62",	"r_xnm63",	"r_xnm64",	"r_xnm640",	"r_xnm641",	"r_xnm642",	"r_xnm643",	"r_xnm644",	"r_xnm645",	"r_xnm6451",	"r_xnm646",	"r_xnm647",	"r_xnm648",	"r_xnm649",	"r_xnm65",	"r_xnm650",	"r_xnm651",	"r_xnm6512",	"r_xnm6515",	"r_xnm652",	"r_xnm653",	"r_xnm654",	"r_xnm6541",	"r_xnm6543",	"r_xnm6544",	"r_xnm6545",	"r_xnm6549",	"r_xnm655",	"r_xnm656",	"r_xnm657",	"r_xnm658",	"r_xnm659",	"r_xnm66",	"r_xnm669",	"r_xnm67",	"r_xnm670",	"r_xnm68",	"r_xnm69",	"r_xnm694",	"r_xnm696",	"r_xnm699",	"r_xnm70",	"r_xnm700",	"r_xnm701",	"r_xnm702",	"r_xnm703",	"r_xnm704",	"r_xnm705",	"r_xnm706",	"r_xnm707",	"r_xnm708",	"r_xnm709",	"r_xnm71",	"r_xnm710",	"r_xnm711",	"r_xnm712",	"r_xnm713",	"r_xnm714",	"r_xnm715",	"r_xnm716",	"r_xnm717",	"r_xnm718",	"r_xnm719",	"r_xnm72",	"r_xnm720",	"r_xnm721",	"r_xnm722",	"r_xnm723",	"r_xnm724",	"r_xnm725",	"r_xnm726",	"r_xnm73",	"r_xnm74",	"r_xnm75",	"r_xnm76",	"r_xnm77",	"r_xnm78",	"r_xnm79",	"r_xnm80",	"r_xnm81",	"r_xnm82",	"r_xnm821",	"r_xnm822",	"r_xnm823",	"r_xnm824",	"r_xnm825",	"r_xnm826",	"r_xnm827",	"r_xnm828",	"r_xnm829",	"r_xnm83",	"r_xnm830",	"r_xnm831",	"r_xnm832",	"r_xnm833",	"r_xnm834",	"r_xnm835",	"r_xnm836",	"r_xnm837",	"r_xnm838",	"r_xnm839",	"r_xnm84",	"r_xnm840",	"r_xnm841",	"r_xnm842",	"r_xnm843",	"r_xnm844",	"r_xnm845",	"r_xnm846",	"r_xnm847",	"r_xnm85",	"r_xnm855",	"r_xnm856",	"r_xnm857",	"r_xnm858",	"r_xnm859",	"r_xnm860",	"r_xnm861",	"r_xnm863",	"r_xnm865",	"r_xnm866",	"r_xnm867",	"r_xnm868",	"r_xnm869",	"r_xnm87",	"r_xnm870",	"r_xnm871",	"r_xnm872",	"r_xnm873",	"r_xnm874",	"r_xnm876",	"r_xnm877",	"r_xnm878",	"r_xnm879",	"r_xnm880",	"r_xnm881",	"r_xnm882",	"r_xnm883",	"r_xnm884",	"r_xnm885",	"r_xnm887",	"r_xnm888",	"r_xnm889",	"r_xnm89",	"r_xnm890",	"r_xnm891",	"r_xnm892",	"r_xnm893",	"r_xnm894",	"r_xnm895",	"r_xnm896",	"r_xnm897",	"r_xnm899",	"r_xnm90",	"r_xnm900",	"r_xnm901",	"r_xnm902",	"r_xnm903",	"r_xnm904",	"r_xnm905",	"r_xnm906",	"r_xnm907",	"r_xnm908",	"r_xnm909",	"r_xnm91",	"r_xnm911",	"r_xnm912",	"r_xnm913",	"r_xnm914",	"r_xnm915",	"r_xnm916",	"r_xnm917",	"r_xnm918",	"r_xnm919",	"r_xnm92",	"r_xnm920",	"r_xnm921",	"r_xnm923",	"r_xnm924",	"r_xnm925",	"r_xnm926",	"r_xnm927",	"r_xnm928",	"r_xnm929",	"r_xnm93",	"r_xnm930",	"r_xnm931",	"r_xnm932",	"r_xnm933",	"r_xnm935",	"r_xnm936",	"r_xnm937",	"r_xnm938",	"r_xnm939",	"r_xnm94",	"r_xnm940",	"r_xnm941",	"r_xnm942",	"r_xnm943",	"r_xnm944",	"r_xnm945",	"r_xnm947",	"r_xnm948",	"r_xnm949",	"r_xnm95",	"r_xnm950",	"r_xnm951",	"r_xnm952",	"r_xnm953",	"r_xnm954",	"r_xnm955",	"r_xnm956",	"r_xnm957",	"r_xnm959",	"r_xnm96",	"r_xnm960",	"r_xnm961",	"r_xnm962",	"r_xnm963",	"r_xnm964",	"r_xnm965",	"r_xnm966",	"r_xnm967",	"r_xnm968",	"r_xnm969",	"r_xnm97",	"r_xnm971",	"r_xnm972",	"r_xnm973",	"r_xnm974",	"r_xnm975",	"r_xnm976",	"r_xnm977",	"r_xnm978",	"r_xnm979",	"r_xnm98",	"r_xnm980",	"r_xnm982",	"r_xnm983",	"r_xnm984",	"r_xnm985",	"r_xnm986",	"r_xnm987",	"r_xnm988",	"r_xnm989",	"r_xnm99",	"r_xnm990",	"r_xnm991",	"r_xnm992",	"r_xnm994",	"r_xnm995",	"r_xnm996",	"r_xnm997",	"r_xnm998",	"r_xnm999",	"REV0300",	"REV1360",	"REV2320",	"REV3421",	"REV5020",	"REV5620",	"REV5627",	"REV8160",	"RiskScore",	"RTI5020",	"RTI5820",	"RTR3422",	"STU5820",	"vantage",	"VANTAGE_V3_SCORE",
]
c_m_list=["Carrier",	"IncomeSource",	"PaydayMethod",	"r_xnm1",	"r_xnm10",	"r_xnm1005",	"r_xnm1017",	"r_xnm1029",	"r_xnm1041",	"r_xnm1053",	"r_xnm1065",	"r_xnm1077",	"r_xnm1089",	"r_xnm1101",	"r_xnm1113",	"r_xnm1124",	"r_xnm1136",	"r_xnm1148",	"r_xnm1160",	"r_xnm1172",	"r_xnm1184",	"r_xnm1196",	"r_xnm1208",	"r_xnm1220",	"r_xnm1232",	"r_xnm1244",	"r_xnm1256",	"r_xnm1267",	"r_xnm1279",	"r_xnm1291",	"r_xnm1303",	"r_xnm1315",	"r_xnm1326",	"r_xnm1338",	"r_xnm1350",	"r_xnm1362",	"r_xnm1374",	"r_xnm1386",	"r_xnm1398",	"r_xnm1410",	"r_xnm1421",	"r_xnm1433",	"r_xnm1445",	"r_xnm1457",	"r_xnm1469",	"r_xnm1481",	"r_xnm1493",	"r_xnm1505",	"r_xnm1517",	"r_xnm1529",	"r_xnm1541",	"r_xnm1553",	"r_xnm1564",	"r_xnm1576",	"r_xnm1588",	"r_xnm1600",	"r_xnm1612",	"r_xnm1622",	"r_xnm1624",	"r_xnm1662",	"r_xnm1708",	"r_xnm1709",	"r_xnm1711",	"r_xnm1725",	"r_xnm1734",	"r_xnm1743",	"r_xnm1752",	"r_xnm1761",	"r_xnm1770",	"r_xnm1779",	"r_xnm1855",	"r_xnm1864",	"r_xnm1872",	"r_xnm1879",	"r_xnm1885",	"r_xnm1889",	"r_xnm1893",	"r_xnm1896",	"r_xnm1898",	"r_xnm1903",	"r_xnm1913",	"r_xnm1923",	"r_xnm1933",	"r_xnm1943",	"r_xnm1953",	"r_xnm1963",	"r_xnm1973",	"r_xnm1983",	"r_xnm2",	"r_xnm2002",	"r_xnm2075",	"r_xnm21",	"r_xnm2140",	"r_xnm2197",	"r_xnm2198",	"r_xnm2199",	"r_xnm22",	"r_xnm2200",	"r_xnm2207",	"r_xnm2208",	"r_xnm2209",	"r_xnm2210",	"r_xnm2217",	"r_xnm2242",	"r_xnm23",	"r_xnm24",	"r_xnm25",	"r_xnm2533",	"r_xnm2534",	"r_xnm421",	"r_xnm56",	"r_xnm6447",	"r_xnm6450",	"r_xnm6453",	"r_xnm6466",	"r_xnm6497",	"r_xnm6499",	"r_xnm6514",	"r_xnm6529",	"r_xnm6530",	"r_xnm6536",	"r_xnm6538",	"r_xnm6540",	"r_xnm6542",	"r_xnm6546",	"r_xnm6551",	"r_xnm6558",	"r_xnm852",	"r_xnm864",	"r_xnm875",	"r_xnm886",	"r_xnm898",	"r_xnm910",	"r_xnm922",	"r_xnm934",	"r_xnm946",	"r_xnm958",	"r_xnm970",	"r_xnm981",	"r_xnm993",	"UNIT_TYPE",
]

#num_specival_value_checking   
# n_m_r_list must be a list of numerical variable names
summary_stats = df[n_m_r_list].agg(['min', 'max', 'mean', 'count']).transpose().reset_index()
summary_stats.columns = ['variable', 'min', 'max', 'mean', 'n']


pre_attr=["ALL0100",	"ALL0200",	"ALL0216",	"ALL0317",	"ALL0336",	"ALL0400",	"ALL0416",	"ALL0436",	"ALL0446",	"ALL0448",	"ALL0700",	"ALL1380",	"ALL2000",	"ALL2001",	"ALL2002",	"ALL2005",	"ALL2106",	"ALL2136",	"ALL2176",	"ALL2326",	"ALL2327",	"ALL2357",	"ALL2380",	"ALL2387",	"ALL2427",	"ALL2428",	"ALL2480",	"ALL2707",	"ALL2870",	"ALL2907",	"ALL2937",	"ALL2967",	"ALL2990",	"ALL3517",	"ALL4080",	"ALL5047",	"ALL5070",	"ALL5743",	"ALL8026",	"ALL8164",	"ALL8167",	"ALL8325",	"ALX5839",	"AUA0416",	"AUA1300",	"AUA2358",	"AUA5820",	"AUA6160",	"BCA0416",	"BCA5030",	"BCA5070",	"BCA5740",	"BCA6220",	"BCC0436",	"BCC0446",	"BCC1360",	"BCC3342",	"BCC3345",	"BCC3423",	"BCC3512",	"BCC5020",	"BCC5320",	"BCC5421",	"BCC5620",	"BCC5627",	"BCC7117",	"BCC8120",	"BCX3421",	"BCX3422",	"BRC0416",	"BRC1300",	"BRC5320",	"BRC5830",	"BRC7140",	"COL3210",	"COL5060",	"COL8165",	"FIP0300",	"FIP2358",	"IQT9415",	"IQT9420",	"IQT9425",	"IQT9426",	"MTA6160",	"MTF2358",	"MTF8166",	"MTX5839",	"PIL5020",	"REV0300",	"REV1360",	"REV2320",	"REV3421",	"REV5020",	"REV5620",	"REV5627",	"REV8160",	"RTI5020",	"RTI5820",	"RTR3422",	"STU5820",
]

#[1]  xKS, IV***
# Step 1: Rename columns
temp_1a = df.rename(columns={ 'everdef': 'bad'})  #'LoanNumber': 'uniq_id',   already in data, but usually need to do this step
column_df = pd.DataFrame(df.columns.tolist(), columns=["column_name"])


# Step 2: Sort by uniq_id
temp_1a = temp_1a.sort_values(by=['uniq_id'])

temp_2 = pd.melt(
    temp_1a,
    id_vars=['uniq_id', 'bad', 'weight'],
    value_vars=n_m_r_list,   
    var_name='x_nm',
    value_name='x_value'
)

#function
def DR_x_weighted_KS_pandas(in_data_df, bad, weight, out_name='x_ks_result'):
    df = in_data_df.copy()
    
    # Step 1: Group by x_nm and x_value
    grouped = df.groupby(['x_nm', 'x_value']).apply(
        lambda g: pd.Series({
            'bad_wgt': g.loc[g[bad] == 1, weight].sum(),
            'good_wgt': g.loc[g[bad] == 0, weight].sum()
        })
    ).reset_index()
    
    # Steps 2-5 (same as before)
    grouped['cum_bad_wgt'] = grouped.groupby('x_nm')['bad_wgt'].cumsum()
    grouped['cum_good_wgt'] = grouped.groupby('x_nm')['good_wgt'].cumsum()
    
    totals = grouped.groupby('x_nm')[['bad_wgt', 'good_wgt']].sum().reset_index()
    totals.columns = ['x_nm', 'tot_bad_wgt', 'tot_good_wgt']
    merged = grouped.merge(totals, on='x_nm', how='left')
    
    merged['bad_cdf'] = merged['cum_bad_wgt'] / merged['tot_bad_wgt']
    merged['good_cdf'] = merged['cum_good_wgt'] / merged['tot_good_wgt']
    merged['ks_diff'] = (merged['bad_cdf'] - merged['good_cdf']).abs()
    
    out_x_ks = merged.groupby('x_nm')['ks_diff'].max().reset_index()
    out_x_ks['xKS'] = 100 * out_x_ks['ks_diff']
    out_x_ks = out_x_ks[['x_nm', 'xKS']].sort_values('xKS', ascending=False)
    
    globals()[out_name] = out_x_ks
    return out_x_ks     
       
# call the function as above   temp_2 is your long-format DataFrame with columns: ['x_nm', 'x_value', 'bad', 'weight'] 
x_ks_df = DR_x_weighted_KS_pandas(temp_2, bad='bad', weight='weight', out_name='x_ks_df')
print(x_ks_df)


#pulling  xKS>0** x_nm
temp_3 = temp_2.merge(
    x_ks_df[x_ks_df['xKS'] > 0],  # Filter rows where xKS > 0
    on='x_nm',                       # Join on x_nm column
    how='inner'                      # Inner join (like SAS default)
)
temp_3 = temp_3.sort_values(by=temp_3.columns[1], ascending=False)

#autobinning  function
import duckdb
def DR_num_rbin_duckdb(
    df_input: pd.DataFrame,
    uniq_id: str,
    bad: str,
    weight: str,
    max_bin: int,
    min_adj_bin: int
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Equivalent of SAS macro %wh_num_rbin using duckdb in Python.
    Returns:
        out_data_rbin_sql: original dataset with 'rbin' column.
        out_num_rbins: bin statistics per variable.
    """
    # Register input table
    duckdb.register("df_input", df_input)

    # Step 1: flag missing
    wh_rbin_0 = duckdb.query("""
        SELECT *, 
               CASE WHEN x_value IS NULL THEN 1.0 ELSE 0.0 END AS miss_yn
        FROM df_input
    """).to_df()
    duckdb.register("wh_rbin_0", wh_rbin_0)

    # Step 2: summarize missing and determine bin count
    wh_rbin_1 = duckdb.query(f"""
        SELECT 
            x_nm,
            SUM({weight} * miss_yn) / SUM({weight}) AS miss_rate,
            SUM({weight}) AS tot_wgt,
            SUM((1.0 - miss_yn) * {weight}) AS tot_nomiss_wgt,
            CEIL({max_bin} * (1.000 - SUM({weight} * miss_yn) / SUM({weight}))) AS adj_nomiss_bin,
            CASE 
                WHEN {min_adj_bin} > CEIL({max_bin} * (1.000 - SUM({weight} * miss_yn) / SUM({weight}))) 
                THEN {min_adj_bin}
                ELSE CEIL({max_bin} * (1.000 - SUM({weight} * miss_yn) / SUM({weight})))
            END AS nomiss_bin
        FROM wh_rbin_0
        GROUP BY x_nm
    """).to_df()
    duckdb.register("wh_rbin_1", wh_rbin_1)

    # Step 3: group by x_value, join with summary
    wh_rbin_2 = duckdb.query(f"""
        SELECT a.*, b.tot_nomiss_wgt, b.nomiss_bin
        FROM (
            SELECT x_nm, miss_yn, x_value, SUM({weight}) AS x_wgt
            FROM wh_rbin_0
            GROUP BY x_nm, miss_yn, x_value
        ) a
        LEFT JOIN wh_rbin_1 b 
        ON a.x_nm = b.x_nm
        ORDER BY a.x_nm, miss_yn, x_value
    """).to_df()

    # Step 4: cumulative binning logic (in pandas)
    wh_rbin_3a = wh_rbin_2.copy()
    wh_rbin_3a['cum_nomiss_wgt'] = wh_rbin_3a.groupby(['x_nm', 'miss_yn'])['x_wgt'].cumsum()
    wh_rbin_3a['nomiss_bin_w'] = (wh_rbin_3a['tot_nomiss_wgt'] / wh_rbin_3a['nomiss_bin']).round().fillna(1.0)
    wh_rbin_3a['bin_id'] = wh_rbin_3a.apply(
        lambda row: row['x_value'] if row['miss_yn'] == 1 else 1 + int(row['cum_nomiss_wgt'] / (1.0 + row['nomiss_bin_w'])),
        axis=1
    )
    duckdb.register("wh_rbin_3a", wh_rbin_3a)

    # Step 5: count by bin_id
    wh_rbin_3b = duckdb.query("""
        SELECT x_nm, bin_id, COUNT(*) AS tempwh_cnt
        FROM wh_rbin_3a
        WHERE miss_yn = 0
        GROUP BY x_nm, bin_id
        ORDER BY x_nm, bin_id
    """).to_df()

    # Step 6: assign rbin number
    wh_rbin_3b['rbin'] = wh_rbin_3b.groupby('x_nm').cumcount() + 1
    duckdb.register("wh_rbin_3b", wh_rbin_3b)

    # Step 7: final bin assignment
    wh_rbin_3 = duckdb.query("""
        SELECT a.*, 
               CASE WHEN a.miss_yn = 1 THEN a.bin_id ELSE b.rbin END AS rbin
        FROM wh_rbin_3a a
        LEFT JOIN wh_rbin_3b b 
        ON a.x_nm = b.x_nm AND a.bin_id = b.bin_id
    """).to_df()
    duckdb.register("wh_rbin_3", wh_rbin_3)

    # Step 8: join back to original
    out_data_rbin_sql = duckdb.query("""
        SELECT a.*, b.rbin
        FROM df_input a
        LEFT JOIN wh_rbin_3 b 
        ON a.x_nm = b.x_nm AND a.x_value = b.x_value
        ORDER BY a.x_nm, a.x_value
    """).to_df()
    
    duckdb.register("out_data_rbin_sql", out_data_rbin_sql)
    # Step 9: compute final bin stats
    out_num_rbins = duckdb.query(f"""
        SELECT h.*, h.rbin_wgt / g.tot_wgt AS rbin_dist_pct
        FROM (
            SELECT x_nm, rbin, COUNT(*) AS obs_cnt,
                   MIN(x_value) AS min_x,
                   MAX(x_value) AS max_x,
                   SUM(x_value * {weight}) / SUM({weight}) AS mean_x,
                   SUM({weight}) AS rbin_wgt,
                   SUM({weight} * {bad}) AS rbin_bad_wgt,
                   SUM(1.0 * {weight} * {bad}) / SUM({weight}) AS rbin_bad_pct
            FROM out_data_rbin_sql
            GROUP BY x_nm, rbin
        ) h
        LEFT JOIN wh_rbin_1 g
        ON h.x_nm = g.x_nm
        ORDER BY h.x_nm, h.rbin
    """).to_df()

    return out_data_rbin_sql, out_num_rbins

# call the function
out_data_rbin_sql, out_num_rbins = DR_num_rbin_duckdb(
    df_input=temp_3,  # including x_nm, x_value, weight, bad etc. columns
    uniq_id='loan_id',
    bad='bad',
    weight='weight',
    max_bin=5,
    min_adj_bin=5
)


#information value calculation**
import statsmodels.api as sm
from scipy.stats import linregress

def DR_num_iv_woe_duckdb(df_rbin: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    import duckdb
    import numpy as np
    import pandas as pd
    import statsmodels.api as sm
    from scipy.stats import linregress

    duckdb.register("out_data_rbin_sql", df_rbin)

    # Step 1: bin-level WOE/IV calculation
    tempwh_iv_1 = duckdb.query("""
        SELECT h.*, 
               g.tot_wgt,
               h.sum_wgt / g.tot_wgt AS dist_pct,
               g.tot_miss_wgt / g.tot_wgt AS tot_miss_rate,
               g.tot_bad_rate,
               CASE WHEN h.bad_rate = 0.0 THEN 0.5 / h.sum_wgt
                    WHEN h.bad_rate = 1.0 THEN ((h.sum_wgt - 1) - 0.5) / h.sum_wgt
                    ELSE h.bad_rate
               END AS cal_bad_rate,
               CASE WHEN h.bad_rate = 0.0 THEN 0.5 ELSE h.sum_bad_wgt END / g.tot_bad_wgt AS dist_bad_pct,
               CASE WHEN h.bad_rate = 1.0 THEN 0.5 ELSE h.sum_good_wgt END / g.tot_good_wgt AS dist_good_pct,
               LOG(cal_bad_rate / (1.0 - cal_bad_rate)) AS logit_bad_rate,
               LOG(dist_good_pct / dist_bad_pct) AS woe,
               (dist_good_pct - dist_bad_pct) * LOG(dist_good_pct / dist_bad_pct) AS bin_iv
        FROM (
            SELECT x_nm, rbin,
                   MIN(x_value) AS min_x_value,
                   MAX(x_value) AS max_x_value,
                   SUM(weight * x_value) / SUM(weight) AS avg_x_value,
                   SUM(weight) AS sum_wgt,
                   SUM(weight * bad) AS sum_bad_wgt,
                   SUM(weight * (1 - bad)) AS sum_good_wgt,
                   SUM(weight * bad) / SUM(weight) AS bad_rate
            FROM out_data_rbin_sql
            GROUP BY x_nm, rbin
        ) h
        LEFT JOIN (
            SELECT x_nm,
                   SUM(weight) AS tot_wgt,
                   SUM(weight * bad) AS tot_bad_wgt,
                   SUM(weight * (1.0 - bad)) AS tot_good_wgt,
                   SUM(CASE WHEN x_value IS NULL THEN weight ELSE 0.0 END) AS tot_miss_wgt,
                   SUM(weight * bad) / SUM(weight) AS tot_bad_rate
            FROM out_data_rbin_sql
            GROUP BY x_nm
        ) g
        ON h.x_nm = g.x_nm
        ORDER BY h.x_nm, h.rbin
    """).to_df()

    # Step 2: summary statistics
    desc_stats = df_rbin.groupby("x_nm")["x_value"].agg(
        mean='mean', min='min', max='max',
        p1=lambda x: np.percentile(x.dropna(), 1),
        p99=lambda x: np.percentile(x.dropna(), 99)
    ).reset_index()

    # Step 3: logistic regression
    logit_results = []
    for x_nm, group in df_rbin.groupby("x_nm"):
        x_raw = group["x_value"]
        y = group["bad"]
        w = group["weight"]

        if x_raw.isnull().all() or x_raw.nunique() < 2 or y.nunique() < 2:
            est, pval = np.nan, np.nan
        else:
            try:
                # standardiza x_value
                x_scaled = (x_raw - x_raw.mean()) / (x_raw.std(ddof=0) + 1e-6)
                X = sm.add_constant(x_scaled, has_constant='add')
                model = sm.GLM(y, X, family=sm.families.Binomial(), freq_weights=w)
                result = model.fit(disp=0)
                est = result.params.get("x_value", np.nan)
                pval = result.pvalues.get("x_value", np.nan)
            except Exception:
                est, pval = np.nan, np.nan

        logit_results.append({"x_nm": x_nm, "estimate": est, "probchisq": pval})
    df_logit = pd.DataFrame(logit_results)

    # Step 4: Cumulative distribution and KS calculation
    df_woe_bin = tempwh_iv_1.copy()
    df_woe_bin["cum_dist_pct"] = df_woe_bin.groupby("x_nm")["dist_pct"].cumsum()
    df_woe_bin["cum_dist_bad_pct"] = df_woe_bin.groupby("x_nm")["dist_bad_pct"].cumsum()
    df_woe_bin["cum_dist_good_pct"] = df_woe_bin.groupby("x_nm")["dist_good_pct"].cumsum()
    df_woe_bin["diff_cum_bad_good_pct"] = df_woe_bin["cum_dist_bad_pct"] - df_woe_bin["cum_dist_good_pct"]
    duckdb.register("tempwh_iv_1", df_woe_bin)

    # Step 5: logit vs avg_x_value（PROC REG replace）
    reg_results = []
    for x_nm, group in df_woe_bin[df_woe_bin["rbin"].notnull()].groupby("x_nm"):
        x = group["avg_x_value"]
        y = group["logit_bad_rate"]

        if x.isnull().all() or x.nunique() <= 1 or y.nunique() <= 1:
            slope, r_square = np.nan, np.nan
        else:
            try:
                slope, _, r_value, _, _ = linregress(x, y)
                r_square = r_value**2
            except Exception:
                slope, r_square = np.nan, np.nan

        reg_results.append({"x_nm": x_nm, "Beta": slope, "R_Square": r_square})
    df_reg = pd.DataFrame(reg_results)

    # Step 6: Bin count statistics
    df_nomiss_rbin = df_rbin[df_rbin["x_value"].notnull()].groupby("x_nm")["rbin"].nunique().reset_index()
    df_nomiss_rbin.columns = ["x_nm", "nomiss_rbin_cnt"]

    # Step 7: Summary output
    df_iv_summary = duckdb.query("""
        SELECT 
            h.x_nm,
            h.tot_miss_rate AS miss_rate,
            h.rbin_cnt,
            p.nomiss_rbin_cnt,
            h.iv,
            h.ks,
            w.R_Square,
            f.estimate AS Beta,
            f.probchisq,
            g.mean, g.min, g.max, g.p1, g.p99,
            h.tot_wgt
        FROM (
            SELECT x_nm, tot_wgt, tot_miss_rate,
                   COUNT(*) AS rbin_cnt,
                   SUM(bin_iv) AS iv,
                   MAX(ABS(diff_cum_bad_good_pct)) AS ks
            FROM tempwh_iv_1
            GROUP BY x_nm, tot_wgt, tot_miss_rate
        ) h
        LEFT JOIN desc_stats g ON h.x_nm = g.x_nm
        LEFT JOIN df_logit f ON h.x_nm = f.x_nm
        LEFT JOIN df_reg w ON h.x_nm = w.x_nm
        LEFT JOIN df_nomiss_rbin p ON h.x_nm = p.x_nm
        ORDER BY h.x_nm
    """).to_df()

    return df_woe_bin, df_iv_summary
#call the function
df_woe, df_iv = DR_num_iv_woe_duckdb(out_data_rbin_sql)

#######Candidate variables 
temp_4 = duckdb.query(f"""
    SELECT * from x_ks_df a inner join df_iv b on a.x_nm=b.x_nm order by a.xKS DESC
""").to_df()
duckdb.register("temp_4", temp_4)

temp_sel = temp_4[
    ((temp_4['rbin_cnt'] >= 4) & (temp_4['iv'] >= 0.005)) |
    ((temp_4['rbin_cnt'] < 4) & (temp_4['xKS'] >= 4))
].copy()


#***************************[ Auto: MFC/F/C  *************************
temp_5a=temp_sel[(temp_sel['nomiss_rbin_cnt']>=2)]

temp_5=duckdb.query(f"""
                    select a.* from out_data_rbin_sql a inner join temp_5a b on a.x_nm=b.x_nm
                    """).to_df()

temp_6= duckdb.query(f"""
                     select a.* from df_iv a inner join temp_5a b on a.x_nm=b.x_nm
                     """).to_df()
                     
                     
temp_7=duckdb.query(f"""
                    select a.* from df_woe a inner join temp_5a b on a.x_nm=b.x_nm
                    """).to_df()

def DR_num_aMFC(in_data_rbin, in_num_iv, in_num_woe, uniq_id, bad, weight, 
                out_mfc_value_name='out_mfc_value', out_aMFC_SAS_name='out_aMFC_SAS'):
    
    # tempwh_whMFC_1
    tempwh_whMFC_1 = duckdb.query(f"""
        SELECT a.x_nm, a.rbin, a.avg_x_value, a.bad_rate
        FROM in_num_woe a, in_num_iv b
        WHERE a.x_nm = b.x_nm
        ORDER BY a.x_nm, a.rbin
    """).to_df()

    # tempwh_whMFC_2
    tempwh_whMFC_2 = duckdb.query(f"""
        SELECT 
            a.x_nm, a.rbin, a.bad_rate, 
            b.rbin AS b_rbin, 
            b.avg_x_value AS b_avg_x_value, 
            b.bad_rate AS b_bad_rate,
            ABS(a.bad_rate - b.bad_rate) AS abd_bad_diff
        FROM tempwh_whMFC_1 a
        CROSS JOIN tempwh_whMFC_1 b
        WHERE a.x_nm = b.x_nm 
          AND a.rbin IS  NULL 
          AND b.rbin IS NOT NULL
        ORDER BY a.x_nm, a.rbin, abd_bad_diff
    """).to_df()

    # tempwh_whMFC_3 (drop duplicates by x_nm and rbin)
    tempwh_whMFC_3 = tempwh_whMFC_2.drop_duplicates(subset=["x_nm", "rbin"])

   # Step 4: tempwh_whMFC_4
    merged = duckdb.query(f"""
        SELECT 
            a.*, 
            b.b_avg_x_value,
            CASE 
                WHEN a.rbin IS NOT NULL THEN a.x_value
                WHEN b.b_avg_x_value IS NOT NULL THEN b.b_avg_x_value
                ELSE a.mean
            END AS tempwh_x_value,
            CASE 
                WHEN (CASE 
                        WHEN a.rbin IS NOT NULL THEN a.x_value
                        WHEN b.b_avg_x_value IS NOT NULL THEN b.b_avg_x_value
                        ELSE a.mean
                      END) > a.p99 THEN a.p99
                WHEN (CASE 
                        WHEN a.rbin IS NOT NULL THEN a.x_value
                        WHEN b.b_avg_x_value IS NOT NULL THEN b.b_avg_x_value
                        ELSE a.mean
                      END) < a.p1 THEN a.p1
                ELSE (CASE 
                        WHEN a.rbin IS NOT NULL THEN a.x_value
                        WHEN b.b_avg_x_value IS NOT NULL THEN b.b_avg_x_value
                        ELSE a.mean
                      END)
            END AS mfc_x_value
        FROM (
            SELECT h.*, g.mean, g.p1, g.p99
            FROM in_data_rbin h
            JOIN in_num_iv g ON h.x_nm = g.x_nm
        ) a
        LEFT JOIN tempwh_whMFC_3 b
        ON a.x_nm = b.x_nm AND a.rbin = b.rbin
        ORDER BY {uniq_id}, a.x_nm, {bad}, {weight}
    """).to_df()

    # PROC TRANSPOSE: wide format
    temp_pivot = merged.pivot_table(index=[uniq_id, bad, weight],
                                    columns='x_nm',
                                    values='mfc_x_value',
                                    aggfunc='first').reset_index()

    # output
    out_mfc_value = tempwh_whMFC_3.copy()
    out_aMFC_SAS = temp_pivot.copy()

    
    globals()[out_mfc_value_name] = out_mfc_value
    globals()[out_aMFC_SAS_name] = out_aMFC_SAS

    return out_mfc_value, out_aMFC_SAS

#call the function as above
duckdb.register("in_data_rbin", temp_5)
duckdb.register("in_num_iv", temp_6)
duckdb.register("in_num_woe", temp_7)

out_val, out_sas = DR_num_aMFC("in_data_rbin", "in_num_iv", "in_num_woe", 
                               uniq_id="uniq_id", bad="bad", weight="weight")




###***************************[3] Auto: woe *************************
temp_8 =temp_sel[(temp_sel['nomiss_rbin_cnt']<2)] 

temp_9=duckdb.query(f"""
                    select a.* from out_data_rbin_sql a inner join temp_8 b on a.x_nm=b.x_nm
                    """).to_df()
 
temp_10= duckdb.query(f"""
                      select a.*, b.woe from temp_9 a left join df_woe b on a.x_nm=b.x_nm and (a.rbin = b.rbin OR (a.rbin IS NULL AND b.rbin IS NULL))
 order by a.uniq_id, a.x_nm
                      """).to_df()
                      


def transpose_woe(temp_10: pd.DataFrame) -> pd.DataFrame:
    out_woe_data_sas = temp_10.pivot_table(
        index=['uniq_id', 'bad', 'weight'],  
        columns='x_nm',                      
        values='woe',                        
        aggfunc='first'                      
    ).reset_index()

   
    out_woe_data_sas.columns.name = None  
    return out_woe_data_sas
#call the function
out_df = transpose_woe(temp_10)


out_IV_data_sas=duckdb.query(f"""
                             select  * from out_sas a inner join out_df b on a.uniq_id=b.uniq_id and a.bad=b.bad
                             """).to_df()

                            
########Part 2########Part 2########Part 2                         
pre_attr=["ALL0100",	"ALL0200",	"ALL0216",	"ALL0317",	"ALL0336",	"ALL0400",	"ALL0416",	"ALL0436",	"ALL0446",	"ALL0448",	"ALL0700",	"ALL1380",	"ALL2000",	"ALL2001",	"ALL2002",	"ALL2005",	"ALL2106",	"ALL2136",	"ALL2176",	"ALL2326",	"ALL2327",	"ALL2357",	"ALL2380",	"ALL2387",	"ALL2427",	"ALL2428",	"ALL2480",	"ALL2707",	"ALL2870",	"ALL2907",	"ALL2937",	"ALL2967",	"ALL2990",	"ALL3517",	"ALL4080",	"ALL5047",	"ALL5070",	"ALL5743",	"ALL8026",	"ALL8164",	"ALL8167",	"ALL8325",	"ALX5839",	"AUA0416",	"AUA1300",	"AUA2358",	"AUA5820",	"AUA6160",	"BCA0416",	"BCA5030",	"BCA5070",	"BCA5740",	"BCA6220",	"BCC0436",	"BCC0446",	"BCC1360",	"BCC3342",	"BCC3345",	"BCC3423",	"BCC3512",	"BCC5020",	"BCC5320",	"BCC5421",	"BCC5620",	"BCC5627",	"BCC7117",	"BCC8120",	"BCX3421",	"BCX3422",	"BRC0416",	"BRC1300",	"BRC5320",	"BRC5830",	"BRC7140",	"COL3210",	"COL5060",	"COL8165",	"FIP0300",	"FIP2358",	"IQT9415",	"IQT9420",	"IQT9425",	"IQT9426",	"MTA6160",	"MTF2358",	"MTF8166",	"MTX5839",	"PIL5020",	"REV0300",	"REV1360",	"REV2320",	"REV3421",	"REV5020",	"REV5620",	"REV5627",	"REV8160",	"RTI5020",	"RTI5820",	"RTR3422",	"STU5820",
]

#character model attr candidates
c_m_r_list=["r_xnm1",	"r_xnm10",	"r_xnm1005",	"r_xnm1017",	"r_xnm1029",	"r_xnm1041",	"r_xnm1053",	"r_xnm1065",	"r_xnm1077",	"r_xnm1089",	"r_xnm1101",	"r_xnm1113",	"r_xnm1124",	"r_xnm1136",	"r_xnm1148",	"r_xnm1160",	"r_xnm1172",	"r_xnm1184",	"r_xnm1196",	"r_xnm1208",	"r_xnm1220",	"r_xnm1232",	"r_xnm1244",	"r_xnm1256",	"r_xnm1267",	"r_xnm1279",	"r_xnm1291",	"r_xnm1303",	"r_xnm1315",	"r_xnm1326",	"r_xnm1338",	"r_xnm1350",	"r_xnm1362",	"r_xnm1374",	"r_xnm1386",	"r_xnm1398",	"r_xnm1410",	"r_xnm1421",	"r_xnm1433",	"r_xnm1445",	"r_xnm1457",	"r_xnm1469",	"r_xnm1481",	"r_xnm1493",	"r_xnm1505",	"r_xnm1517",	"r_xnm1529",	"r_xnm1541",	"r_xnm1553",	"r_xnm1564",	"r_xnm1576",	"r_xnm1588",	"r_xnm1600",	"r_xnm1612",	"r_xnm1622",	"r_xnm1624",	"r_xnm1662",	"r_xnm1708",	"r_xnm1709",	"r_xnm1711",	"r_xnm1725",	"r_xnm1734",	"r_xnm1743",	"r_xnm1752",	"r_xnm1761",	"r_xnm1770",	"r_xnm1779",	"r_xnm1855",	"r_xnm1864",	"r_xnm1872",	"r_xnm1879",	"r_xnm1885",	"r_xnm1889",	"r_xnm1893",	"r_xnm1896",	"r_xnm1898",	"r_xnm1903",	"r_xnm1913",	"r_xnm1923",	"r_xnm1933",	"r_xnm1943",	"r_xnm1953",	"r_xnm1963",	"r_xnm1973",	"r_xnm1983",	"r_xnm2",	"r_xnm2002",	"r_xnm2075",	"r_xnm21",	"r_xnm2140",	"r_xnm2197",	"r_xnm2198",	"r_xnm2199",	"r_xnm22",	"r_xnm2200",	"r_xnm2207",	"r_xnm2208",	"r_xnm2209",	"r_xnm2210",	"r_xnm2217",	"r_xnm2242",	"r_xnm23",	"r_xnm24",	"r_xnm25",	"r_xnm2533",	"r_xnm2534",	"r_xnm421",	"r_xnm56",	"r_xnm6447",	"r_xnm6450",	"r_xnm6453",	"r_xnm6466",	"r_xnm6497",	"r_xnm6499",	"r_xnm6514",	"r_xnm6529",	"r_xnm6530",	"r_xnm6536",	"r_xnm6538",	"r_xnm6540",	"r_xnm6542",	"r_xnm6546",	"r_xnm6551",	"r_xnm6558",	"r_xnm852",	"r_xnm864",	"r_xnm875",	"r_xnm886",	"r_xnm898",	"r_xnm910",	"r_xnm922",	"r_xnm934",	"r_xnm946",	"r_xnm958",	"r_xnm970",	"r_xnm981",	"r_xnm993",	"Carrier",	"IncomeSource",	"PaydayMethod",	"UNIT_TYPE",
]
#numeric model attr candidates
n_m_r_list=["ALL0100",	"ALL0200",	"ALL0216",	"ALL0317",	"ALL0336",	"ALL0400",	"ALL0416",	"ALL0436",	"ALL0446",	"ALL0448",	"ALL0700",	"ALL1380",	"ALL2000",	"ALL2001",	"ALL2002",	"ALL2005",	"ALL2106",	"ALL2136",	"ALL2176",	"ALL2326",	"ALL2327",	"ALL2357",	"ALL2380",	"ALL2387",	"ALL2427",	"ALL2428",	"ALL2480",	"ALL2707",	"ALL2870",	"ALL2907",	"ALL2937",	"ALL2967",	"ALL2990",	"ALL3517",	"ALL4080",	"ALL5047",	"ALL5070",	"ALL5743",	"ALL8026",	"ALL8164",	"ALL8167",	"ALL8325",	"ALX5839",	"AUA0416",	"AUA1300",	"AUA2358",	"AUA5820",	"AUA6160",	"BCA0416",	"BCA5030",	"BCA5070",	"BCA5740",	"BCA6220",	"BCC0436",	"BCC0446",	"BCC1360",	"BCC3342",	"BCC3345",	"BCC3423",	"BCC3512",	"BCC5020",	"BCC5320",	"BCC5421",	"BCC5620",	"BCC5627",	"BCC7117",	"BCC8120",	"BCX3421",	"BCX3422",	"BRC0416",	"BRC1300",	"BRC5320",	"BRC5830",	"BRC7140",	"COL3210",	"COL5060",	"COL8165",	"FIP0300",	"FIP2358",	"IQT9415",	"IQT9420",	"IQT9425",	"IQT9426",	"International",	"MTA6160",	"MTF2358",	"MTF8166",	"MTX5839",	"PIL5020",	"REV0300",	"REV1360",	"REV2320",	"REV3421",	"REV5020",	"REV5620",	"REV5627",	"REV8160",	"RTI5020",	"RTI5820",	"RTR3422",	"RiskScore",	"STU5820",	"VANTAGE_V3_SCORE",	"anual_income",	"r_xnm100",	"r_xnm1000",	"r_xnm1001",	"r_xnm1002",	"r_xnm1003",	"r_xnm1004",	"r_xnm1006",	"r_xnm1007",	"r_xnm1008",	"r_xnm1009",	"r_xnm101",	"r_xnm1010",	"r_xnm1011",	"r_xnm1012",	"r_xnm1013",	"r_xnm1014",	"r_xnm1015",	"r_xnm1016",	"r_xnm1018",	"r_xnm1019",	"r_xnm102",	"r_xnm1020",	"r_xnm1021",	"r_xnm1022",	"r_xnm1023",	"r_xnm1024",	"r_xnm1025",	"r_xnm1026",	"r_xnm1027",	"r_xnm1028",	"r_xnm103",	"r_xnm1030",	"r_xnm1031",	"r_xnm1032",	"r_xnm1033",	"r_xnm1034",	"r_xnm1035",	"r_xnm1036",	"r_xnm1037",	"r_xnm1038",	"r_xnm1039",	"r_xnm104",	"r_xnm1040",	"r_xnm1042",	"r_xnm1043",	"r_xnm1044",	"r_xnm1045",	"r_xnm1046",	"r_xnm1047",	"r_xnm1048",	"r_xnm1049",	"r_xnm105",	"r_xnm1050",	"r_xnm1051",	"r_xnm1052",	"r_xnm1054",	"r_xnm1055",	"r_xnm1056",	"r_xnm1057",	"r_xnm1058",	"r_xnm1059",	"r_xnm106",	"r_xnm1060",	"r_xnm1061",	"r_xnm1062",	"r_xnm1063",	"r_xnm1064",	"r_xnm1066",	"r_xnm1067",	"r_xnm1068",	"r_xnm1069",	"r_xnm107",	"r_xnm1070",	"r_xnm1071",	"r_xnm1072",	"r_xnm1073",	"r_xnm1074",	"r_xnm1075",	"r_xnm1076",	"r_xnm1078",	"r_xnm1079",	"r_xnm108",	"r_xnm1080",	"r_xnm1081",	"r_xnm1082",	"r_xnm1083",	"r_xnm1084",	"r_xnm1085",	"r_xnm1086",	"r_xnm1087",	"r_xnm1088",	"r_xnm109",	"r_xnm1090",	"r_xnm1091",	"r_xnm1092",	"r_xnm1093",	"r_xnm1094",	"r_xnm1095",	"r_xnm1096",	"r_xnm1097",	"r_xnm1098",	"r_xnm1099",	"r_xnm110",	"r_xnm1100",	"r_xnm1102",	"r_xnm1103",	"r_xnm1104",	"r_xnm1105",	"r_xnm1106",	"r_xnm1107",	"r_xnm1108",	"r_xnm1109",	"r_xnm111",	"r_xnm1110",	"r_xnm1111",	"r_xnm1112",	"r_xnm1114",	"r_xnm1115",	"r_xnm1116",	"r_xnm1117",	"r_xnm1118",	"r_xnm1119",	"r_xnm112",	"r_xnm1120",	"r_xnm1121",	"r_xnm1122",	"r_xnm1123",	"r_xnm1125",	"r_xnm1126",	"r_xnm1127",	"r_xnm1128",	"r_xnm1129",	"r_xnm113",	"r_xnm1130",	"r_xnm1131",	"r_xnm1132",	"r_xnm1133",	"r_xnm1134",	"r_xnm1135",	"r_xnm1137",	"r_xnm1138",	"r_xnm1139",	"r_xnm114",	"r_xnm1140",	"r_xnm1141",	"r_xnm1142",	"r_xnm1143",	"r_xnm1144",	"r_xnm1145",	"r_xnm1146",	"r_xnm1147",	"r_xnm1149",	"r_xnm115",	"r_xnm1150",	"r_xnm1151",	"r_xnm1152",	"r_xnm1153",	"r_xnm1154",	"r_xnm1155",	"r_xnm1156",	"r_xnm1157",	"r_xnm1158",	"r_xnm1159",	"r_xnm116",	"r_xnm1161",	"r_xnm1162",	"r_xnm1163",	"r_xnm1164",	"r_xnm1165",	"r_xnm1166",	"r_xnm1167",	"r_xnm1168",	"r_xnm1169",	"r_xnm117",	"r_xnm1170",	"r_xnm1171",	"r_xnm1173",	"r_xnm1174",	"r_xnm1175",	"r_xnm1176",	"r_xnm1177",	"r_xnm1178",	"r_xnm1179",	"r_xnm118",	"r_xnm1180",	"r_xnm1181",	"r_xnm1182",	"r_xnm1183",	"r_xnm1185",	"r_xnm1186",	"r_xnm1187",	"r_xnm1188",	"r_xnm1189",	"r_xnm119",	"r_xnm1190",	"r_xnm1191",	"r_xnm1192",	"r_xnm1193",	"r_xnm1194",	"r_xnm1195",	"r_xnm1197",	"r_xnm1198",	"r_xnm1199",	"r_xnm120",	"r_xnm1200",	"r_xnm1201",	"r_xnm1202",	"r_xnm1203",	"r_xnm1204",	"r_xnm1205",	"r_xnm1206",	"r_xnm1207",	"r_xnm1209",	"r_xnm121",	"r_xnm1210",	"r_xnm1211",	"r_xnm1212",	"r_xnm1213",	"r_xnm1214",	"r_xnm1215",	"r_xnm1216",	"r_xnm1217",	"r_xnm1218",	"r_xnm1219",	"r_xnm122",	"r_xnm1221",	"r_xnm1222",	"r_xnm1223",	"r_xnm1224",	"r_xnm1225",	"r_xnm1226",	"r_xnm1227",	"r_xnm1228",	"r_xnm1229",	"r_xnm123",	"r_xnm1230",	"r_xnm1231",	"r_xnm1233",	"r_xnm1234",	"r_xnm1235",	"r_xnm1236",	"r_xnm1237",	"r_xnm1238",	"r_xnm1239",	"r_xnm124",	"r_xnm1240",	"r_xnm1241",	"r_xnm1242",	"r_xnm1243",	"r_xnm1245",	"r_xnm1246",	"r_xnm1247",	"r_xnm1248",	"r_xnm1249",	"r_xnm125",	"r_xnm1250",	"r_xnm1251",	"r_xnm1252",	"r_xnm1253",	"r_xnm1254",	"r_xnm1255",	"r_xnm1257",	"r_xnm1258",	"r_xnm1259",	"r_xnm126",	"r_xnm1260",	"r_xnm1261",	"r_xnm1262",	"r_xnm1263",	"r_xnm1264",	"r_xnm1265",	"r_xnm1266",	"r_xnm1268",	"r_xnm1269",	"r_xnm127",	"r_xnm1270",	"r_xnm1271",	"r_xnm1272",	"r_xnm1273",	"r_xnm1274",	"r_xnm1275",	"r_xnm1276",	"r_xnm1277",	"r_xnm1278",	"r_xnm128",	"r_xnm1280",	"r_xnm1281",	"r_xnm1282",	"r_xnm1283",	"r_xnm1284",	"r_xnm1285",	"r_xnm1286",	"r_xnm1287",	"r_xnm1288",	"r_xnm1289",	"r_xnm129",	"r_xnm1290",	"r_xnm1292",	"r_xnm1293",	"r_xnm1294",	"r_xnm1295",	"r_xnm1296",	"r_xnm1297",	"r_xnm1298",	"r_xnm1299",	"r_xnm13",	"r_xnm130",	"r_xnm1300",	"r_xnm1301",	"r_xnm1302",	"r_xnm1304",	"r_xnm1305",	"r_xnm1306",	"r_xnm1307",	"r_xnm1308",	"r_xnm1309",	"r_xnm131",	"r_xnm1310",	"r_xnm1311",	"r_xnm1312",	"r_xnm1313",	"r_xnm1314",	"r_xnm1316",	"r_xnm1317",	"r_xnm1318",	"r_xnm1319",	"r_xnm132",	"r_xnm1320",	"r_xnm1321",	"r_xnm1322",	"r_xnm1323",	"r_xnm1324",	"r_xnm1325",	"r_xnm1327",	"r_xnm1328",	"r_xnm1329",	"r_xnm133",	"r_xnm1330",	"r_xnm1331",	"r_xnm1332",	"r_xnm1333",	"r_xnm1334",	"r_xnm1335",	"r_xnm1336",	"r_xnm1337",	"r_xnm1339",	"r_xnm134",	"r_xnm1340",	"r_xnm1341",	"r_xnm1342",	"r_xnm1343",	"r_xnm1344",	"r_xnm1345",	"r_xnm1346",	"r_xnm1347",	"r_xnm1348",	"r_xnm1349",	"r_xnm135",	"r_xnm1351",	"r_xnm1352",	"r_xnm1353",	"r_xnm1354",	"r_xnm1355",	"r_xnm1356",	"r_xnm1357",	"r_xnm1358",	"r_xnm1359",	"r_xnm136",	"r_xnm1360",	"r_xnm1361",	"r_xnm1363",	"r_xnm1364",	"r_xnm1365",	"r_xnm1366",	"r_xnm1367",	"r_xnm1368",	"r_xnm1369",	"r_xnm137",	"r_xnm1370",	"r_xnm1371",	"r_xnm1372",	"r_xnm1373",	"r_xnm1375",	"r_xnm1376",	"r_xnm1377",	"r_xnm1378",	"r_xnm1379",	"r_xnm138",	"r_xnm1380",	"r_xnm1381",	"r_xnm1382",	"r_xnm1383",	"r_xnm1384",	"r_xnm1385",	"r_xnm1387",	"r_xnm1388",	"r_xnm1389",	"r_xnm139",	"r_xnm1390",	"r_xnm1391",	"r_xnm1392",	"r_xnm1393",	"r_xnm1394",	"r_xnm1395",	"r_xnm1396",	"r_xnm1397",	"r_xnm1399",	"r_xnm14",	"r_xnm140",	"r_xnm1400",	"r_xnm1401",	"r_xnm1402",	"r_xnm1403",	"r_xnm1404",	"r_xnm1405",	"r_xnm1406",	"r_xnm1407",	"r_xnm1408",	"r_xnm1409",	"r_xnm141",	"r_xnm1411",	"r_xnm1412",	"r_xnm1413",	"r_xnm1414",	"r_xnm1415",	"r_xnm1416",	"r_xnm1417",	"r_xnm1418",	"r_xnm1419",	"r_xnm142",	"r_xnm1420",	"r_xnm1422",	"r_xnm1423",	"r_xnm1424",	"r_xnm1425",	"r_xnm1426",	"r_xnm1427",	"r_xnm1428",	"r_xnm1429",	"r_xnm143",	"r_xnm1430",	"r_xnm1431",	"r_xnm1432",	"r_xnm1434",	"r_xnm1435",	"r_xnm1436",	"r_xnm1437",	"r_xnm1438",	"r_xnm1439",	"r_xnm144",	"r_xnm1440",	"r_xnm1441",	"r_xnm1442",	"r_xnm1443",	"r_xnm1444",	"r_xnm1446",	"r_xnm1447",	"r_xnm1448",	"r_xnm1449",	"r_xnm145",	"r_xnm1450",	"r_xnm1451",	"r_xnm1452",	"r_xnm1453",	"r_xnm1454",	"r_xnm1455",	"r_xnm1456",	"r_xnm1458",	"r_xnm1459",	"r_xnm146",	"r_xnm1460",	"r_xnm1461",	"r_xnm1462",	"r_xnm1463",	"r_xnm1464",	"r_xnm1465",	"r_xnm1466",	"r_xnm1467",	"r_xnm1468",	"r_xnm147",	"r_xnm1470",	"r_xnm1471",	"r_xnm1472",	"r_xnm1473",	"r_xnm1474",	"r_xnm1475",	"r_xnm1476",	"r_xnm1477",	"r_xnm1478",	"r_xnm1479",	"r_xnm148",	"r_xnm1480",	"r_xnm1482",	"r_xnm1483",	"r_xnm1484",	"r_xnm1485",	"r_xnm1486",	"r_xnm1487",	"r_xnm1488",	"r_xnm1489",	"r_xnm149",	"r_xnm1490",	"r_xnm1491",	"r_xnm1492",	"r_xnm1494",	"r_xnm1495",	"r_xnm1496",	"r_xnm1497",	"r_xnm1498",	"r_xnm1499",	"r_xnm15",	"r_xnm150",	"r_xnm1500",	"r_xnm1501",	"r_xnm1502",	"r_xnm1503",	"r_xnm1504",	"r_xnm1506",	"r_xnm1507",	"r_xnm1508",	"r_xnm1509",	"r_xnm151",	"r_xnm1510",	"r_xnm1511",	"r_xnm1512",	"r_xnm1513",	"r_xnm1514",	"r_xnm1515",	"r_xnm1516",	"r_xnm1518",	"r_xnm1519",	"r_xnm152",	"r_xnm1520",	"r_xnm1521",	"r_xnm1522",	"r_xnm1523",	"r_xnm1524",	"r_xnm1525",	"r_xnm1526",	"r_xnm1527",	"r_xnm1528",	"r_xnm153",	"r_xnm1530",	"r_xnm1531",	"r_xnm1532",	"r_xnm1533",	"r_xnm1534",	"r_xnm1535",	"r_xnm1536",	"r_xnm1537",	"r_xnm1538",	"r_xnm1539",	"r_xnm154",	"r_xnm1540",	"r_xnm1542",	"r_xnm1543",	"r_xnm1544",	"r_xnm1545",	"r_xnm1546",	"r_xnm1547",	"r_xnm1548",	"r_xnm1549",	"r_xnm155",	"r_xnm1550",	"r_xnm1551",	"r_xnm1552",	"r_xnm1554",	"r_xnm1555",	"r_xnm1556",	"r_xnm1557",	"r_xnm1558",	"r_xnm1559",	"r_xnm156",	"r_xnm1560",	"r_xnm1561",	"r_xnm1562",	"r_xnm1563",	"r_xnm1565",	"r_xnm1566",	"r_xnm1567",	"r_xnm1568",	"r_xnm1569",	"r_xnm157",	"r_xnm1570",	"r_xnm1571",	"r_xnm1572",	"r_xnm1573",	"r_xnm1574",	"r_xnm1575",	"r_xnm1577",	"r_xnm1578",	"r_xnm1579",	"r_xnm158",	"r_xnm1580",	"r_xnm1581",	"r_xnm1582",	"r_xnm1583",	"r_xnm1584",	"r_xnm1585",	"r_xnm1586",	"r_xnm1587",	"r_xnm1589",	"r_xnm159",	"r_xnm1590",	"r_xnm1591",	"r_xnm1592",	"r_xnm1593",	"r_xnm1594",	"r_xnm1595",	"r_xnm1596",	"r_xnm1597",	"r_xnm1598",	"r_xnm1599",	"r_xnm160",	"r_xnm1601",	"r_xnm1602",	"r_xnm1603",	"r_xnm1604",	"r_xnm1605",	"r_xnm1606",	"r_xnm1607",	"r_xnm1608",	"r_xnm1609",	"r_xnm161",	"r_xnm1610",	"r_xnm1611",	"r_xnm1613",	"r_xnm1614",	"r_xnm1615",	"r_xnm1616",	"r_xnm1617",	"r_xnm1618",	"r_xnm1619",	"r_xnm162",	"r_xnm1620",	"r_xnm1621",	"r_xnm1623",	"r_xnm1625",	"r_xnm1626",	"r_xnm1627",	"r_xnm1628",	"r_xnm1629",	"r_xnm163",	"r_xnm1630",	"r_xnm1631",	"r_xnm1632",	"r_xnm1633",	"r_xnm1634",	"r_xnm1637",	"r_xnm1638",	"r_xnm1639",	"r_xnm164",	"r_xnm1640",	"r_xnm1641",	"r_xnm1642",	"r_xnm1643",	"r_xnm1646",	"r_xnm1647",	"r_xnm1648",	"r_xnm1649",	"r_xnm165",	"r_xnm1650",	"r_xnm1651",	"r_xnm1652",	"r_xnm1653",	"r_xnm166",	"r_xnm1661",	"r_xnm1669",	"r_xnm167",	"r_xnm1670",	"r_xnm1671",	"r_xnm1672",	"r_xnm1673",	"r_xnm1674",	"r_xnm1675",	"r_xnm1676",	"r_xnm1677",	"r_xnm1679",	"r_xnm168",	"r_xnm1680",	"r_xnm1681",	"r_xnm1683",	"r_xnm1710",	"r_xnm1717",	"r_xnm1719",	"r_xnm1721",	"r_xnm1723",	"r_xnm1726",	"r_xnm1728",	"r_xnm173",	"r_xnm1730",	"r_xnm1732",	"r_xnm1735",	"r_xnm1737",	"r_xnm1739",	"r_xnm174",	"r_xnm1741",	"r_xnm1744",	"r_xnm1746",	"r_xnm1748",	"r_xnm175",	"r_xnm1750",	"r_xnm1753",	"r_xnm1755",	"r_xnm1757",	"r_xnm1759",	"r_xnm176",	"r_xnm1762",	"r_xnm1764",	"r_xnm1766",	"r_xnm1768",	"r_xnm177",	"r_xnm1771",	"r_xnm1773",	"r_xnm1775",	"r_xnm1777",	"r_xnm178",	"r_xnm1780",	"r_xnm1785",	"r_xnm1816",	"r_xnm1824",	"r_xnm1827",	"r_xnm1828",	"r_xnm1829",	"r_xnm183",	"r_xnm1830",	"r_xnm1831",	"r_xnm1832",	"r_xnm1833",	"r_xnm1834",	"r_xnm1835",	"r_xnm1836",	"r_xnm1837",	"r_xnm1838",	"r_xnm1839",	"r_xnm184",	"r_xnm1840",	"r_xnm1841",	"r_xnm1842",	"r_xnm1843",	"r_xnm1844",	"r_xnm1845",	"r_xnm1846",	"r_xnm1847",	"r_xnm1848",	"r_xnm1849",	"r_xnm185",	"r_xnm1850",	"r_xnm1851",	"r_xnm1852",	"r_xnm1853",	"r_xnm1854",	"r_xnm1856",	"r_xnm1857",	"r_xnm1858",	"r_xnm1859",	"r_xnm186",	"r_xnm1860",	"r_xnm1861",	"r_xnm1862",	"r_xnm1863",	"r_xnm1865",	"r_xnm1866",	"r_xnm1867",	"r_xnm1868",	"r_xnm1869",	"r_xnm187",	"r_xnm1870",	"r_xnm1871",	"r_xnm1873",	"r_xnm1874",	"r_xnm1875",	"r_xnm1876",	"r_xnm1877",	"r_xnm1878",	"r_xnm188",	"r_xnm1880",	"r_xnm1881",	"r_xnm1882",	"r_xnm1883",	"r_xnm1884",	"r_xnm1886",	"r_xnm1887",	"r_xnm1888",	"r_xnm1890",	"r_xnm1891",	"r_xnm1892",	"r_xnm1894",	"r_xnm1895",	"r_xnm1897",	"r_xnm1901",	"r_xnm1902",	"r_xnm1904",	"r_xnm1905",	"r_xnm1906",	"r_xnm1907",	"r_xnm1908",	"r_xnm1909",	"r_xnm1910",	"r_xnm1911",	"r_xnm1912",	"r_xnm1914",	"r_xnm1915",	"r_xnm1916",	"r_xnm1917",	"r_xnm1918",	"r_xnm1919",	"r_xnm1920",	"r_xnm1921",	"r_xnm1922",	"r_xnm1924",	"r_xnm1925",	"r_xnm1926",	"r_xnm1927",	"r_xnm1928",	"r_xnm1929",	"r_xnm193",	"r_xnm1930",	"r_xnm1931",	"r_xnm1932",	"r_xnm1934",	"r_xnm1935",	"r_xnm1936",	"r_xnm1937",	"r_xnm1938",	"r_xnm1939",	"r_xnm194",	"r_xnm1940",	"r_xnm1941",	"r_xnm1942",	"r_xnm1944",	"r_xnm1945",	"r_xnm1946",	"r_xnm1947",	"r_xnm1948",	"r_xnm1949",	"r_xnm195",	"r_xnm1950",	"r_xnm1951",	"r_xnm1952",	"r_xnm1954",	"r_xnm1955",	"r_xnm1956",	"r_xnm1957",	"r_xnm1958",	"r_xnm1959",	"r_xnm196",	"r_xnm1960",	"r_xnm1961",	"r_xnm1962",	"r_xnm1964",	"r_xnm1965",	"r_xnm1966",	"r_xnm1967",	"r_xnm1968",	"r_xnm1969",	"r_xnm197",	"r_xnm1970",	"r_xnm1971",	"r_xnm1972",	"r_xnm1974",	"r_xnm1975",	"r_xnm1976",	"r_xnm1977",	"r_xnm1978",	"r_xnm1979",	"r_xnm198",	"r_xnm1980",	"r_xnm1981",	"r_xnm1982",	"r_xnm1984",	"r_xnm1985",	"r_xnm1986",	"r_xnm1987",	"r_xnm1988",	"r_xnm1989",	"r_xnm199",	"r_xnm1990",	"r_xnm1991",	"r_xnm1992",	"r_xnm1993",	"r_xnm200",	"r_xnm201",	"r_xnm202",	"r_xnm203",	"r_xnm204",	"r_xnm205",	"r_xnm206",	"r_xnm207",	"r_xnm208",	"r_xnm209",	"r_xnm2091",	"r_xnm2092",	"r_xnm2093",	"r_xnm2094",	"r_xnm2095",	"r_xnm2096",	"r_xnm2097",	"r_xnm210",	"r_xnm2100",	"r_xnm2102",	"r_xnm2103",	"r_xnm2104",	"r_xnm2105",	"r_xnm2106",	"r_xnm2107",	"r_xnm2108",	"r_xnm211",	"r_xnm2112",	"r_xnm2114",	"r_xnm2115",	"r_xnm2116",	"r_xnm2117",	"r_xnm2118",	"r_xnm2119",	"r_xnm212",	"r_xnm2120",	"r_xnm2121",	"r_xnm2123",	"r_xnm2124",	"r_xnm2125",	"r_xnm2126",	"r_xnm2127",	"r_xnm2128",	"r_xnm2129",	"r_xnm213",	"r_xnm2131",	"r_xnm2133",	"r_xnm2134",	"r_xnm2135",	"r_xnm2136",	"r_xnm2137",	"r_xnm2138",	"r_xnm2139",	"r_xnm214",	"r_xnm2144",	"r_xnm2149",	"r_xnm215",	"r_xnm216",	"r_xnm217",	"r_xnm2170",	"r_xnm218",	"r_xnm2192",	"r_xnm2193",	"r_xnm2194",	"r_xnm2201",	"r_xnm2218",	"r_xnm2224",	"r_xnm2225",	"r_xnm2226",	"r_xnm2227",	"r_xnm2228",	"r_xnm2229",	"r_xnm2230",	"r_xnm2231",	"r_xnm2232",	"r_xnm2233",	"r_xnm2234",	"r_xnm2235",	"r_xnm2236",	"r_xnm2237",	"r_xnm2238",	"r_xnm2239",	"r_xnm224",	"r_xnm225",	"r_xnm226",	"r_xnm228",	"r_xnm229",	"r_xnm230",	"r_xnm231",	"r_xnm232",	"r_xnm233",	"r_xnm234",	"r_xnm235",	"r_xnm236",	"r_xnm237",	"r_xnm238",	"r_xnm242",	"r_xnm243",	"r_xnm244",	"r_xnm245",	"r_xnm246",	"r_xnm247",	"r_xnm248",	"r_xnm249",	"r_xnm251",	"r_xnm252",	"r_xnm2521",	"r_xnm2522",	"r_xnm2523",	"r_xnm2524",	"r_xnm2525",	"r_xnm2530",	"r_xnm2531",	"r_xnm2532",	"r_xnm2535",	"r_xnm2536",	"r_xnm2537",	"r_xnm2538",	"r_xnm2539",	"r_xnm2540",	"r_xnm2544",	"r_xnm2545",	"r_xnm2547",	"r_xnm2559",	"r_xnm258",	"r_xnm259",	"r_xnm2595",	"r_xnm2596",	"r_xnm2598",	"r_xnm26",	"r_xnm2603",	"r_xnm2604",	"r_xnm2605",	"r_xnm2606",	"r_xnm2607",	"r_xnm2608",	"r_xnm2610",	"r_xnm2611",	"r_xnm262",	"r_xnm2620",	"r_xnm2623",	"r_xnm2625",	"r_xnm263",	"r_xnm2631",	"r_xnm2632",	"r_xnm2634",	"r_xnm2639",	"r_xnm264",	"r_xnm2640",	"r_xnm2641",	"r_xnm2642",	"r_xnm2643",	"r_xnm2644",	"r_xnm2646",	"r_xnm2647",	"r_xnm265",	"r_xnm2656",	"r_xnm2659",	"r_xnm266",	"r_xnm2661",	"r_xnm267",	"r_xnm268",	"r_xnm269",	"r_xnm27",	"r_xnm270",	"r_xnm271",	"r_xnm272",	"r_xnm276",	"r_xnm277",	"r_xnm278",	"r_xnm279",	"r_xnm28",	"r_xnm280",	"r_xnm281",	"r_xnm282",	"r_xnm283",	"r_xnm285",	"r_xnm286",	"r_xnm29",	"r_xnm292",	"r_xnm293",	"r_xnm296",	"r_xnm297",	"r_xnm298",	"r_xnm299",	"r_xnm30",	"r_xnm300",	"r_xnm301",	"r_xnm302",	"r_xnm303",	"r_xnm304",	"r_xnm305",	"r_xnm306",	"r_xnm31",	"r_xnm310",	"r_xnm311",	"r_xnm312",	"r_xnm313",	"r_xnm314",	"r_xnm315",	"r_xnm316",	"r_xnm317",	"r_xnm32",	"r_xnm320",	"r_xnm3207",	"r_xnm3208",	"r_xnm3210",	"r_xnm3214",	"r_xnm3215",	"r_xnm3216",	"r_xnm3217",	"r_xnm3218",	"r_xnm3219",	"r_xnm3220",	"r_xnm3222",	"r_xnm3223",	"r_xnm3232",	"r_xnm3235",	"r_xnm3237",	"r_xnm326",	"r_xnm327",	"r_xnm33",	"r_xnm330",	"r_xnm331",	"r_xnm332",	"r_xnm333",	"r_xnm334",	"r_xnm335",	"r_xnm336",	"r_xnm337",	"r_xnm338",	"r_xnm339",	"r_xnm34",	"r_xnm340",	"r_xnm344",	"r_xnm345",	"r_xnm346",	"r_xnm347",	"r_xnm348",	"r_xnm349",	"r_xnm35",	"r_xnm350",	"r_xnm351",	"r_xnm36",	"r_xnm3604",	"r_xnm3606",	"r_xnm3611",	"r_xnm3612",	"r_xnm3613",	"r_xnm3614",	"r_xnm3615",	"r_xnm3616",	"r_xnm3618",	"r_xnm3619",	"r_xnm3628",	"r_xnm3631",	"r_xnm3632",	"r_xnm3633",	"r_xnm37",	"r_xnm38",	"r_xnm39",	"r_xnm40",	"r_xnm4000",	"r_xnm4002",	"r_xnm4007",	"r_xnm4008",	"r_xnm4009",	"r_xnm4011",	"r_xnm4012",	"r_xnm4014",	"r_xnm4015",	"r_xnm4024",	"r_xnm4027",	"r_xnm41",	"r_xnm42",	"r_xnm428",	"r_xnm429",	"r_xnm43",	"r_xnm432",	"r_xnm433",	"r_xnm434",	"r_xnm435",	"r_xnm436",	"r_xnm437",	"r_xnm438",	"r_xnm439",	"r_xnm4396",	"r_xnm4398",	"r_xnm44",	"r_xnm440",	"r_xnm4403",	"r_xnm4404",	"r_xnm4405",	"r_xnm4407",	"r_xnm4408",	"r_xnm441",	"r_xnm4410",	"r_xnm4411",	"r_xnm442",	"r_xnm4420",	"r_xnm4423",	"r_xnm443",	"r_xnm444",	"r_xnm445",	"r_xnm446",	"r_xnm447",	"r_xnm448",	"r_xnm449",	"r_xnm45",	"r_xnm450",	"r_xnm451",	"r_xnm452",	"r_xnm453",	"r_xnm454",	"r_xnm455",	"r_xnm456",	"r_xnm457",	"r_xnm458",	"r_xnm459",	"r_xnm46",	"r_xnm460",	"r_xnm461",	"r_xnm462",	"r_xnm463",	"r_xnm464",	"r_xnm465",	"r_xnm466",	"r_xnm467",	"r_xnm468",	"r_xnm469",	"r_xnm47",	"r_xnm470",	"r_xnm471",	"r_xnm472",	"r_xnm473",	"r_xnm474",	"r_xnm477",	"r_xnm478",	"r_xnm4792",	"r_xnm4794",	"r_xnm4799",	"r_xnm48",	"r_xnm4800",	"r_xnm4801",	"r_xnm4803",	"r_xnm4804",	"r_xnm4806",	"r_xnm4807",	"r_xnm4816",	"r_xnm4819",	"r_xnm483",	"r_xnm488",	"r_xnm489",	"r_xnm49",	"r_xnm490",	"r_xnm491",	"r_xnm492",	"r_xnm493",	"r_xnm494",	"r_xnm495",	"r_xnm496",	"r_xnm497",	"r_xnm498",	"r_xnm499",	"r_xnm50",	"r_xnm500",	"r_xnm501",	"r_xnm502",	"r_xnm503",	"r_xnm504",	"r_xnm505",	"r_xnm506",	"r_xnm507",	"r_xnm508",	"r_xnm509",	"r_xnm51",	"r_xnm510",	"r_xnm511",	"r_xnm512",	"r_xnm513",	"r_xnm514",	"r_xnm5188",	"r_xnm5190",	"r_xnm5195",	"r_xnm5196",	"r_xnm5197",	"r_xnm5199",	"r_xnm52",	"r_xnm5200",	"r_xnm5202",	"r_xnm5203",	"r_xnm5212",	"r_xnm5215",	"r_xnm5216",	"r_xnm53",	"r_xnm538",	"r_xnm54",	"r_xnm542",	"r_xnm55",	"r_xnm5586",	"r_xnm5591",	"r_xnm5592",	"r_xnm5593",	"r_xnm5595",	"r_xnm5596",	"r_xnm5598",	"r_xnm5599",	"r_xnm5608",	"r_xnm5611",	"r_xnm58",	"r_xnm59",	"r_xnm5980",	"r_xnm5982",	"r_xnm5991",	"r_xnm5992",	"r_xnm5994",	"r_xnm5995",	"r_xnm60",	"r_xnm6004",	"r_xnm6007",	"r_xnm6008",	"r_xnm61",	"r_xnm613",	"r_xnm614",	"r_xnm62",	"r_xnm63",	"r_xnm64",	"r_xnm640",	"r_xnm641",	"r_xnm642",	"r_xnm643",	"r_xnm644",	"r_xnm645",	"r_xnm6451",	"r_xnm646",	"r_xnm647",	"r_xnm648",	"r_xnm649",	"r_xnm65",	"r_xnm650",	"r_xnm651",	"r_xnm6512",	"r_xnm6515",	"r_xnm652",	"r_xnm653",	"r_xnm654",	"r_xnm6541",	"r_xnm6543",	"r_xnm6544",	"r_xnm6545",	"r_xnm6549",	"r_xnm655",	"r_xnm656",	"r_xnm657",	"r_xnm658",	"r_xnm659",	"r_xnm66",	"r_xnm669",	"r_xnm67",	"r_xnm670",	"r_xnm68",	"r_xnm69",	"r_xnm694",	"r_xnm696",	"r_xnm699",	"r_xnm70",	"r_xnm700",	"r_xnm701",	"r_xnm702",	"r_xnm703",	"r_xnm704",	"r_xnm705",	"r_xnm706",	"r_xnm707",	"r_xnm708",	"r_xnm709",	"r_xnm71",	"r_xnm710",	"r_xnm711",	"r_xnm712",	"r_xnm713",	"r_xnm714",	"r_xnm715",	"r_xnm716",	"r_xnm717",	"r_xnm718",	"r_xnm719",	"r_xnm72",	"r_xnm720",	"r_xnm721",	"r_xnm722",	"r_xnm723",	"r_xnm724",	"r_xnm725",	"r_xnm726",	"r_xnm73",	"r_xnm74",	"r_xnm75",	"r_xnm76",	"r_xnm77",	"r_xnm78",	"r_xnm79",	"r_xnm80",	"r_xnm81",	"r_xnm82",	"r_xnm821",	"r_xnm822",	"r_xnm823",	"r_xnm824",	"r_xnm825",	"r_xnm826",	"r_xnm827",	"r_xnm828",	"r_xnm829",	"r_xnm83",	"r_xnm830",	"r_xnm831",	"r_xnm832",	"r_xnm833",	"r_xnm834",	"r_xnm835",	"r_xnm836",	"r_xnm837",	"r_xnm838",	"r_xnm839",	"r_xnm84",	"r_xnm840",	"r_xnm841",	"r_xnm842",	"r_xnm843",	"r_xnm844",	"r_xnm845",	"r_xnm846",	"r_xnm847",	"r_xnm85",	"r_xnm855",	"r_xnm856",	"r_xnm857",	"r_xnm858",	"r_xnm859",	"r_xnm860",	"r_xnm861",	"r_xnm863",	"r_xnm865",	"r_xnm866",	"r_xnm867",	"r_xnm868",	"r_xnm869",	"r_xnm87",	"r_xnm870",	"r_xnm871",	"r_xnm872",	"r_xnm873",	"r_xnm874",	"r_xnm876",	"r_xnm877",	"r_xnm878",	"r_xnm879",	"r_xnm880",	"r_xnm881",	"r_xnm882",	"r_xnm883",	"r_xnm884",	"r_xnm885",	"r_xnm887",	"r_xnm888",	"r_xnm889",	"r_xnm89",	"r_xnm890",	"r_xnm891",	"r_xnm892",	"r_xnm893",	"r_xnm894",	"r_xnm895",	"r_xnm896",	"r_xnm897",	"r_xnm899",	"r_xnm90",	"r_xnm900",	"r_xnm901",	"r_xnm902",	"r_xnm903",	"r_xnm904",	"r_xnm905",	"r_xnm906",	"r_xnm907",	"r_xnm908",	"r_xnm909",	"r_xnm91",	"r_xnm911",	"r_xnm912",	"r_xnm913",	"r_xnm914",	"r_xnm915",	"r_xnm916",	"r_xnm917",	"r_xnm918",	"r_xnm919",	"r_xnm92",	"r_xnm920",	"r_xnm921",	"r_xnm923",	"r_xnm924",	"r_xnm925",	"r_xnm926",	"r_xnm927",	"r_xnm928",	"r_xnm929",	"r_xnm93",	"r_xnm930",	"r_xnm931",	"r_xnm932",	"r_xnm933",	"r_xnm935",	"r_xnm936",	"r_xnm937",	"r_xnm938",	"r_xnm939",	"r_xnm94",	"r_xnm940",	"r_xnm941",	"r_xnm942",	"r_xnm943",	"r_xnm944",	"r_xnm945",	"r_xnm947",	"r_xnm948",	"r_xnm949",	"r_xnm95",	"r_xnm950",	"r_xnm951",	"r_xnm952",	"r_xnm953",	"r_xnm954",	"r_xnm955",	"r_xnm956",	"r_xnm957",	"r_xnm959",	"r_xnm96",	"r_xnm960",	"r_xnm961",	"r_xnm962",	"r_xnm963",	"r_xnm964",	"r_xnm965",	"r_xnm966",	"r_xnm967",	"r_xnm968",	"r_xnm969",	"r_xnm97",	"r_xnm971",	"r_xnm972",	"r_xnm973",	"r_xnm974",	"r_xnm975",	"r_xnm976",	"r_xnm977",	"r_xnm978",	"r_xnm979",	"r_xnm98",	"r_xnm980",	"r_xnm982",	"r_xnm983",	"r_xnm984",	"r_xnm985",	"r_xnm986",	"r_xnm987",	"r_xnm988",	"r_xnm989",	"r_xnm99",	"r_xnm990",	"r_xnm991",	"r_xnm992",	"r_xnm994",	"r_xnm995",	"r_xnm996",	"r_xnm997",	"r_xnm998",	"r_xnm999",	"vantage",
]
#premier special value replacement***

df_renamed = df.rename(columns={'loannumber': 'uniq_id', 'everdef': 'bad'})
temp_1a = df_renamed.sort_values(by='uniq_id')

## PROC TRANSPOSE equivalent (first one)
# Melt the DataFrame to transform numeric variables
temp2_ = pd.melt(temp_1a, 
                  id_vars=[ 'uniq_id', 'bad', 'weight'],
                  value_vars=pre_attr,
                  var_name='x_nm',
                  value_name='x_value')

PREMIER_ATTR_LU_1203=pd.read_csv(r'D:\HA_study_plan_tasks\banking_financial_scorecard_dev\model_1\lu\PREMIER_ATTR_LU_1203.csv')

df_attr = duckdb.query("SELECT * FROM lu.PREMIER_ATTR_LU_1203").to_df()

# claim
length_sv_map = {}

# Iterate over lengths from 2 to 9 (skipping 8)）
for length in range(2, 10):
    if length == 8:
        continue  # skipping length=8
    subset = df_attr[df_attr['length'] == (length)]  
    quoted_names = ["'" + attr + "'" for attr in subset['attr_name']]
    sv_string = ",".join(quoted_names)
    length_sv_map[f"_{length}_sv"] = sv_string
    
def parse_sv_dict(sv_dict):
    parsed = {}
    for k, v in sv_dict.items():
        parsed[k] = [x.strip().strip("'").strip('"') for x in v.split(',')]
    return parsed

length_sv_map = parse_sv_dict(length_sv_map)
    
#print(length_sv_map["_2_sv"])
# output: 'age','sex','zip'（depending on data）

##special values bear in mind, no need to treated as in SAS program
def convert_special_missing(temp_2, sv_dict):
    def map_special(row):
        xnm = row['x_nm']
        val = row['x_value']
        
        for length in range(2, 10):
            if length == 8:
                continue  # skipping 8

            sv_list = sv_dict.get(f"_{length}_sv", [])
            base = int('9' * length)
            if xnm in sv_list:
                if val == base-9 + 1: return -10  # .a
                elif val == base-9 + 2: return -20  # .b
                elif val == base-9 + 3: return -30  # .c
                elif val == base-9 + 4: return  -40  # .d
                elif val == base-9 + 5: return  -50  # .e
                elif val == base-9 + 6: return  -60  # .f
                elif val == base-9 + 7: return  -70  # .g
                elif val == base-9 + 8: return   -80 # .h
                elif val == base-9 + 9: return   -90 # .i
        return val

    temp_2['x_value_'] = temp_2.apply(map_special, axis=1)
    return temp_2  
#call the function
sv_data = convert_special_missing(temp2_, length_sv_map)

#rmianing no rule of special value replacement:
# define the length of x_nm
group_dict = {
    2: {'ALL0317','ALL0448','ALL1380','ALL2002','ALL2005','ALL2427','ALL2428','ALL2907',
        'ALL2937','ALL2967','BCA0416','BCC0446','BCC3342','BCC3345','BCC3512','BCX3421',
        'BCX3422','BRC0416','BRC1300','FIP0300','FIP2358','MTF2358'},
    3: {'ALL4080', 'BCC7117', 'BRC7140'},
    4: {'ALL8325', 'MTF8166'},
    9: {'BCA5030','BCA5070','BCA5740','BRC5320','BRC5830','REV5620','ALL5743','ALL5070','ALL5047'}
}

# define a mapping function
def map_special_sas_like(temp_2, sv_dict):
    def map_special(row):
        xnm = row['x_nm']
        val = row['x_value']
        xnm = row['x_nm']
        val = row['x_value']
        for length in (2, 3,4,9):

            sv_list = sv_dict.get(length, [])
            base = int('9' * length)
            if xnm in sv_list:
                if val == base-9 + 1: return -10  # .a
                elif val == base-9 + 2: return -20  # .b
                elif val == base-9 + 3: return -30  # .c
                elif val == base-9 + 4: return  -40  # .d
                elif val == base-9 + 5: return  -50  # .e
                elif val == base-9 + 6: return  -60  # .f
                elif val == base-9 + 7: return  -70  # .g
                elif val == base-9 + 8: return   -80 # .h
                elif val == base-9 + 9: return   -90 # .i
        return val

    temp_2['x_value_'] = temp_2.apply(map_special, axis=1)
    return temp_2  

#call function
sv_data_2= map_special_sas_like(sv_data, group_dict)
#transpose to wide format
df_pivot = sv_data_2.pivot(index='uniq_id', columns='x_nm', values='x_value_').reset_index()

# Step 2: summary stats (such as proc means)
summary = df_pivot[pre_attr].describe(include='all')  
print(summary)
# still some special values not replaced, due to they will not be in model candidate**


#check camp 
check_camp=duckdb.query(f"""
                        select camp, count(weight) as cnt from temp_1a group by camp
                        """).to_df()

#data sererated into dev and OOT 
model_OOT = temp_1a[temp_1a['camp'] == '2016-10'][['uniq_id', 'weight', 'bad'] + c_m_r_list + n_m_r_list]
model_dev = temp_1a[temp_1a['camp'] != '2016-10'][['uniq_id', 'weight', 'bad'] + c_m_r_list + n_m_r_list]


###full dev fast modeling prep
#[1]  xKS, IV***
# Sort by uniq_id
model_dev = model_dev.sort_values(by=['uniq_id'])

temp_2_dev = pd.melt(
    model_dev,
    id_vars=['uniq_id', 'bad', 'weight'],
    value_vars=n_m_r_list,   
    var_name='x_nm',
    value_name='x_value'
)
devx_ks_df = DR_x_weighted_KS_pandas(temp_2_dev, bad='bad', weight='weight', out_name='devx_ks_df')

#pulling  xKS>0** x_nm
temp_3_dev = temp_2_dev.merge(
    devx_ks_df[devx_ks_df['xKS'] > 0],  # Filter rows where xKS > 0
    on='x_nm',                       # Join on x_nm column
    how='inner'    )                  # Inner join (like SAS default)

temp_3_dev = temp_3_dev.sort_values(by=temp_3_dev.columns[1], ascending=False)

out_devdata_rbin_sql, out_devnum_rbins = DR_num_rbin_duckdb(
    df_input=temp_3_dev,  # including x_nm, x_value, weight, bad etc. columns
    uniq_id='loan_id',
    bad='bad',
    weight='weight',
    max_bin=5,
    min_adj_bin=2
)

dfdev_woe, dfdev_iv = DR_num_iv_woe_duckdb(out_devdata_rbin_sql)

#######Candidate variables 
temp_4_dev = duckdb.query(f"""
    SELECT * from devx_ks_df a inner join dfdev_iv b on a.x_nm=b.x_nm order by a.xKS DESC
""").to_df()
duckdb.register("temp_4_dev", temp_4_dev)

temp_sel_dev = temp_4_dev[
    ((temp_4_dev['rbin_cnt'] >= 4) & (temp_4_dev['iv'] >= 0.005)) |
    ((temp_4_dev['rbin_cnt'] < 4) & (temp_4_dev['xKS'] >= 4))
].copy()


#***************************[ Auto: MFC/F/C  *************************
temp_5a_dev=temp_sel_dev[(temp_sel_dev['nomiss_rbin_cnt']>=2)]

temp_5_dev=duckdb.query(f"""
                    select a.* from out_devdata_rbin_sql a inner join temp_5a_dev b on a.x_nm=b.x_nm
                    """).to_df()

temp_6_dev= duckdb.query(f"""
                     select a.* from dfdev_iv a inner join temp_5a_dev b on a.x_nm=b.x_nm
                     """).to_df()
                     
                     
temp_7_dev=duckdb.query(f"""
                    select a.* from dfdev_woe a inner join temp_5a_dev b on a.x_nm=b.x_nm
                    """).to_df()

#call the function as above
duckdb.register("in_data_rbin", temp_5_dev)
duckdb.register("in_num_iv", temp_6_dev)
duckdb.register("in_num_woe", temp_7_dev)

out_val_dev, out_sas_dev = DR_num_aMFC("in_data_rbin", "in_num_iv", "in_num_woe", 
                               uniq_id="uniq_id", bad="bad", weight="weight")




###***************************[3] Auto: woe *************************
temp_8_dev =temp_sel_dev[(temp_sel_dev['nomiss_rbin_cnt']<2)] 

temp_9_dev=duckdb.query(f"""
                    select a.* from out_devdata_rbin_sql a inner join temp_8_dev b on a.x_nm=b.x_nm
                    """).to_df()
 
temp_10_dev= duckdb.query(f"""
                      select a.*, b.woe from temp_9_dev a left join dfdev_woe b on a.x_nm=b.x_nm and (a.rbin = b.rbin OR (a.rbin IS NULL AND b.rbin IS NULL))
 order by a.uniq_id, a.x_nm
                      """).to_df()

out_dfdev = transpose_woe(temp_10_dev)


out_IV_data_sas_dev=duckdb.query(f"""
                             select  * from out_sas_dev a inner join out_dfdev b on a.uniq_id=b.uniq_id and a.bad=b.bad
                             """).to_df()
                             
                             
##char to IV and woe***##char to IV and woe***##char to IV and woe***##char to IV and woe***                             
##char to IV and woe***##char to IV and woe***##char to IV and woe***##char to IV and woe***
model_dev = model_dev.sort_values(['uniq_id', 'bad', 'weight'])
model_dev_charwoe = model_dev.rename(columns={ 'uniq_id': 'obs_id'})  #'everdef': 'bad' already in data, but usually need to do this step
charwoe_dev_attr = pd.melt(model_dev_charwoe,
                  id_vars=[ 'obs_id', 'bad', 'weight'],
                  value_vars=c_m_r_list,
                  var_name='x_nm',
                  value_name='x_value')

                            

#function of char to iv and woe **
def DR_char_iv_woe_duckdb(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    
    duckdb.register("char_data", df)

    # Step 1: calculate bin-level IV/WOE（including log protection）
    df_bin = duckdb.query("""
        SELECT 
            h.*, 
            g.tot_wgt,
            h.sum_wgt / g.tot_wgt AS dist_pct,
            h.sum_bad_wgt / g.tot_bad_wgt AS dist_bad_pct,
            h.sum_good_wgt / g.tot_good_wgt AS dist_good_pct,
            g.tot_miss_wgt / g.tot_wgt AS tot_miss_rate,
            g.tot_bad_rate,
            CASE 
                WHEN h.bad_rate = 0.0 THEN 0.5 / h.sum_wgt
                WHEN h.bad_rate = 1.0 THEN ((h.sum_wgt - 1) - 0.5) / h.sum_wgt
                ELSE h.bad_rate
            END AS cal_bad_rate,
            CASE 
                WHEN h.bad_rate = 0.0 THEN NULL 
                WHEN h.bad_rate = 1.0 THEN NULL 
                ELSE LOG(h.bad_rate / (1.0 - h.bad_rate))
            END AS logit_bad_rate,
            CASE 
                WHEN dist_good_pct > 0 AND dist_bad_pct > 0 THEN 
                    LOG(dist_good_pct / dist_bad_pct)
                ELSE NULL 
            END AS woe,
            CASE 
                WHEN dist_good_pct > 0 AND dist_bad_pct > 0 THEN 
                    (dist_good_pct - dist_bad_pct) * LOG(dist_good_pct / dist_bad_pct)
                ELSE 0.0 
            END AS bin_iv
        FROM (
            SELECT 
                x_nm, x_value,
                SUM(weight) AS sum_wgt,
                SUM(weight * bad) AS sum_bad_wgt,
                SUM(weight) - SUM(weight * bad) AS sum_good_wgt,
                SUM(weight * bad) / SUM(weight) AS bad_rate
            FROM char_data
            GROUP BY x_nm, x_value
        ) h
        LEFT JOIN (
            SELECT 
                x_nm,
                SUM(weight) AS tot_wgt,
                SUM(weight * bad) AS tot_bad_wgt,
                SUM(weight * (1.0 - bad)) AS tot_good_wgt,
                SUM(CASE WHEN x_value IS NULL THEN weight ELSE 0.0 END) AS tot_miss_wgt,
                SUM(weight * bad) / SUM(weight) AS tot_bad_rate
            FROM char_data
            GROUP BY x_nm
        ) g
        ON h.x_nm = g.x_nm
        ORDER BY h.x_nm, cal_bad_rate
    """).to_df()

    # Step 2: replace missing WOE with the maximum WOE of each variable
    non_missing = df_bin[df_bin["woe"].notnull()].copy()
    non_missing["woe_"] = non_missing["woe"]

    max_woe = non_missing.sort_values(["x_nm", "bad_rate"]).drop_duplicates("x_nm", keep="last")[["x_nm", "woe"]]

    missing = df_bin[df_bin["woe"].isnull()].copy()
    missing = missing.merge(max_woe, on="x_nm", how="left", suffixes=("", "_filled"))
    missing["woe_"] = missing["woe_filled"]

    # merge data
    df_woe = pd.concat([non_missing, missing], axis=0, ignore_index=True)
    df_woe = df_woe.sort_values(["x_nm", "cal_bad_rate"])

    # Step 3: cumulative distribution
    df_woe["cum_dist_pct"] = df_woe.groupby("x_nm")["dist_pct"].cumsum()
    df_woe["cum_dist_bad_pct"] = df_woe.groupby("x_nm")["dist_bad_pct"].cumsum()
    df_woe["cum_dist_good_pct"] = df_woe.groupby("x_nm")["dist_good_pct"].cumsum()
    df_woe["diff_cum_bad_good_pct"] = df_woe["cum_dist_bad_pct"] - df_woe["cum_dist_good_pct"]

    # Step 4: summarized output
    df_iv = df_woe.groupby(["x_nm", "tot_wgt", "tot_miss_rate"]).agg(
        x_grp_cnt=("x_value", "count"),
        IV=("bin_iv", "sum"),
        KS=("diff_cum_bad_good_pct", lambda x: np.max(np.abs(x)))
    ).reset_index().sort_values(by="x_nm")

    return df_woe, df_iv

def DR_char_grp_to_woe_duckdb(df_raw: pd.DataFrame, df_woe_map: pd.DataFrame) -> pd.DataFrame:
    import duckdb
    import pandas as pd

    duckdb.register("raw_data", df_raw)
    duckdb.register("woe_map", df_woe_map)

    df_merged = duckdb.query("""
        SELECT a.*, b.woe_
        FROM raw_data a
        LEFT JOIN woe_map b
        ON a.x_nm = b.x_nm AND a.x_value = b.x_value
        ORDER BY a.obs_id, a.x_nm
    """).to_df()

    # pivot the WOE of each variable into a single row
    df_transformed = df_merged.pivot_table(
        index=["obs_id", "bad", "weight"], 
        columns="x_nm", 
        values="woe_"
    ).reset_index()

    return df_transformed

#call the two function
df_char_woe, df_char_iv = DR_char_iv_woe_duckdb(charwoe_dev_attr)

df_char_woe_ready = DR_char_grp_to_woe_duckdb(charwoe_dev_attr, df_char_woe)

#remove 100% constant values and only keep below attr**
c_raw_keep=["r_xnm1622",
"r_xnm1662",
"r_xnm2002",
"r_xnm2075",
"r_xnm21",
"r_xnm2140",
"r_xnm23",
"r_xnm2533",
"r_xnm2534",
"r_xnm6453",
"r_xnm6540",
"r_xnm6542",
"r_xnm1",
"r_xnm421",
"PaydayMethod",
"r_xnm6536",
"IncomeSource",
"r_xnm6546",
"Carrier",]

model_dev = model_dev.sort_values(['uniq_id', 'bad', 'weight'])
charwoe_dev_attr_ = pd.melt(model_dev_charwoe,
                  id_vars=[ 'obs_id', 'bad', 'weight'],
                  value_vars=c_raw_keep,
                  var_name='x_nm',
                  value_name='x_value')

df_char_woe_2, df_char_iv_2 = DR_char_iv_woe_duckdb(charwoe_dev_attr_)
df_char_woe_ready = DR_char_grp_to_woe_duckdb(charwoe_dev_attr_, df_char_woe_2)



#premier exclusively attr model sel**#premier exclusively attr model sel**
#premier exclusively attr model sel**#premier exclusively attr model sel**
sv_data=duckdb.query(f"""
                     select a.*, b.bad, b.weight from df_pivot a inner join model_dev b on a.uniq_id=b.uniq_id
                     """).to_df()

sv_data = sv_data.sort_values(by=['uniq_id'])
sv_data2 = pd.melt(	
    sv_data,	
    id_vars=['uniq_id', 'bad', 'weight'],	
    value_vars=pre_attr,   	
    var_name='x_nm',	
    value_name='x_value'	
)	


prex_ks_df = DR_x_weighted_KS_pandas(sv_data2, bad='bad', weight='weight', out_name='prex_ks_df')	



#pulling  xKS>0** x_nm								
sv_data3 = sv_data2.merge(								
    prex_ks_df[prex_ks_df['xKS'] > 0],  # Filter rows where xKS > 0								
    on='x_nm',                       # Join on x_nm column								
    how='inner'                      # Inner join (like SAS default)								
)		
						
sv_data3 = sv_data3.sort_values(by=sv_data3.columns[1], ascending=False)								
								
out_data_rbin_sqlpre, out_num_rbinspre = DR_num_rbin_duckdb(								
    df_input=sv_data3,  # including x_nm, x_value, weight, bad etc. columns								
    uniq_id='loan_id',								
    bad='bad',								
    weight='weight',								
    max_bin=5,								
    min_adj_bin=5								
)								
								
predf_woe, predf_iv = DR_num_iv_woe_duckdb(out_data_rbin_sqlpre)								
								
#######Candidate variables 								
temp_4pre = duckdb.query(f"""								
    SELECT * from prex_ks_df a inner join predf_iv b on a.x_nm=b.x_nm order by a.xKS DESC								
""").to_df()
duckdb.register("temp_4pre", temp_4pre)							
								
temp_selpre = temp_4pre[								
    ((temp_4pre['rbin_cnt'] >= 4) & (temp_4pre['iv'] >= 0.005)) |								
    ((temp_4pre['rbin_cnt'] < 4) & (temp_4pre['xKS'] >= 4))								
].copy()								
								
								
#***************************[ Auto: MFC/F/C  *************************								
temp_5apre=temp_selpre[(temp_selpre['nomiss_rbin_cnt']>=2)]								
								
temp_5pre=duckdb.query(f"""								
                    select a.* from out_data_rbin_sqlpre a inner join temp_5apre b on a.x_nm=b.x_nm								
                    """).to_df()								
								
temp_6pre= duckdb.query(f"""								
                     select a.* from predf_iv a inner join temp_5apre b on a.x_nm=b.x_nm								
                     """).to_df()								
                     								
                     								
temp_7pre=duckdb.query(f"""								
                    select a.* from predf_woe a inner join temp_5apre b on a.x_nm=b.x_nm								
                    """).to_df()								
								
#call the function as above								
duckdb.register("in_data_rbin", temp_5pre)								
duckdb.register("in_num_iv", temp_6pre)								
duckdb.register("in_num_woe", temp_7pre)								
								
preout_val, preout_sas = DR_num_aMFC("in_data_rbin", "in_num_iv", "in_num_woe", 								
                               uniq_id="uniq_id", bad="bad", weight="weight")								
								
								
								
								
###***************************[3] Auto: woe *************************								
temp_8pre =temp_selpre[(temp_selpre['nomiss_rbin_cnt']<2)] 								
								
temp_9pre=duckdb.query(f"""								
                    select a.* from out_data_rbin_sqlpre a inner join temp_8pre b on a.x_nm=b.x_nm								
                    """).to_df()								
 								
temp_10pre= duckdb.query(f"""								
                      select a.*, b.woe from temp_9pre a left join predf_woe b on a.x_nm=b.x_nm and (a.rbin = b.rbin OR (a.rbin IS NULL AND b.rbin IS NULL))								
 order by a.uniq_id, a.x_nm								
                      """).to_df()								
								
preout_df = transpose_woe(temp_10pre)								
								
								
preout_IV_data_sas=duckdb.query(f"""								
                             select  * from preout_sas a inner join preout_df b on a.uniq_id=b.uniq_id and a.bad=b.bad								
                             """).to_df()								


drop_list=["ALL0100",	"ALL0200",	"ALL0216",	"ALL0317",	"ALL0336",	"ALL0400",	"ALL0416",	"ALL0436",	"ALL0446",	"ALL0448",	"ALL0700",	"ALL1380",	"ALL2000",	"ALL2001",	"ALL2002",	"ALL2005",	"ALL2327",	"ALL2380",	"ALL2480",	"ALL4080",	"ALL8164",	"ALL8325",	"ALX5839",	"AUA0416",	"AUA1300",	"AUA2358",	"AUA5820",	"AUA6160",	"BCA0416",	"BCA5030",	"BCA6220",	"BCC0436",	"BCC0446",	"BCC1360",	"BCC3342",	"BCC3345",	"BCC3423",	"BCC3512",	"BCC5020",	"BCC5320",	"BCC5421",	"BCC5620",	"BCC5627",	"BCC7117",	"BCC8120",	"BCX3421",	"BCX3422",	"BRC0416",	"BRC1300",	"BRC5320",	"BRC5830",	"BRC7140",	"COL3210",	"COL5060",	"COL8165",	"FIP0300",	"FIP2358",	"IQT9415",	"IQT9420",	"IQT9425",	"IQT9426",	"MTA6160",	"MTF2358",	"MTF8166",	"MTX5839",	"PIL5020",	"REV0300",	"REV1360",	"REV2320",	"REV3421",	"REV5020",	"REV5620",	"REV5627",	"REV8160",	"RTI5020",	"RTI5820",	"RTR3422",	"STU5820",
]

##finalo model full data for dimnetion reduction (like proc varclus  etc)
cols_to_keep = [col for col in out_IV_data_sas_dev.columns if col not in drop_list]
out_IV_data_sas_dev2 = out_IV_data_sas_dev[cols_to_keep]

# register tables
duckdb.register("out_IV_data_sas_dev2", out_IV_data_sas_dev2)
duckdb.register("df_char_woe_ready", df_char_woe_ready)
duckdb.register("model_raw_data", df          )
duckdb.register("preout_IV_data_sas", preout_IV_data_sas)

# SQL join search
auto_model_data = duckdb.query("""
    SELECT 
        a.*, 
        b.*, 
        c.camp,d.*
    FROM out_IV_data_sas_dev2 a
    LEFT JOIN df_char_woe_ready b
        ON a.uniq_id = b.obs_id 
        AND a.bad = b.bad 
        AND a.weight = b.weight
    LEFT JOIN model_raw_data c
        ON a.uniq_id = c.loannumber 
        AND a.bad = c.everdef
       LEFT JOIN preout_sas d
           ON a.uniq_id = d.uniq_id
   
""").to_df()


 # LEFT JOIN preout_IV_data_sas d
 #     ON a.uniq_id = d.uniq_id

##checking point

checka=duckdb.query(f"""
                    select camp, count(*) as cnt, sum(bad)/count(*) as badrate  from auto_model_data
                    group by camp
                    """).to_df()
#for dimention reduction  asigning pival variable
# set tag variable
auto_model_data["tag"] = 4  # The default is tag = 4

auto_model_data.loc[auto_model_data["camp"].isin(['2016-05', '2016-06']), "tag"] = 1
auto_model_data.loc[auto_model_data["camp"].isin(['2016-07', '2016-08']), "tag"] = 2
auto_model_data.loc[auto_model_data["camp"].isin(['2016-12', '2017-01', '2017-02', '2017-04']), "tag"] = 3

# delete columns
auto_model_data2 = auto_model_data.drop(columns=["vantage", "obs_id", "w", "camp", "_NAME_",
                                                 "uniq_id_1",	"bad_1",	"weight_1","bad_2",	"weight_2",
                                                 "uniq_id_1_1",	"bad_1_1",	"weight_1_1","uniq_id_2",	"bad_3",	"weight_3"
], errors="ignore")

#check point
checkb=duckdb.query(f"""
                    select tag, count(*) as cnt , sum(bad) /count(*) as badrate  from auto_model_data2 group by tag""").to_df()

#function 
#dr_woe_dim_reduction_more
from sklearn.decomposition import PCA
from sklearn.linear_model import LogisticRegression
from sklearn.feature_selection import SequentialFeatureSelector
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from sklearn import set_config
set_config(enable_metadata_routing=True)


def dr_woe_dim_reduction_more(
    df_woe: pd.DataFrame,
    uniq_id: str,
    bad: int,
    weight: int,
    data_subgrp: int,
    clus_cnt: int = 10,
    top_n: int = 1,
    sle: float = 0.05,
    forward_stop_n: int = 10
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Python version of %dr_woe_dim_reduction_more.
    Returns:
        - df_dimR: final reduced feature dataset
        - x_list_df: summary of selected variables and sources
    """
    exclude_cols = [uniq_id, bad, weight, data_subgrp]
    all_x_list = [col for col in df_woe.columns if col not in exclude_cols]

    X = df_woe[all_x_list].fillna(0)
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    # Step 1: PCA for clustering proxy
    pca = PCA(n_components=clus_cnt)
    pca.fit(X_scaled)
    components = pca.components_.T
    cluster_assignments = np.argmax(np.abs(components), axis=1)

    clustering_df = pd.DataFrame({
        "Variable": all_x_list,
        "Cluster": cluster_assignments,
        "RSquareRatio": np.var(X_scaled, axis=0)  # proxy
    })

    # Step 2: pick top N variables from each cluster
    clustering_df = clustering_df.sort_values(["Cluster", "RSquareRatio"], ascending=[True, False])
    clustering_df["cnt"] = clustering_df.groupby("Cluster").cumcount() + 1
    df_cluster_topn = clustering_df[clustering_df["cnt"] <= top_n].copy()
    df_cluster_topn["Source"] = "Cluster"

    # Step 3: candidates not in cluster representatives
    remaining_candidates = list(set(all_x_list) - set(df_cluster_topn["Variable"]))

    def forward_selection(X_df, y, weights, label):
        model = LogisticRegression(max_iter=1000, solver="liblinear")
        model.set_fit_request(sample_weight=True)
        sfs = SequentialFeatureSelector(model,
                                        n_features_to_select=min(forward_stop_n, X_df.shape[1]),
                                        direction="forward",
                                        scoring="neg_log_loss",
                                        cv=3,
                                        n_jobs=-1)
        sfs.fit(X_df, y, sample_weight=weights)
        selected_vars = list(X_df.columns[sfs.get_support()])
        return pd.DataFrame({"Variable": selected_vars, "Source": label})

    # Prepare X/y/weights
    y = df_woe[bad].values
    sample_weight = df_woe[weight].values

    selected_after_clu = forward_selection(df_woe[remaining_candidates].fillna(0), y, sample_weight, "After Clu")
    selected_all = forward_selection(df_woe[all_x_list].fillna(0), y, sample_weight, "All Forward")
    selected_all_by_grp = []
    for grp, df_grp in df_woe.groupby(data_subgrp):
        try:
            sel = forward_selection(df_grp[all_x_list].fillna(0),
                                     df_grp[bad].values,
                                     df_grp[weight].values,
                                     "All by Forward")
            selected_all_by_grp.append(sel)
        except Exception:
            continue
    selected_all_by_grp = pd.concat(selected_all_by_grp, ignore_index=True) if selected_all_by_grp else pd.DataFrame(columns=["Variable", "Source"])

    # Combine
    candidate_all = pd.concat([
        df_cluster_topn[["Variable", "Source"]],
        selected_after_clu,
        selected_all,
        selected_all_by_grp
    ]).drop_duplicates()

    selected_vars = candidate_all["Variable"].unique().tolist()
    df_dimR = df_woe[[uniq_id, bad, weight] + selected_vars].copy()

    # Summary
    summary = pd.DataFrame({"Variable": selected_vars})
    for src in ["Cluster", "After Clu", "All Forward", "All by Forward"]:
        summary[src.replace(" ", "_")] = summary["Variable"].isin(
            candidate_all[candidate_all["Source"] == src]["Variable"]
        ).astype(int)

    return df_dimR, summary


auto_df_dimR, autosel_summary=dr_woe_dim_reduction_more(
    df_woe=auto_model_data2,
    uniq_id= 'uniq_id',
    bad='bad',
    weight='weight',
    data_subgrp='tag',
    clus_cnt = 20,
    top_n= 2,
    sle  = 0.2,
    forward_stop_n  = 18
)

##manual MFC**
temp_mfc_all2 =duckdb.query(f"""
                            select a.* , c.F1  as description   
                            from dfdev_iv  a
                            join  autosel_summary          b on  a.x_nm=b.variable
                            join  RSVP_RISK_NEW_LU c on  a.x_nm =c._x_nm
                            """).to_df()    #copy the data to excel sorting IV largest to smallest and keep over 0.09

temp_mfc_all2a =duckdb.query(f"""
                             select a.* , c.description   
                             from predf_iv  a
                             join  autosel_summary          b on  a.x_nm=b.variable
                             join  PREMIER_ATTR_LU_1203 c on  a.x_nm =c.attr_name
                             """).to_df()  #copy the data to excel sorting IV largest to smallest and keep over 0.09
                             
num_keepc=["r_xnm1901",
"r_xnm428",
"r_xnm15",
"r_xnm845",
"r_xnm1710",
"r_xnm507",
"r_xnm488",
"r_xnm251",
"r_xnm185",
"r_xnm1837",
"r_xnm263",
"r_xnm1785",
"r_xnm229",
"r_xnm694",
"r_xnm1614",
"r_xnm1873",
"r_xnm703",
"r_xnm34",
"r_xnm1640",
"r_xnm164",
"r_xnm1649",
"r_xnm1555",
"r_xnm262",
"r_xnm228",
"r_xnm231",
"r_xnm266",
"r_xnm1601",
"r_xnm1643",
"r_xnm1608",
"r_xnm2523",
"r_xnm1890",
"r_xnm1992",
"r_xnm1828",
"r_xnm712",
"r_xnm1603",
"r_xnm1669",
"r_xnm1512",
"r_xnm230",]
premier_keep=["REV0300",
"BCC3512",
"ALL2001",
"RTR3422",
"BCC5020","IQT9415"]

in_clause = ",".join(f"'{col}'" for col in num_keepc)

stemp_mfc_1a = duckdb.query(f"""
    SELECT *
    FROM dfdev_iv
    WHERE x_nm IN ({in_clause})
    ORDER BY 1
""").to_df()

stemp_mfc_1b = duckdb.query("""
    SELECT b.*
    FROM stemp_mfc_1a a
    LEFT JOIN dfdev_woe b
    ON a.x_nm = b.x_nm
""").to_df()

import os
import pandas as pd
import statsmodels.api as sm
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows

def logit_full_report_to_excel(df_woe, output_excel='logit_full_report.xlsx'):
    os.makedirs("logit_full_report_imgs", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Logit_Report"

    current_row = 1

    for var in df_woe['x_nm'].dropna().unique():
        sub_df = df_woe[df_woe['x_nm'] == var].copy()
        sub_df_clean = sub_df.dropna(subset=['avg_x_value', 'logit_bad_rate', 'sum_wgt'])

        if sub_df_clean.shape[0] < 2:
            continue

        # regression modeling
        X = sm.add_constant(sub_df_clean['avg_x_value'])
        y = sub_df_clean['logit_bad_rate']
        weights = sub_df_clean['sum_wgt']
        model = sm.WLS(y, X, weights=weights).fit()

        # insert variable headers
        ws.cell(row=current_row, column=1, value=f"Variable: {var}")
        current_row += 1

        # Plot Figure
        plt.figure(figsize=(6, 4))
        sns.scatterplot(data=sub_df, x='avg_x_value', y='logit_bad_rate', size='sum_wgt', legend=False)
        plt.plot(sub_df_clean['avg_x_value'], model.predict(X), color='blue', linestyle='--')
        plt.title(f"{var} logit_bad_rate vs avg_x_value")
        plt.xlabel('avg_x_value')
        plt.ylabel('logit_bad_rate')
        plot_path = f"logit_full_report_imgs/{var}.png"
        plt.tight_layout()
        plt.savefig(plot_path)
        plt.close()

        # Insert Image
        img = ExcelImage(plot_path)
        img.anchor = f"A{current_row}"
        ws.add_image(img)
        current_row += 20

        # Insert regression summary information
        ws.cell(row=current_row, column=1, value=f"R-squared: {model.rsquared:.4f}")
        ws.cell(row=current_row + 1, column=1, value=f"Adj R-squared: {model.rsquared_adj:.4f}")
        ws.cell(row=current_row + 2, column=1, value=f"RMSE: {model.mse_resid**0.5:.4f}")
        current_row += 4

        # Insert coefficient table
        ws.cell(row=current_row, column=1, value="Coefficients:")
        coef_df = model.summary2().tables[1].reset_index()
        for row in dataframe_to_rows(coef_df, index=False, header=True):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=current_row, column=col_idx, value=val)
            current_row += 1
        current_row += 1

        # Insert raw data
        ws.cell(row=current_row, column=1, value="Raw Table:")
        current_row += 1
        for row in dataframe_to_rows(sub_df, index=False, header=True):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=current_row, column=col_idx, value=val)
            current_row += 1

        current_row += 3

    wb.save(output_excel)
    print(f"All logit outputs saved to one sheet: {output_excel}")
#call the plots function
logit_full_report_to_excel(stemp_mfc_1b, output_excel=r'D:\HA_study_plan_tasks\banking_financial_scorecard_dev\model_1\mfc\full_logit_report2.xlsx')

in_clause = ",".join(f"'{col}'" for col in premier_keep)

stemp_mfc_2a = duckdb.query(f"""
    SELECT *
    FROM dfdev_iv
    WHERE x_nm IN ({in_clause})
    ORDER BY 1
""").to_df()

stemp_mfc_2b = duckdb.query("""
    SELECT b.*
    FROM stemp_mfc_2a a
    LEFT JOIN dfdev_woe b
    ON a.x_nm = b.x_nm
""").to_df()
logit_full_report_to_excel(stemp_mfc_2b, output_excel=r'D:\HA_study_plan_tasks\banking_financial_scorecard_dev\model_1\mfc\full_logit_report_premier.xlsx')


##check premier  special value bad distribution
type(auto_model_data2['REV0300'])            # output：<class 'pandas.core.series.Series'>
print(auto_model_data2['REV0300'].dtype) 
rev0300_df = duckdb.query("""
    SELECT 
        CASE 
            WHEN REV0300 = 99 THEN 99
            ELSE 8888
        END AS normalvalue,
        COUNT(*) AS cnt,
        SUM(bad * weight) / SUM(weight) AS bad_rate
    FROM df_renamed
    GROUP BY normalvalue
""").to_df()

rev0300_df = duckdb.query("""
    SELECT 
         REV0300,
        COUNT(*) AS cnt,
        SUM(bad * weight) / SUM(weight) AS bad_rate
    FROM df_renamed
    GROUP BY REV0300
""").to_df()

bcc3512_df = duckdb.query("""
    SELECT 
         BCC3512,
        COUNT(*) AS cnt,
        SUM(bad * weight) / SUM(weight) AS bad_rate
    FROM df_renamed
    GROUP BY BCC3512
""").to_df()

ALL2001_df = duckdb.query("""
    SELECT 
         ALL2001,
        COUNT(*) AS cnt,
        SUM(bad * weight) / SUM(weight) AS bad_rate
    FROM df_renamed
    GROUP BY ALL2001
""").to_df()



RTR3422_df = duckdb.query("""
    SELECT 
         RTR3422,
        COUNT(*) AS cnt,
        SUM(bad * weight) / SUM(weight) AS bad_rate
    FROM df_renamed
    GROUP BY RTR3422
""").to_df()


BCC5020_df = duckdb.query("""
    SELECT 
         BCC5020,
        COUNT(*) AS cnt,
        SUM(bad * weight) / SUM(weight) AS bad_rate
    FROM df_renamed
    GROUP BY BCC5020 order by BCC5020
""").to_df()


######final model manully selection 
def mfc(in_x, mrv, floor, cap):
    if pd.isna(in_x):
        return mrv
    elif in_x < floor:
        return floor
    elif in_x > cap:
        return cap
    else:
        return in_x

def mfc2(in_x, mv1, mrv1, mv2, mrv2, floor, cap):
    if in_x == mv1:
        return mrv1
    elif in_x == mv2:
        return mrv2
    else:
        return mfc(in_x, None, floor, cap)

def mfc3(in_x, mv1, mrv1, mv2, mrv2, mv3, mrv3, floor, cap):
    if in_x == mv1:
        return mrv1
    elif in_x == mv2:
        return mrv2
    elif in_x == mv3:
        return mrv3
    else:
        return mfc(in_x, None, floor, cap)

def mfc4(in_x, mv1, mrv1, mv2, mrv2, mv3, mrv3, mv4, mrv4, floor, cap):
    if in_x == mv1:
        return mrv1
    elif in_x == mv2:
        return mrv2
    elif in_x == mv3:
        return mrv3
    elif in_x == mv4:
        return mrv4
    else:
        return mfc(in_x, None, floor, cap)

# Apply to DataFrame df
# Numeric MFC transformations
df_renamed['m_r_xnm1901'] = df_renamed['r_xnm1901'].apply(lambda x: mfc(x, 724.359375, 720, 963))
df_renamed['m_r_xnm428'] = df_renamed['r_xnm428'].apply(lambda x: mfc(x, 869.044487427466, 750, 990))
df_renamed['m_r_xnm15'] = df_renamed['r_xnm15'].apply(lambda x: mfc(x, 1337.74131274131, 0, 1953))
df_renamed['m_r_xnm845'] = df_renamed['r_xnm845'].apply(lambda x: 0 if x == 0 else 1)
df_renamed['m_r_xnm507'] = df_renamed['r_xnm507'].apply(lambda x: 0 if pd.isna(x) else 1)
df_renamed['m_r_xnm251'] = df_renamed['r_xnm251'].apply(lambda x: mfc(x, 6.53018867924528, 0, 7))
df_renamed['m_r_xnm185'] = df_renamed['r_xnm185'].apply(lambda x: 0 if x == 0 else 1)
df_renamed['m_r_xnm1837'] = df_renamed['r_xnm1837'].apply(lambda x: mfc(x, 1.4310355987055, 1, 1.45))
df_renamed['m_r_xnm263'] = df_renamed['r_xnm263'].apply(lambda x: mfc(x, 5.98042084168337, 6, 12))
df_renamed['m_r_xnm1785'] = df_renamed['r_xnm1785'].apply(lambda x: mfc(x, 13.7211457455771, 1, 14))
df_renamed['m_r_xnm229'] = df_renamed['r_xnm229'].apply(lambda x: mfc(x, 12.0031467181467, 5, 13))
df_renamed['m_r_xnm694'] = df_renamed['r_xnm694'].apply(lambda x: 0 if pd.isna(x) else 1)
df_renamed['m_r_xnm1614'] = df_renamed['r_xnm1614'].apply(lambda x: mfc(x, 6.04676258992806, 0, 7))
df_renamed['m_r_xnm1873'] = df_renamed['r_xnm1873'].apply(lambda x: 0.38853 if pd.isna(x) else (1 if x <= 0.45 else 0))
df_renamed['m_r_xnm34'] = df_renamed['r_xnm34'].apply(lambda x: mfc(x, 0, 0, 1.4))
df_renamed['m_r_xnm1640'] = df_renamed['r_xnm1640'].apply(lambda x: 1 if pd.isna(x) else 0)
df_renamed['m_r_xnm164'] = df_renamed['r_xnm164'].apply(lambda x: mfc(x, 0, 0, 2))
df_renamed['m_r_xnm1649'] = df_renamed['r_xnm1649'].apply(lambda x: 1 if pd.isna(x) else 0)
df_renamed['m_r_xnm1555'] = df_renamed['r_xnm1555'].apply(lambda x: mfc(x, 1.45465393794749, 0, 9))
df_renamed['m_r_xnm262'] = df_renamed['r_xnm262'].apply(lambda x: 0 if pd.isna(x) else 1)
df_renamed['m_r_xnm228'] = df_renamed['r_xnm228'].apply(lambda x: mfc(x, 10.8292857142857, 4, 11))
df_renamed['m_r_xnm266'] = df_renamed['r_xnm266'].apply(lambda x: 0 if pd.isna(x) else 1)
df_renamed['m_r_xnm1601'] = df_renamed['r_xnm1601'].apply(lambda x: 0 if x == 1 else 1)
df_renamed['m_r_xnm1643'] = df_renamed['r_xnm1643'].apply(lambda x: mfc(x, 17.8313008130081, 0, 200))
df_renamed['m_r_xnm1608'] = df_renamed['r_xnm1608'].apply(lambda x: mfc(x, 0, 0, 2))
df_renamed['m_r_xnm2523'] = df_renamed['r_xnm2523'].apply(lambda x: mfc(x, 717.088353413655, 1, 1600))
df_renamed['m_r_xnm1890'] = df_renamed['r_xnm1890'].apply(lambda x: mfc(x, 0.448526912181303, 0.21, 1))
df_renamed['m_r_xnm1992'] = df_renamed['r_xnm1992'].apply(lambda x: mfc(x, 1.35898496240602, 1, 1.5))
df_renamed['m_r_xnm1828'] = df_renamed['r_xnm1828'].apply(lambda x: mfc(x, 7.14592933947773, 1, 8))
df_renamed['m_r_xnm712'] = df_renamed['r_xnm712'].apply(lambda x: mfc(x, 5.9265873015873, 0, 6))
df_renamed['m_r_xnm1603'] = df_renamed['r_xnm1603'].apply(lambda x: mfc(x, 0, 0, 2))
df_renamed['m_r_xnm1669'] = df_renamed['r_xnm1669'].apply(lambda x: mfc(x, 2.48137254901961, 1, 3))
df_renamed['m_r_xnm1512'] = df_renamed['r_xnm1512'].apply(lambda x: mfc(x, 67.289124668435, 0, 10000))
df_renamed['m_r_xnm230'] = df_renamed['r_xnm230'].apply(lambda x: mfc(x, 1.25740458015267, 0, 1.3))
df_renamed['m_REV0300'] = df_renamed['REV0300'].apply(lambda x: mfc(x, 0, 0, 18))
df_renamed['m_IQT9415'] = df_renamed['IQT9415'].apply(lambda x: mfc(x, 2.16644474034621, 0, 3))
df_renamed['m_ALL2001'] = df_renamed['ALL2001'].apply(lambda x: mfc(x, 1, 0, 5))
df_renamed['m_BCC3512'] = df_renamed['BCC3512'].apply(lambda x: mfc3(x, 97, 3, 98, 3, 99, 3, 0, 50))
df_renamed['m_RTR3422'] = df_renamed['RTR3422'].apply(lambda x: mfc3(x, 97, 0, 98, 8, 99, 8, 0, 90))
df_renamed['m_BCC5020'] = df_renamed['BCC5020'].apply(lambda x: mfc4(x, 999999996, 0, 999999997, 277, 999999998, 9991, 999999999, 277, 0, 152336))

# Character WOE mappings
df_renamed['m_Carrier'] = df_renamed['Carrier'].map({
    'CellCom': 0.6924716418,
    'Comcast': 0.6924716418,
    'SouthernLINC': 0.6924716418,
    'Pinger/Bandwidth.com (Sybase)': 0.6924716418,
    'TracFone (AT&T)': 0.6924716418,
    'United States Cellular Corp': 0.6924716418,
    'Unknown': 0.1784753742,
    'Sprint': 0.1784753742,
    'Verizon': 0.087095077,
    'AT&T': -0.007380284,
    'Aio Wireless': -0.226911945,
    'Google Voice': -0.226911945,
    'T-Mobile': -0.226911945,
    'MetroPCS/T-Mobile US': -0.731249125,
    '': -0.731249125,
    'C Spire Wireless (aka Cellular South)': -0.731249125,
    'Republic Wireless/Bandwidth.com (Sybase)': -0.731249125,
    'bandwidth.com': -0.731249125
}).fillna(-0.731249125)

df_renamed['m_IncomeSource'] = df_renamed['IncomeSource'].map({
    'Job': 0.109456169,
    'Retirement/Benefits': -0.232054318,
    'Self Employed': -0.232054318,
    'Social Security/Disability': -0.627773624,
    'Social Security': -0.627773624,
    'Unemployment/Other Public Assistance': -0.896763563,
    'Disability': -0.896763563
}).fillna(-0.896763563)

df_renamed['m_PaydayMethod'] = df_renamed['PaydayMethod'].map({
    'Every 2 weeks': 0.280064472,
    'Weekly': -0.021294826,
    'Twice Monthly': -0.021294826,
    'Monthly': -0.500032135
}).fillna(-0.500032135)

# Flag variables (boolean/text to binary)
df_renamed['m_r_xnm1622'] = df_renamed['r_xnm1622'].apply(lambda x: 0 if x == 'TRUE' else 1)
df_renamed['m_r_xnm1662'] = df_renamed['r_xnm1662'].apply(lambda x: 0 if x == 'TRUE' else 1)
df_renamed['m_r_xnm2002'] = df_renamed['r_xnm2002'].apply(lambda x: 0 if x == 'TRUE' else 1)
df_renamed['m_r_xnm21'] = df_renamed['r_xnm21'].apply(lambda x: 1 if x == 'TRUE' else 0)
df_renamed['m_r_xnm23'] = df_renamed['r_xnm23'].apply(lambda x: 0 if x == 'TRUE' else 1)
df_renamed['m_r_xnm2533'] = df_renamed['r_xnm2533'].apply(lambda x: 1 if x == 'TRUE' else 0)
df_renamed['m_r_xnm2534'] = df_renamed['r_xnm2534'].apply(lambda x: 1 if x == 'TRUE' else 0)
df_renamed['m_r_xnm6540'] = df_renamed['r_xnm6540'].apply(lambda x: 1 if x == 'TRUE' else 0)
df_renamed['m_r_xnm6542'] = df_renamed['r_xnm6542'].apply(lambda x: 1 if x == 'TRUE' else 0)

# Columns to keep
keep_cols = [
    'uniq_id', 'weight', 'bad', 'camp',
    'm_r_xnm1622', 'm_r_xnm1662', 'm_r_xnm2002', 'm_r_xnm21', 'm_r_xnm23',
    'm_r_xnm2533', 'm_r_xnm2534', 'm_r_xnm6540', 'm_r_xnm6542',
    'm_Carrier', 'm_IncomeSource', 'm_PaydayMethod',
    'm_r_xnm1901', 'm_r_xnm428', 'm_r_xnm15', 'm_r_xnm845', 'm_r_xnm507',
    'm_r_xnm251', 'm_r_xnm185', 'm_r_xnm1837', 'm_r_xnm263', 'm_r_xnm1785',
    'm_r_xnm229', 'm_r_xnm694', 'm_r_xnm1614', 'm_r_xnm1873', 'm_r_xnm34',
    'm_r_xnm1640', 'm_r_xnm164', 'm_r_xnm1649', 'm_r_xnm1555', 'm_r_xnm262',
    'm_r_xnm228', 'm_r_xnm266', 'm_r_xnm1601', 'm_r_xnm1643', 'm_r_xnm1608',
    'm_r_xnm2523', 'm_r_xnm1890', 'm_r_xnm1992', 'm_r_xnm1828', 'm_r_xnm712',
    'm_r_xnm1603', 'm_r_xnm1669', 'm_r_xnm1512', 'm_r_xnm230',
    'm_REV0300', 'm_IQT9415', 'm_ALL2001', 'm_BCC3512',
    'm_RTR3422', 'm_BCC5020'
]

df_renamed_final = df_renamed[keep_cols].copy()

sel_model_dev=duckdb.query(f"""select * from df_renamed_final where uniq_id in (select uniq_id from model_dev)""").to_df()

remove = ['camp']

sel_model_dev = sel_model_dev[[col for col in keep_cols if col not in remove]]                      
sel_model_dev_long = pd.melt(sel_model_dev, 
                  id_vars=[ 'uniq_id', 'bad', 'weight'],
                 
                  var_name='x_nm',
                  value_name='x_value')

                          
selout_data_rbin_sql, selout_num_rbins = DR_num_rbin_duckdb(
    df_input=sel_model_dev_long,  # including x_nm, x_value, weight, bad etc. columns
    uniq_id='uniq_id',
    bad='bad',
    weight='weight',
    max_bin=10,
    min_adj_bin=5
)
sel_mfc_woe, sel_mfc_iv = DR_num_iv_woe_duckdb(selout_data_rbin_sql)       

#Import From mfc Excel    
sel_mfc_lu= pd.read_excel(r"D:\HA_study_plan_tasks\banking_financial_scorecard_dev\model_1\mfc\MFC.xlsx", sheet_name="mfc_LU")

sel_rsvp_LU = """
    SELECT DISTINCT a.x_nm AS var_name,
           a.mm_x_nm,
           a."Trans./Flag" AS imputation,
           CASE WHEN c.Beta > 0 THEN '+' ELSE '-' END AS marg_sign,
           b.description AS descrip
    FROM sel_mfc_lu a
    INNER JOIN PREMIER_ATTR_LU_1203 b
        ON TRIM(a.x_nm) = TRIM(b.attr_name)
    LEFT JOIN sel_mfc_iv c
        ON a.mm_x_nm = c.x_nm

    UNION ALL

    SELECT DISTINCT a.x_nm AS var_name,
           a.mm_x_nm,
           a."Trans./Flag" AS imputation,
           CASE WHEN c.Beta > 0 THEN '+' ELSE '-' END AS marg_sign,
           b.f1 AS descrip
    FROM sel_mfc_lu a
    INNER JOIN RSVP_RISK_NEW_LU b
        ON TRIM(a.x_nm) = TRIM(b._x_nm)
    LEFT JOIN sel_mfc_iv c
        ON a.mm_x_nm = c.x_nm
"""

# excute and change to DataFrame
sel_rsvp_LU = duckdb.query(sel_rsvp_LU).to_df()


sel_sp_LU=duckdb.query(f"""
select distinct(a.x_nm) as var_name, a.mm_x_nm, a."Trans./Flag" as imputation, 
       case when c.Beta>0 then '+' else '-' end as marg_sign, '' as descrip 
from sel_mfc_lu a  
left join sel_mfc_iv        c  on a.mm_x_nm    =c.x_nm  
where a.x_nm in 
('Carrier',
'IncomeSource',
'PaydayMethod')""").to_df()

sel_revp_lu_all =duckdb.query(f"""select * from sel_rsvp_LU union all
                              select * from sel_sp_LU""").to_df()
                              
#model selection 
m_list=[
 'm_r_xnm1622', 'm_r_xnm1662', 'm_r_xnm2002', 'm_r_xnm21', 'm_r_xnm23',
 'm_r_xnm2533', 'm_r_xnm2534', 'm_r_xnm6540', 'm_r_xnm6542',
 'm_Carrier', 'm_IncomeSource', 'm_PaydayMethod',
 'm_r_xnm1901', 'm_r_xnm428', 'm_r_xnm15', 'm_r_xnm845', 'm_r_xnm507',
 'm_r_xnm251', 'm_r_xnm185', 'm_r_xnm1837', 'm_r_xnm263', 'm_r_xnm1785',
 'm_r_xnm229', 'm_r_xnm694', 'm_r_xnm1614', 'm_r_xnm1873', 'm_r_xnm34',
 'm_r_xnm1640', 'm_r_xnm164', 'm_r_xnm1649', 'm_r_xnm1555', 'm_r_xnm262',
 'm_r_xnm228', 'm_r_xnm266', 'm_r_xnm1601', 'm_r_xnm1643', 'm_r_xnm1608',
 'm_r_xnm2523', 'm_r_xnm1890', 'm_r_xnm1992', 'm_r_xnm1828', 'm_r_xnm712',
 'm_r_xnm1603', 'm_r_xnm1669', 'm_r_xnm1512', 'm_r_xnm230',
 'm_REV0300', 'm_IQT9415', 'm_ALL2001', 'm_BCC3512',
 'm_RTR3422', 'm_BCC5020'] #  the list will be adjusted manully after each cycle

#****************************selection criteria  **********************
# [1] Statistically significant
# [3] Marg_sign=model_sign
# [4] vif < 2.0
# [5] Business reasonable

#seed variable set
from statsmodels.stats.outliers_influence import variance_inflation_factor
def forward_logistic_selection(data, target, candidate_vars, weight_col=None, max_vars=10):
    selected = []
    remaining = candidate_vars.copy()
    y = data[target]
    weights = data[weight_col] if weight_col else None

    while len(selected) < max_vars and remaining:
        best_score = -float('inf')
        best_var = None
        for var in remaining:
            try:
                X = sm.add_constant(data[selected + [var]])
                model = sm.Logit(y, X, weights=weights).fit(disp=False)
                score = model.llf  # log-likelihood
                if score > best_score:
                    best_score = score
                    best_var = var
            except:
                continue
        if best_var is not None:
            selected.append(best_var)
            remaining.remove(best_var)
        else:
            break
    return selected


selected_vars = forward_logistic_selection(sel_model_dev, 'bad', m_list, weight_col='weight', max_vars=10)

X = sm.add_constant(sel_model_dev[selected_vars])
y = sel_model_dev['bad']
weights = sel_model_dev['weight']

logit_model = sm.Logit(y, X, weights=weights).fit()
sel_model_dev['pred'] = logit_model.predict(X)

# equal to ODS OUTPUT parameterestimates
sel_params = pd.DataFrame({
    'variable': logit_model.params.index,
    'estimate': logit_model.params.values,
    'std_err': logit_model.bse.values,
    'waldchisq': (logit_model.params / logit_model.bse) ** 2,
    'probchisq': logit_model.pvalues.values
})

duckdb.register("sel_params", sel_params)

vif_x_list = duckdb.query("""
    SELECT DISTINCT variable
    FROM sel_params
    WHERE variable != 'const'
""").df()['variable'].tolist()

X_vif = sm.add_constant(sel_model_dev[vif_x_list])

vif_data = pd.DataFrame()
vif_data["variable"] = X_vif.columns
vif_data["VIF"] = [variance_inflation_factor(X_vif.values, i) for i in range(X_vif.shape[1])]
sel_vif = vif_data[vif_data["variable"] != "const"]

sel_pred = sel_model_dev[['pred'] + vif_x_list + ['bad', 'weight','uniq_id']]

#weighted KS for checking lift KS of final score or key var **
def weighted_ks(in_data_pred: pd.DataFrame, bad: str, weight: str, pred: str) -> pd.DataFrame:
    # Step 1: add good_wgt / bad_wgt and sort by pred (DuckDB)
    duckdb.register("in_data_pred", in_data_pred)

    tempwh_ks_0 = duckdb.query(f"""
        SELECT *, 
               CASE WHEN {bad} = 1.0 THEN {weight} ELSE 0.0 END AS bad_wgt,
               CASE WHEN {bad} = 0.0 THEN {weight} ELSE 0.0 END AS good_wgt
        FROM in_data_pred
        ORDER BY {pred}
    """).to_df()

    # Step 2: cumulative weighted sum
    tempwh_ks_0["cum_bad_wgt"] = tempwh_ks_0["bad_wgt"].cumsum()
    tempwh_ks_0["cum_good_wgt"] = tempwh_ks_0["good_wgt"].cumsum()

    # Step 3: total weight
    total_bad_wgt = tempwh_ks_0["bad_wgt"].sum()
    total_good_wgt = tempwh_ks_0["good_wgt"].sum()

    tempwh_ks_0["bad_cdf"] = tempwh_ks_0["cum_bad_wgt"] / total_bad_wgt if total_bad_wgt > 0 else 0
    tempwh_ks_0["good_cdf"] = tempwh_ks_0["cum_good_wgt"] / total_good_wgt if total_good_wgt > 0 else 0

    # Step 4: calculate KS
    tempwh_ks_0["ks_gap"] = abs(tempwh_ks_0["bad_cdf"] - tempwh_ks_0["good_cdf"])
    ks_value = tempwh_ks_0["ks_gap"].max() * 100

    out_KS = pd.DataFrame({
        "variable": [f"KS of {pred}"],
        "KS": [ks_value]
    })

    return out_KS
#call it
sel_ks = weighted_ks(in_data_pred=sel_pred, bad="bad", weight="weight", pred="pred")

duckdb.register("sel_params", sel_params)
duckdb.register("sel_vif", sel_vif)
duckdb.register("sel_rsvplu2", sel_revp_lu_all)

temp_sel_out_O = duckdb.query("""
    SELECT a.variable,
           a.estimate,
           a.waldchisq,
           a.probchisq,
           b.VIF AS vif,
           c.imputation,
           CASE WHEN a.estimate > 0 THEN '+' ELSE '-' END AS model_sign,
           c.marg_sign,
           c.descrip,
           c.var_name
    FROM sel_params a
    LEFT JOIN sel_vif b
        ON a.variable = b.variable
    LEFT JOIN sel_rsvplu2 c
        ON a.variable = TRIM(c.mm_x_nm)
    ORDER BY waldchisq DESC
""").to_df()

#model lift table generation**
def weighted_lift_ks_roc(model_desc, in_data, indepvar, bad, weight, bin_size, descending_yn=1):
    # register data to DuckDB
    duckdb.register("in_data", in_data)

    sort_order = "DESC" if descending_yn == 1 else "ASC"

    # Step 1: Group by score → bad_wgt / good_wgt
    tempwh_xks_0s = duckdb.query(f"""
        SELECT {indepvar},
               SUM(CASE WHEN {bad}=1.0 THEN {weight} ELSE 0.0 END) AS bad_wgt,
               SUM(CASE WHEN {bad}=0.0 THEN {weight} ELSE 0.0 END) AS good_wgt
        FROM in_data
        WHERE {weight} IS NOT NULL AND {bad} IS NOT NULL AND {indepvar} IS NOT NULL
        GROUP BY {indepvar}
        ORDER BY {indepvar} {sort_order}
    """).to_df()

    # Step 2: Add total stats
    tempwh_ks_roc_0 = tempwh_xks_0s.copy()
    tempwh_ks_roc_0["tot_wgt"] = tempwh_ks_roc_0["bad_wgt"] + tempwh_ks_roc_0["good_wgt"]
    tot_bad_wgt = tempwh_ks_roc_0["bad_wgt"].sum()
    tot_good_wgt = tempwh_ks_roc_0["good_wgt"].sum()
    tot_wgt = tot_bad_wgt + tot_good_wgt

    # Step 3: Cumulative & binning
    df = tempwh_ks_roc_0.copy()
    df["cum_wgt"] = (df["bad_wgt"] + df["good_wgt"]).cumsum()
    df["cum_bad_wgt"] = df["bad_wgt"].cumsum()
    df["cum_good_wgt"] = df["good_wgt"].cumsum()
    bin_w = tot_wgt / bin_size
    df["bin"] = 1 + (df["cum_wgt"] / (1.0 + bin_w)).apply(np.floor).astype(int)

    # Step 4: KS
    df["bad_cdf"] = df["cum_bad_wgt"] / tot_bad_wgt
    df["good_cdf"] = df["cum_good_wgt"] / tot_good_wgt
    df["ks_gap"] = abs(df["bad_cdf"] - df["good_cdf"])
    ks_value = df["ks_gap"].max() * 100

    # Step 5: ROC - sensitivity vs 1-specificity
    df["id"] = range(1, len(df)+1)
    df["sensitivity"] = (tot_bad_wgt - df["cum_bad_wgt"]) / tot_bad_wgt
    df["specifity"] = df["cum_good_wgt"] / tot_good_wgt
    df["x"] = 1.0 - df["specifity"]
    df["y"] = df["sensitivity"]

    df["area"] = (df["x"].diff().fillna(0)) * (df["y"] + df["y"].shift(1).fillna(0)) / 2
    roc_auc = df["area"].sum()

    # Step 6: binning metrics output
    grouped = df.groupby("bin").agg(
        min_indepvar=(indepvar, "min"),
        max_indepvar=(indepvar, "max"),
        avg_indepvar=(indepvar, lambda x: np.average(x, weights=(df.loc[x.index, "bad_wgt"] + df.loc[x.index, "good_wgt"]))),
        marg_all=("cum_wgt", "max"),
        marg_bad=("cum_bad_wgt", "max"),
        marg_good=("cum_good_wgt", "max"),
        cum_wgt=("cum_wgt", "max"),
        cum_bad_wgt=("cum_bad_wgt", "max"),
        cum_good_wgt=("cum_good_wgt", "max"),
        one_minus_specifity=("x", "max"),
        sensitivity=("y", "max")
    ).reset_index()

    grouped["marg_dist_all_rate"] = grouped["marg_all"] / tot_wgt
    grouped["marg_dist_bad_rate"] = grouped["marg_bad"] / tot_bad_wgt
    grouped["marg_dist_good_rate"] = grouped["marg_good"] / tot_good_wgt
    grouped["marg_bad_rate"] = grouped["marg_bad"] / grouped["marg_all"]
    grouped["woe"] = np.log(grouped["marg_dist_bad_rate"] / grouped["marg_dist_good_rate"])
    grouped["cum_dist_all_rate"] = grouped["cum_wgt"] / tot_wgt
    grouped["cum_dist_bad_rate"] = grouped["cum_bad_wgt"] / tot_bad_wgt
    grouped["cum_dist_good_rate"] = grouped["cum_good_wgt"] / tot_good_wgt
    grouped["cum_bad_rate"] = grouped["cum_bad_wgt"] / grouped["cum_wgt"]

    out_ks_roc = pd.DataFrame({
        "model": [model_desc],
        "indepvar_order": [sort_order.lower()],
        "KS": [ks_value],
        "ROC": [roc_auc],
        "tot_wgt": [tot_wgt],
        "tot_bad_rate": [tot_bad_wgt / tot_wgt]
    })

    return out_ks_roc, grouped

def plot_roc(df_bins):
    df_plot = pd.concat([
        pd.DataFrame({"sensitivity": [0], "one_minus_specifity": [0]}),
        df_bins[["sensitivity", "one_minus_specifity"]]
    ])
    plt.plot(df_plot["one_minus_specifity"], df_plot["sensitivity"], marker="o")
    plt.title("ROC Curve")
    plt.xlabel("1 - Specificity")
    plt.ylabel("Sensitivity")
    plt.grid(True)
    plt.show()

def plot_lorenz(df_bins):
    df_plot = pd.concat([
        pd.DataFrame({"cum_dist_bad_rate": [0], "cum_dist_all_rate": [0]}),
        df_bins[["cum_dist_bad_rate", "cum_dist_all_rate"]]
    ])
    plt.plot(df_plot["cum_dist_all_rate"], df_plot["cum_dist_bad_rate"], marker="o", color="orange")
    plt.title("Lorenz Curve")
    plt.xlabel("Cumulative Distribution of Population")
    plt.ylabel("Cumulative Distribution of Bad")
    plt.grid(True)
    plt.show()
    
sel_out_ks_roc, sel_out_bins = weighted_lift_ks_roc(
    model_desc="Model A",
    in_data=sel_pred,
    indepvar="pred",
    bad="bad",
    weight="weight",
    bin_size=10,
    descending_yn=1
)

print(sel_out_ks_roc)
print(sel_out_bins.head())

plot_roc(sel_out_bins)
plot_lorenz(sel_out_bins)

##manually sel each time add one var 
sel_x_list=["m_r_xnm1901",
"m_PaydayMethod",
"m_r_xnm845",
"m_Carrier",
"m_r_xnm428",
"m_r_xnm2523",
"m_r_xnm1785",
"m_r_xnm507",
"m_r_xnm1890",
"m_r_xnm263",
]



import statsmodels.api as sm
import duckdb
from statsmodels.stats.outliers_influence import variance_inflation_factor

def sel_addone_bin(
    model_data: pd.DataFrame,
    uniq_id: str,
    bad: int,
    weight: int,
    bin_cnt: int,
    m_x_list: list,
    sel_x_list: list,
):
    # 1. temp_addOne_1: keep uniq_id, weight, bad, m_x_list; drop sel_x_list
    cols_to_keep = [uniq_id, weight, bad] + m_x_list
    temp_addOne_1 = model_data[cols_to_keep].copy()
    temp_addOne_1.drop(columns=sel_x_list, errors='ignore', inplace=True)

    # 2. sort by uniq_id, weight, bad (mimic proc sort)
    temp_addOne_1 = temp_addOne_1.sort_values(by=[uniq_id, weight, bad])

    # 3. transpose: convert from wide to long
    # SAS proc transpose by uniq_id, weight, bad, _name_ to x_nm, values to x_value
    temp_addOne_2 = (
        temp_addOne_1
        .set_index([uniq_id, weight, bad])
        .stack()
        .reset_index()
        .rename(columns={'level_3': 'x_nm', 0: 'x_value'})
    )

    # 4. temp_addOne_3: keep uniq_id, weight, bad, sel_x_list from original model_data
    temp_addOne_3 = model_data[[uniq_id, weight, bad] + sel_x_list].copy()

    # 5. SQL join temp_addOne_3 with temp_addOne_2 on uniq_id and bad (left join)
    con = duckdb.connect(database=':memory:')
    con.register('temp_addOne_3', temp_addOne_3)
    con.register('temp_addOne_2', temp_addOne_2)

    temp_addOne_4 = con.query(f"""
        SELECT b.x_nm, b.x_value, a.*
        FROM temp_addOne_3 a
        LEFT JOIN temp_addOne_2 b
          ON a.{uniq_id} = b.{uniq_id} AND a.{bad} = b.{bad}
        ORDER BY x_nm, a.{uniq_id}, a.{weight}, a.{bad}
    """).to_df()

    # 6. logistic regression for each x_nm group: bad ~ sel_x_list + x_value, weighted by weight
    # statistical paremeters，vif，predicted probabilities
    all_results = []
    all_vif = []
    all_preds = []

    for x_nm, group in temp_addOne_4.groupby('x_nm'):
        # dropna for needed columns
        cols_for_model = sel_x_list + ['x_value']
        temp_df = group[[bad, weight] + cols_for_model].dropna()

        # skip groups with too few samples
        if temp_df.shape[0] < 10:
            continue

        # independent variable X, add intercept term
        X = temp_df[cols_for_model]
        X = sm.add_constant(X, has_constant='add')

        y = temp_df[bad]
        w = temp_df[weight]

        try:
            # weighted logitistic model
            logit_model = sm.Logit(y, X, freq_weights=w).fit(disp=0)

            # parameters estimation
            params = logit_model.params.to_frame('Estimate').reset_index().rename(columns={'index': 'Variable'})

            # Wald Chi-square
            wald_chi = (logit_model.params / logit_model.bse)**2
            from scipy.stats import chi2
            prob_chi = chi2.sf(wald_chi, df=1)

            params['WaldChiSq'] = wald_chi.values
            params['ProbChiSq'] = prob_chi
            params['x_nm'] = x_nm

            # VIF calculation
            vif_data = pd.DataFrame()
            if len(cols_for_model) > 1:
                X_vif = temp_df[cols_for_model].copy()
                # calculate VIF for each variable
                vif_data = pd.DataFrame()
                vif_data['Variable'] = X_vif.columns
                vif_data['VarianceInflation'] = [variance_inflation_factor(X_vif.values, i) for i in range(X_vif.shape[1])]
            else:
                # When only one variable x_value，set VIF to nan
                vif_data = pd.DataFrame({'Variable': cols_for_model, 'VarianceInflation': [np.nan]*len(cols_for_model)})

            vif_data['x_nm'] = x_nm

            all_results.append(params)
            all_vif.append(vif_data)

            # predict probabilities
            pred = logit_model.predict(X)
            pred_df = temp_df[[bad, weight]].copy()
            pred_df['pred'] = pred
            pred_df['x_nm'] = x_nm

            all_preds.append(pred_df)

        except Exception as e:
            # skip for error
            print(f"Warning: logit failed for variable {x_nm}: {e}")
            continue

    # merge results
    res_params = pd.concat(all_results, ignore_index=True)
    res_vif = pd.concat(all_vif, ignore_index=True)
    pred_all = pd.concat(all_preds, ignore_index=True)

    # 7. Calculate KS (cumulative weighted difference between bad and good) by sorting cumulatively according to x_nm and pred
    con.register('temp_pred', pred_all)
    ks_df = con.query(f"""
        WITH ordered AS (
            SELECT
                x_nm, pred,
                SUM(CASE WHEN {bad} = 1 THEN {weight} ELSE 0 END) AS bad_wgt,
                SUM(CASE WHEN {bad} = 0 THEN {weight} ELSE 0 END) AS good_wgt
            FROM temp_pred
            GROUP BY x_nm, pred
        ),
        cum AS (
            SELECT
                x_nm, pred,
                SUM(bad_wgt) OVER (PARTITION BY x_nm ORDER BY pred) AS cum_bad_wgt,
                SUM(good_wgt) OVER (PARTITION BY x_nm ORDER BY pred) AS cum_good_wgt,
                SUM(bad_wgt) OVER (PARTITION BY x_nm) AS tot_bad_wgt,
                SUM(good_wgt) OVER (PARTITION BY x_nm) AS tot_good_wgt
            FROM ordered
        )
        SELECT
            x_nm,
            100 * MAX(ABS(cum_bad_wgt / tot_bad_wgt - cum_good_wgt / tot_good_wgt)) AS KS
        FROM cum
        GROUP BY x_nm
        ORDER BY KS DESC
    """).to_df()

    # 8. calculate bin_bad_pct for each bin
    # calculate each record’s weight and cumulative sum after sorting
    con.register('pred_bin', pred_all)
    bin_df = con.query(f"""
        SELECT
            x_nm, pred, {bad}, SUM({weight}) AS wgt
        FROM pred_bin
        GROUP BY x_nm, pred, {bad}
    """).to_df()

    # calcualte total weight by x_nm
    total_wgt = bin_df.groupby('x_nm')['wgt'].sum().rename('tot_wgt').reset_index()
    bin_df = bin_df.merge(total_wgt, on='x_nm', how='left')

    # calculate cumulative weights by x_nm, pred in descending order
    bin_df = bin_df.sort_values(['x_nm', 'pred'], ascending=[True, False])
    bin_df['cum_wgt'] = bin_df.groupby('x_nm')['wgt'].cumsum()

    # calculate width for each bin
    bin_df['bin_w'] = np.floor(0.5 + bin_df['tot_wgt'] / bin_cnt)
    bin_df['bin'] = 1 + np.floor(bin_df['cum_wgt'] / (1.0 + bin_df['bin_w']))

    # calculate bin_bad_pct
    bin_bad_pct_df = (
        bin_df.groupby(['x_nm', 'bin'])
        .apply(lambda x: np.sum(x['wgt'] * x[bad]) / np.sum(x['wgt']))
        .reset_index(name='bin_bad_pct')
    )

    # 9. transport bin_bad_pct，id=bin，row is x_nm，columns are bin_1, bin_2 ...
    bin_bad_pct_wide = bin_bad_pct_df.pivot(index='x_nm', columns='bin', values='bin_bad_pct')
    bin_bad_pct_wide.columns = [f'bin_{int(c)}' for c in bin_bad_pct_wide.columns]
    bin_bad_pct_wide = bin_bad_pct_wide.reset_index()

    # 10. summarize all results：merge KS，parameters，VIF，bin_bad_pct
    # add a flag int_yn in the coefficient table to mark the Intercept
    res_params['int_yn'] = res_params['Variable'].apply(lambda x: 1 if x.lower() == 'const' or x.lower() == 'intercept' else 0)

    # merge all tables
    final_df = (
        res_params
        .merge(res_vif, on=['x_nm', 'Variable'], how='left')
        .merge(ks_df, on='x_nm', how='left')
        .merge(bin_bad_pct_wide, on='x_nm', how='left')
    )

    # sorting by ks，x_nm，int_yn，WaldChiSq in descending order
    final_df = final_df.sort_values(by=['KS', 'x_nm', 'int_yn', 'WaldChiSq'], ascending=[False, True, False, False])

    # drop int_yn
    final_df = final_df.drop(columns=['int_yn'])

    # close connection
    con.close()

    return final_df.reset_index(drop=True)

result_df = sel_addone_bin(
    model_data=sel_model_dev,  # DataFrame
    uniq_id="uniq_id",         # 
    bad="bad",                 # 
    weight="weight",           # 
    bin_cnt=10,                # 
    m_x_list=m_list,           # 
    sel_x_list=sel_x_list      # 
)


duckdb.register("temp_addone_out", result_df)
duckdb.register("rsvplu2", sel_revp_lu_all)

temp_sel_out = duckdb.query("""
    SELECT 
        a.x_nm, a.KS, a.Variable, a.Estimate, a.WaldChiSq, a.ProbChiSq, 
        a.bin_1, a.bin_2, a.bin_10,
        CASE WHEN a.Estimate > 0 THEN '+' ELSE '-' END AS model_sign,
        CASE WHEN c.marg_sign IS NOT NULL THEN c.marg_sign ELSE d.marg_sign END AS marg_sign,
        CASE WHEN c.descrip IS NOT NULL THEN c.descrip ELSE d.descrip END AS description
    FROM temp_addone_out a
    LEFT JOIN rsvplu2 c ON a.Variable = TRIM(c.mm_x_nm)
    LEFT JOIN rsvplu2 d ON a.x_nm = TRIM(d.mm_x_nm)
    ORDER BY a.KS DESC
""").to_df()


#####scorecard to OOT /insample validation data ** then check the KS and lift ROC etc**

sel_model_oot=duckdb.query(f"""select * from df_renamed_final where uniq_id in (select uniq_id from model_OOT)""").to_df()
val = sel_model_oot

# calculate xbeta output the temp_sel_out_O to excel then using the formula below to cal xbeta and pred for checking the KS and lift regarding the OOT data
val['xbeta'] = (
    val['m_PaydayMethod'] * -1.139401345 +
    1 * 5.551720464 +                                                         # intercept
    val['m_r_xnm1901']*(-0.00430408838209394) +
val['m_Carrier']*(-0.880389990526734) +
val['m_r_xnm2523']*(-0.000678811613440738) +
val['m_r_xnm428']*(-0.00305180330290924) +
val['m_r_xnm1785']*(0.0316676337785248) +
val['m_r_xnm845']*(0.370308152700866) +
val['m_r_xnm507']*(0.420487365131244) +
val['m_r_xnm1890']*(-0.564449313264391) +
val['m_r_xnm263']*(0.061204768225097) )

# calculate predicted probabilities
val['pred'] = np.exp(val['xbeta']) / (1 + np.exp(val['xbeta']))

#OOT validation KS
oot_out_ks_roc, oot_out_bins = weighted_lift_ks_roc(
    model_desc="Model OOT",
    in_data=val,
    indepvar="pred",
    bad="bad",
    weight="weight",
    bin_size=10,
    descending_yn=1
)

print(oot_out_ks_roc)
print(oot_out_bins.head())

plot_roc(oot_out_bins)
plot_lorenz(oot_out_bins)
oot_ks = weighted_ks(in_data_pred=val, bad="bad", weight="weight", pred="pred")
