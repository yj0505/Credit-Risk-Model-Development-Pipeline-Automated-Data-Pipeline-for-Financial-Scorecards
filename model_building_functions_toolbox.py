


#function A:
def DR_x_weighted_KS_pandas(in_data_df, bad, weight, out_name='x_ks_result'):
    df = in_data_df.copy()
    
    # Step 1: Group by x_nm and x_value
    grouped = df.groupby(['x_nm', 'x_value']).apply(
        lambda g: pd.Series({
            'bad_wgt': g.loc[g[bad] == 1, weight].sum(),
            'good_wgt': g.loc[g[bad] == 0, weight].sum()
        })
    ).reset_index()
    
    # Steps 2-5 (same as before)
    grouped['cum_bad_wgt'] = grouped.groupby('x_nm')['bad_wgt'].cumsum()
    grouped['cum_good_wgt'] = grouped.groupby('x_nm')['good_wgt'].cumsum()
    
    totals = grouped.groupby('x_nm')[['bad_wgt', 'good_wgt']].sum().reset_index()
    totals.columns = ['x_nm', 'tot_bad_wgt', 'tot_good_wgt']
    merged = grouped.merge(totals, on='x_nm', how='left')
    
    merged['bad_cdf'] = merged['cum_bad_wgt'] / merged['tot_bad_wgt']
    merged['good_cdf'] = merged['cum_good_wgt'] / merged['tot_good_wgt']
    merged['ks_diff'] = (merged['bad_cdf'] - merged['good_cdf']).abs()
    
    out_x_ks = merged.groupby('x_nm')['ks_diff'].max().reset_index()
    out_x_ks['xKS'] = 100 * out_x_ks['ks_diff']
    out_x_ks = out_x_ks[['x_nm', 'xKS']].sort_values('xKS', ascending=False)
    
    globals()[out_name] = out_x_ks
    return out_x_ks     
       
# call the function as above   temp_2 is your long-format DataFrame with columns: ['x_nm', 'x_value', 'bad', 'weight'] 
x_ks_df = DR_x_weighted_KS_pandas(temp_2, bad='bad', weight='weight', out_name='x_ks_df')
print(x_ks_df)


#autobinning  function B:
import duckdb
def DR_num_rbin_duckdb(
    df_input: pd.DataFrame,
    uniq_id: str,
    bad: str,
    weight: str,
    max_bin: int,
    min_adj_bin: int
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Equivalent of SAS macro %wh_num_rbin using duckdb in Python.
    Returns:
        out_data_rbin_sql: original dataset with 'rbin' column.
        out_num_rbins: bin statistics per variable.
    """
    # Register input table
    duckdb.register("df_input", df_input)

    # Step 1: flag missing
    wh_rbin_0 = duckdb.query("""
        SELECT *, 
               CASE WHEN x_value IS NULL THEN 1.0 ELSE 0.0 END AS miss_yn
        FROM df_input
    """).to_df()
    duckdb.register("wh_rbin_0", wh_rbin_0)

    # Step 2: summarize missing and determine bin count
    wh_rbin_1 = duckdb.query(f"""
        SELECT 
            x_nm,
            SUM({weight} * miss_yn) / SUM({weight}) AS miss_rate,
            SUM({weight}) AS tot_wgt,
            SUM((1.0 - miss_yn) * {weight}) AS tot_nomiss_wgt,
            CEIL({max_bin} * (1.000 - SUM({weight} * miss_yn) / SUM({weight}))) AS adj_nomiss_bin,
            CASE 
                WHEN {min_adj_bin} > CEIL({max_bin} * (1.000 - SUM({weight} * miss_yn) / SUM({weight}))) 
                THEN {min_adj_bin}
                ELSE CEIL({max_bin} * (1.000 - SUM({weight} * miss_yn) / SUM({weight})))
            END AS nomiss_bin
        FROM wh_rbin_0
        GROUP BY x_nm
    """).to_df()
    duckdb.register("wh_rbin_1", wh_rbin_1)

    # Step 3: group by x_value, join with summary
    wh_rbin_2 = duckdb.query(f"""
        SELECT a.*, b.tot_nomiss_wgt, b.nomiss_bin
        FROM (
            SELECT x_nm, miss_yn, x_value, SUM({weight}) AS x_wgt
            FROM wh_rbin_0
            GROUP BY x_nm, miss_yn, x_value
        ) a
        LEFT JOIN wh_rbin_1 b 
        ON a.x_nm = b.x_nm
        ORDER BY a.x_nm, miss_yn, x_value
    """).to_df()

    # Step 4: cumulative binning logic (in pandas)
    wh_rbin_3a = wh_rbin_2.copy()
    wh_rbin_3a['cum_nomiss_wgt'] = wh_rbin_3a.groupby(['x_nm', 'miss_yn'])['x_wgt'].cumsum()
    wh_rbin_3a['nomiss_bin_w'] = (wh_rbin_3a['tot_nomiss_wgt'] / wh_rbin_3a['nomiss_bin']).round().fillna(1.0)
    wh_rbin_3a['bin_id'] = wh_rbin_3a.apply(
        lambda row: row['x_value'] if row['miss_yn'] == 1 else 1 + int(row['cum_nomiss_wgt'] / (1.0 + row['nomiss_bin_w'])),
        axis=1
    )
    duckdb.register("wh_rbin_3a", wh_rbin_3a)

    # Step 5: count by bin_id
    wh_rbin_3b = duckdb.query("""
        SELECT x_nm, bin_id, COUNT(*) AS tempwh_cnt
        FROM wh_rbin_3a
        WHERE miss_yn = 0
        GROUP BY x_nm, bin_id
        ORDER BY x_nm, bin_id
    """).to_df()

    # Step 6: assign rbin number
    wh_rbin_3b['rbin'] = wh_rbin_3b.groupby('x_nm').cumcount() + 1
    duckdb.register("wh_rbin_3b", wh_rbin_3b)

    # Step 7: final bin assignment
    wh_rbin_3 = duckdb.query("""
        SELECT a.*, 
               CASE WHEN a.miss_yn = 1 THEN a.bin_id ELSE b.rbin END AS rbin
        FROM wh_rbin_3a a
        LEFT JOIN wh_rbin_3b b 
        ON a.x_nm = b.x_nm AND a.bin_id = b.bin_id
    """).to_df()
    duckdb.register("wh_rbin_3", wh_rbin_3)

    # Step 8: join back to original
    out_data_rbin_sql = duckdb.query("""
        SELECT a.*, b.rbin
        FROM df_input a
        LEFT JOIN wh_rbin_3 b 
        ON a.x_nm = b.x_nm AND a.x_value = b.x_value
        ORDER BY a.x_nm, a.x_value
    """).to_df()
    
    duckdb.register("out_data_rbin_sql", out_data_rbin_sql)
    # Step 9: compute final bin stats
    out_num_rbins = duckdb.query(f"""
        SELECT h.*, h.rbin_wgt / g.tot_wgt AS rbin_dist_pct
        FROM (
            SELECT x_nm, rbin, COUNT(*) AS obs_cnt,
                   MIN(x_value) AS min_x,
                   MAX(x_value) AS max_x,
                   SUM(x_value * {weight}) / SUM({weight}) AS mean_x,
                   SUM({weight}) AS rbin_wgt,
                   SUM({weight} * {bad}) AS rbin_bad_wgt,
                   SUM(1.0 * {weight} * {bad}) / SUM({weight}) AS rbin_bad_pct
            FROM out_data_rbin_sql
            GROUP BY x_nm, rbin
        ) h
        LEFT JOIN wh_rbin_1 g
        ON h.x_nm = g.x_nm
        ORDER BY h.x_nm, h.rbin
    """).to_df()

    return out_data_rbin_sql, out_num_rbins

# call the function
out_data_rbin_sql, out_num_rbins = DR_num_rbin_duckdb(
    df_input=temp_3,  # Including x_nm, x_value, weight, bad etc. columns
    uniq_id='loan_id',
    bad='bad',
    weight='weight',
    max_bin=5,
    min_adj_bin=5
)


#information value calculation** function C:
import statsmodels.api as sm
from scipy.stats import linregress

def DR_num_iv_woe_duckdb(df_rbin: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    import duckdb
    import numpy as np
    import pandas as pd
    import statsmodels.api as sm
    from scipy.stats import linregress

    duckdb.register("out_data_rbin_sql", df_rbin)

    # Step 1: WOE/IV calculation 
    tempwh_iv_1 = duckdb.query("""
        SELECT h.*, 
               g.tot_wgt,
               h.sum_wgt / g.tot_wgt AS dist_pct,
               g.tot_miss_wgt / g.tot_wgt AS tot_miss_rate,
               g.tot_bad_rate,
               CASE WHEN h.bad_rate = 0.0 THEN 0.5 / h.sum_wgt
                    WHEN h.bad_rate = 1.0 THEN ((h.sum_wgt - 1) - 0.5) / h.sum_wgt
                    ELSE h.bad_rate
               END AS cal_bad_rate,
               CASE WHEN h.bad_rate = 0.0 THEN 0.5 ELSE h.sum_bad_wgt END / g.tot_bad_wgt AS dist_bad_pct,
               CASE WHEN h.bad_rate = 1.0 THEN 0.5 ELSE h.sum_good_wgt END / g.tot_good_wgt AS dist_good_pct,
               LOG(cal_bad_rate / (1.0 - cal_bad_rate)) AS logit_bad_rate,
               LOG(dist_good_pct / dist_bad_pct) AS woe,
               (dist_good_pct - dist_bad_pct) * LOG(dist_good_pct / dist_bad_pct) AS bin_iv
        FROM (
            SELECT x_nm, rbin,
                   MIN(x_value) AS min_x_value,
                   MAX(x_value) AS max_x_value,
                   SUM(weight * x_value) / SUM(weight) AS avg_x_value,
                   SUM(weight) AS sum_wgt,
                   SUM(weight * bad) AS sum_bad_wgt,
                   SUM(weight * (1 - bad)) AS sum_good_wgt,
                   SUM(weight * bad) / SUM(weight) AS bad_rate
            FROM out_data_rbin_sql
            GROUP BY x_nm, rbin
        ) h
        LEFT JOIN (
            SELECT x_nm,
                   SUM(weight) AS tot_wgt,
                   SUM(weight * bad) AS tot_bad_wgt,
                   SUM(weight * (1.0 - bad)) AS tot_good_wgt,
                   SUM(CASE WHEN x_value IS NULL THEN weight ELSE 0.0 END) AS tot_miss_wgt,
                   SUM(weight * bad) / SUM(weight) AS tot_bad_rate
            FROM out_data_rbin_sql
            GROUP BY x_nm
        ) g
        ON h.x_nm = g.x_nm
        ORDER BY h.x_nm, h.rbin
    """).to_df()

    # Step 2: descriptive statistical summary
    desc_stats = df_rbin.groupby("x_nm")["x_value"].agg(
        mean='mean', min='min', max='max',
        p1=lambda x: np.percentile(x.dropna(), 1),
        p99=lambda x: np.percentile(x.dropna(), 99)
    ).reset_index()

    # Step 3: estimation of bivariate-logistic
    logit_results = []
    for x_nm, group in df_rbin.groupby("x_nm"):
        x_raw = group["x_value"]
        y = group["bad"]
        w = group["weight"]

        if x_raw.isnull().all() or x_raw.nunique() < 2 or y.nunique() < 2:
            est, pval = np.nan, np.nan
        else:
            try:
                #  x_value standardzation
                x_scaled = (x_raw - x_raw.mean()) / (x_raw.std(ddof=0) + 1e-6)
                X = sm.add_constant(x_scaled, has_constant='add')
                model = sm.GLM(y, X, family=sm.families.Binomial(), freq_weights=w)
                result = model.fit(disp=0)
                est = result.params.get("x_value", np.nan)
                pval = result.pvalues.get("x_value", np.nan)
            except Exception:
                est, pval = np.nan, np.nan

        logit_results.append({"x_nm": x_nm, "estimate": est, "probchisq": pval})
    df_logit = pd.DataFrame(logit_results)

    # Step 4: KS calculation
    df_woe_bin = tempwh_iv_1.copy()
    df_woe_bin["cum_dist_pct"] = df_woe_bin.groupby("x_nm")["dist_pct"].cumsum()
    df_woe_bin["cum_dist_bad_pct"] = df_woe_bin.groupby("x_nm")["dist_bad_pct"].cumsum()
    df_woe_bin["cum_dist_good_pct"] = df_woe_bin.groupby("x_nm")["dist_good_pct"].cumsum()
    df_woe_bin["diff_cum_bad_good_pct"] = df_woe_bin["cum_dist_bad_pct"] - df_woe_bin["cum_dist_good_pct"]
    duckdb.register("tempwh_iv_1", df_woe_bin)

    # Step 5: logit vs avg_x_value simulation（PROC REG ）
    reg_results = []
    for x_nm, group in df_woe_bin[df_woe_bin["rbin"].notnull()].groupby("x_nm"):
        x = group["avg_x_value"]
        y = group["logit_bad_rate"]

        if x.isnull().all() or x.nunique() <= 1 or y.nunique() <= 1:
            slope, r_square = np.nan, np.nan
        else:
            try:
                slope, _, r_value, _, _ = linregress(x, y)
                r_square = r_value**2
            except Exception:
                slope, r_square = np.nan, np.nan

        reg_results.append({"x_nm": x_nm, "Beta": slope, "R_Square": r_square})
    df_reg = pd.DataFrame(reg_results)

    # Step 6: binning
    df_nomiss_rbin = df_rbin[df_rbin["x_value"].notnull()].groupby("x_nm")["rbin"].nunique().reset_index()
    df_nomiss_rbin.columns = ["x_nm", "nomiss_rbin_cnt"]

    # Step 7: summary output
    df_iv_summary = duckdb.query("""
        SELECT 
            h.x_nm,
            h.tot_miss_rate AS miss_rate,
            h.rbin_cnt,
            p.nomiss_rbin_cnt,
            h.iv,
            h.ks,
            w.R_Square,
            f.estimate AS Beta,
            f.probchisq,
            g.mean, g.min, g.max, g.p1, g.p99,
            h.tot_wgt
        FROM (
            SELECT x_nm, tot_wgt, tot_miss_rate,
                   COUNT(*) AS rbin_cnt,
                   SUM(bin_iv) AS iv,
                   MAX(ABS(diff_cum_bad_good_pct)) AS ks
            FROM tempwh_iv_1
            GROUP BY x_nm, tot_wgt, tot_miss_rate
        ) h
        LEFT JOIN desc_stats g ON h.x_nm = g.x_nm
        LEFT JOIN df_logit f ON h.x_nm = f.x_nm
        LEFT JOIN df_reg w ON h.x_nm = w.x_nm
        LEFT JOIN df_nomiss_rbin p ON h.x_nm = p.x_nm
        ORDER BY h.x_nm
    """).to_df()

    return df_woe_bin, df_iv_summary
#call the function
df_woe, df_iv = DR_num_iv_woe_duckdb(out_data_rbin_sql)

#function D    automaticallt numeric attributes to MFC:
def DR_num_aMFC(in_data_rbin, in_num_iv, in_num_woe, uniq_id, bad, weight, 
                out_mfc_value_name='out_mfc_value', out_aMFC_SAS_name='out_aMFC_SAS'):
    
    # tempwh_whMFC_1
    tempwh_whMFC_1 = duckdb.query(f"""
        SELECT a.x_nm, a.rbin, a.avg_x_value, a.bad_rate
        FROM in_num_woe a, in_num_iv b
        WHERE a.x_nm = b.x_nm
        ORDER BY a.x_nm, a.rbin
    """).to_df()

    # tempwh_whMFC_2
    tempwh_whMFC_2 = duckdb.query(f"""
        SELECT 
            a.x_nm, a.rbin, a.bad_rate, 
            b.rbin AS b_rbin, 
            b.avg_x_value AS b_avg_x_value, 
            b.bad_rate AS b_bad_rate,
            ABS(a.bad_rate - b.bad_rate) AS abd_bad_diff
        FROM tempwh_whMFC_1 a
        CROSS JOIN tempwh_whMFC_1 b
        WHERE a.x_nm = b.x_nm 
          AND a.rbin IS  NULL 
          AND b.rbin IS NOT NULL
        ORDER BY a.x_nm, a.rbin, abd_bad_diff
    """).to_df()

    # tempwh_whMFC_3 (drop duplicates by x_nm and rbin)
    tempwh_whMFC_3 = tempwh_whMFC_2.drop_duplicates(subset=["x_nm", "rbin"])

   # Step 4: tempwh_whMFC_4
    merged = duckdb.query(f"""
        SELECT 
            a.*, 
            b.b_avg_x_value,
            CASE 
                WHEN a.rbin IS NOT NULL THEN a.x_value
                WHEN b.b_avg_x_value IS NOT NULL THEN b.b_avg_x_value
                ELSE a.mean
            END AS tempwh_x_value,
            CASE 
                WHEN (CASE 
                        WHEN a.rbin IS NOT NULL THEN a.x_value
                        WHEN b.b_avg_x_value IS NOT NULL THEN b.b_avg_x_value
                        ELSE a.mean
                      END) > a.p99 THEN a.p99
                WHEN (CASE 
                        WHEN a.rbin IS NOT NULL THEN a.x_value
                        WHEN b.b_avg_x_value IS NOT NULL THEN b.b_avg_x_value
                        ELSE a.mean
                      END) < a.p1 THEN a.p1
                ELSE (CASE 
                        WHEN a.rbin IS NOT NULL THEN a.x_value
                        WHEN b.b_avg_x_value IS NOT NULL THEN b.b_avg_x_value
                        ELSE a.mean
                      END)
            END AS mfc_x_value
        FROM (
            SELECT h.*, g.mean, g.p1, g.p99
            FROM in_data_rbin h
            JOIN in_num_iv g ON h.x_nm = g.x_nm
        ) a
        LEFT JOIN tempwh_whMFC_3 b
        ON a.x_nm = b.x_nm AND a.rbin = b.rbin
        ORDER BY {uniq_id}, a.x_nm, {bad}, {weight}
    """).to_df()

    # PROC TRANSPOSE: wide format
    temp_pivot = merged.pivot_table(index=[uniq_id, bad, weight],
                                    columns='x_nm',
                                    values='mfc_x_value',
                                    aggfunc='first').reset_index()

    # output
    out_mfc_value = tempwh_whMFC_3.copy()
    out_aMFC_SAS = temp_pivot.copy()

    # saving CSV 
    globals()[out_mfc_value_name] = out_mfc_value
    globals()[out_aMFC_SAS_name] = out_aMFC_SAS

    return out_mfc_value, out_aMFC_SAS

#call the function as above
duckdb.register("in_data_rbin", temp_5)
duckdb.register("in_num_iv", temp_6)
duckdb.register("in_num_woe", temp_7)

out_val, out_sas = DR_num_aMFC("in_data_rbin", "in_num_iv", "in_num_woe", 
                               uniq_id="uniq_id", bad="bad", weight="weight")




#num_woe function E:
def transpose_woe(temp_10: pd.DataFrame) -> pd.DataFrame:
    out_woe_data_sas = temp_10.pivot_table(
        index=['uniq_id', 'bad', 'weight'],  #  BY
        columns='x_nm',                      #  ID
        values='woe',                        #  VAR
        aggfunc='first'                      # dedup and only keep the first one by default
    ).reset_index()

    # optinoal: with list name( go back to MultiIndex）
    out_woe_data_sas.columns.name = None  # drop columns' name
    return out_woe_data_sas
#call the function
out_df = transpose_woe(temp_10)



#semi function F  mapping to special values for bureau-wised attributes:
# claim 
length_sv_map = {}

# Iterate over lengths from 2 to 9 (skipping 8)
for length in range(2, 10):
    if length == 8:
        continue  # skipping length=8
    subset = df_attr[df_attr['length'] == (length)]  
    quoted_names = ["'" + attr + "'" for attr in subset['attr_name']]
    sv_string = ",".join(quoted_names)
    length_sv_map[f"_{length}_sv"] = sv_string
    
def parse_sv_dict(sv_dict):
    parsed = {}
    for k, v in sv_dict.items():
        parsed[k] = [x.strip().strip("'").strip('"') for x in v.split(',')]
    return parsed

length_sv_map = parse_sv_dict(length_sv_map)
    
#print(length_sv_map["_2_sv"])
# output: 'age','sex','zip'（depending on data）

##special values bear in mind, no need to treated as in SAS program
def convert_special_missing(temp_2, sv_dict):
    def map_special(row):
        xnm = row['x_nm']
        val = row['x_value']
        
        for length in range(2, 10):
            if length == 8:
                continue  # skipping 8

            sv_list = sv_dict.get(f"_{length}_sv", [])
            base = int('9' * length)
            if xnm in sv_list:
                if val == base-9 + 1: return -10  # .a
                elif val == base-9 + 2: return -20  # .b
                elif val == base-9 + 3: return -30  # .c
                elif val == base-9 + 4: return  -40  # .d
                elif val == base-9 + 5: return  -50  # .e
                elif val == base-9 + 6: return  -60  # .f
                elif val == base-9 + 7: return  -70  # .g
                elif val == base-9 + 8: return   -80 # .h
                elif val == base-9 + 9: return   -90 # .i
        return val

    temp_2['x_value_'] = temp_2.apply(map_special, axis=1)
    return temp_2  
#call the function
sv_data = convert_special_missing(temp2_, length_sv_map)

#rmianing no rule of special value replacement:
# define the length of x_nm
group_dict = {
    2: {'ALL0317','ALL0448','ALL1380','ALL2002','ALL2005','ALL2427','ALL2428','ALL2907',
        'ALL2937','ALL2967','BCA0416','BCC0446','BCC3342','BCC3345','BCC3512','BCX3421',
        'BCX3422','BRC0416','BRC1300','FIP0300','FIP2358','MTF2358'},
    3: {'ALL4080', 'BCC7117', 'BRC7140'},
    4: {'ALL8325', 'MTF8166'},
    9: {'BCA5030','BCA5070','BCA5740','BRC5320','BRC5830','REV5620','ALL5743','ALL5070','ALL5047'}
}

# define a mapping function
def map_special_sas_like(temp_2, sv_dict):
    def map_special(row):
        xnm = row['x_nm']
        val = row['x_value']
        xnm = row['x_nm']
        val = row['x_value']
        for length in (2, 3,4,9):

            sv_list = sv_dict.get(length, [])
            base = int('9' * length)
            if xnm in sv_list:
                if val == base-9 + 1: return -10  # .a
                elif val == base-9 + 2: return -20  # .b
                elif val == base-9 + 3: return -30  # .c
                elif val == base-9 + 4: return  -40  # .d
                elif val == base-9 + 5: return  -50  # .e
                elif val == base-9 + 6: return  -60  # .f
                elif val == base-9 + 7: return  -70  # .g
                elif val == base-9 + 8: return   -80 # .h
                elif val == base-9 + 9: return   -90 # .i
        return val

    temp_2['x_value_'] = temp_2.apply(map_special, axis=1)
    return temp_2  

#call function
sv_data_2= map_special_sas_like(sv_data, group_dict)
#transpose to wide format
df_pivot = sv_data_2.pivot(index='uniq_id', columns='x_nm', values='x_value_').reset_index()

# Step 2: summary stats (such as proc means)
summary = df_pivot[pre_attr].describe(include='all')
print(summary)
# still some special values not replaced, due to they will not be in model candidate**

                            

#function G  of char to iv and woe **  char to woe ;
def DR_char_iv_woe_duckdb(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
   
    duckdb.register("char_data", df)

    # Step 1: calculate bin-level IV/WOE（including log protection）
    df_bin = duckdb.query("""
        SELECT 
            h.*, 
            g.tot_wgt,
            h.sum_wgt / g.tot_wgt AS dist_pct,
            h.sum_bad_wgt / g.tot_bad_wgt AS dist_bad_pct,
            h.sum_good_wgt / g.tot_good_wgt AS dist_good_pct,
            g.tot_miss_wgt / g.tot_wgt AS tot_miss_rate,
            g.tot_bad_rate,
            CASE 
                WHEN h.bad_rate = 0.0 THEN 0.5 / h.sum_wgt
                WHEN h.bad_rate = 1.0 THEN ((h.sum_wgt - 1) - 0.5) / h.sum_wgt
                ELSE h.bad_rate
            END AS cal_bad_rate,
            CASE 
                WHEN h.bad_rate = 0.0 THEN NULL 
                WHEN h.bad_rate = 1.0 THEN NULL 
                ELSE LOG(h.bad_rate / (1.0 - h.bad_rate))
            END AS logit_bad_rate,
            CASE 
                WHEN dist_good_pct > 0 AND dist_bad_pct > 0 THEN 
                    LOG(dist_good_pct / dist_bad_pct)
                ELSE NULL 
            END AS woe,
            CASE 
                WHEN dist_good_pct > 0 AND dist_bad_pct > 0 THEN 
                    (dist_good_pct - dist_bad_pct) * LOG(dist_good_pct / dist_bad_pct)
                ELSE 0.0 
            END AS bin_iv
        FROM (
            SELECT 
                x_nm, x_value,
                SUM(weight) AS sum_wgt,
                SUM(weight * bad) AS sum_bad_wgt,
                SUM(weight) - SUM(weight * bad) AS sum_good_wgt,
                SUM(weight * bad) / SUM(weight) AS bad_rate
            FROM char_data
            GROUP BY x_nm, x_value
        ) h
        LEFT JOIN (
            SELECT 
                x_nm,
                SUM(weight) AS tot_wgt,
                SUM(weight * bad) AS tot_bad_wgt,
                SUM(weight * (1.0 - bad)) AS tot_good_wgt,
                SUM(CASE WHEN x_value IS NULL THEN weight ELSE 0.0 END) AS tot_miss_wgt,
                SUM(weight * bad) / SUM(weight) AS tot_bad_rate
            FROM char_data
            GROUP BY x_nm
        ) g
        ON h.x_nm = g.x_nm
        ORDER BY h.x_nm, cal_bad_rate
    """).to_df()

    # Step 2: replace missing WOE with the maximum WOE of each variable
    non_missing = df_bin[df_bin["woe"].notnull()].copy()
    non_missing["woe_"] = non_missing["woe"]

    max_woe = non_missing.sort_values(["x_nm", "bad_rate"]).drop_duplicates("x_nm", keep="last")[["x_nm", "woe"]]

    missing = df_bin[df_bin["woe"].isnull()].copy()
    missing = missing.merge(max_woe, on="x_nm", how="left", suffixes=("", "_filled"))
    missing["woe_"] = missing["woe_filled"]

    # merge data
    df_woe = pd.concat([non_missing, missing], axis=0, ignore_index=True)
    df_woe = df_woe.sort_values(["x_nm", "cal_bad_rate"])

    # Step 3: cumulative distribution
    df_woe["cum_dist_pct"] = df_woe.groupby("x_nm")["dist_pct"].cumsum()
    df_woe["cum_dist_bad_pct"] = df_woe.groupby("x_nm")["dist_bad_pct"].cumsum()
    df_woe["cum_dist_good_pct"] = df_woe.groupby("x_nm")["dist_good_pct"].cumsum()
    df_woe["diff_cum_bad_good_pct"] = df_woe["cum_dist_bad_pct"] - df_woe["cum_dist_good_pct"]

    # Step 4: summarized output
    df_iv = df_woe.groupby(["x_nm", "tot_wgt", "tot_miss_rate"]).agg(
        x_grp_cnt=("x_value", "count"),
        IV=("bin_iv", "sum"),
        KS=("diff_cum_bad_good_pct", lambda x: np.max(np.abs(x)))
    ).reset_index().sort_values(by="x_nm")

    return df_woe, df_iv

#function H:
def DR_char_grp_to_woe_duckdb(df_raw: pd.DataFrame, df_woe_map: pd.DataFrame) -> pd.DataFrame:
    import duckdb
    import pandas as pd

    duckdb.register("raw_data", df_raw)
    duckdb.register("woe_map", df_woe_map)

    df_merged = duckdb.query("""
        SELECT a.*, b.woe_
        FROM raw_data a
        LEFT JOIN woe_map b
        ON a.x_nm = b.x_nm AND a.x_value = b.x_value
        ORDER BY a.obs_id, a.x_nm
    """).to_df()

    # pivot the WOE of each variable into a single row
    df_transformed = df_merged.pivot_table(
        index=["obs_id", "bad", "weight"], 
        columns="x_nm", 
        values="woe_"
    ).reset_index()

    return df_transformed

#call the two function
df_char_woe, df_char_iv = DR_char_iv_woe_duckdb(charwoe_dev_attr)
df_char_woe_ready = DR_char_grp_to_woe_duckdb(charwoe_dev_attr, df_char_woe)


#function  I: woe dimentional reduction 
#dr_woe_dim_reduction_more
from sklearn.decomposition import PCA
from sklearn.linear_model import LogisticRegression
from sklearn.feature_selection import SequentialFeatureSelector
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from sklearn import set_config
set_config(enable_metadata_routing=True)


def dr_woe_dim_reduction_more(
    df_woe: pd.DataFrame,
    uniq_id: str,
    bad: int,
    weight: int,
    data_subgrp: int,
    clus_cnt: int = 10,
    top_n: int = 1,
    sle: float = 0.05,
    forward_stop_n: int = 10
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Python version of %dr_woe_dim_reduction_more.
    Returns:
        - df_dimR: final reduced feature dataset
        - x_list_df: summary of selected variables and sources
    """
    exclude_cols = [uniq_id, bad, weight, data_subgrp]
    all_x_list = [col for col in df_woe.columns if col not in exclude_cols]

    X = df_woe[all_x_list].fillna(0)
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    # Step 1: PCA for clustering proxy
    pca = PCA(n_components=clus_cnt)
    pca.fit(X_scaled)
    components = pca.components_.T
    cluster_assignments = np.argmax(np.abs(components), axis=1)

    clustering_df = pd.DataFrame({
        "Variable": all_x_list,
        "Cluster": cluster_assignments,
        "RSquareRatio": np.var(X_scaled, axis=0)  # proxy
    })

    # Step 2: pick top N variables from each cluster
    clustering_df = clustering_df.sort_values(["Cluster", "RSquareRatio"], ascending=[True, False])
    clustering_df["cnt"] = clustering_df.groupby("Cluster").cumcount() + 1
    df_cluster_topn = clustering_df[clustering_df["cnt"] <= top_n].copy()
    df_cluster_topn["Source"] = "Cluster"

    # Step 3: candidates not in cluster representatives
    remaining_candidates = list(set(all_x_list) - set(df_cluster_topn["Variable"]))

    def forward_selection(X_df, y, weights, label):
        model = LogisticRegression(max_iter=1000, solver="liblinear")
        model.set_fit_request(sample_weight=True)
        sfs = SequentialFeatureSelector(model,
                                        n_features_to_select=min(forward_stop_n, X_df.shape[1]),
                                        direction="forward",
                                        scoring="neg_log_loss",
                                        cv=3,
                                        n_jobs=-1)
        sfs.fit(X_df, y, sample_weight=weights)
        selected_vars = list(X_df.columns[sfs.get_support()])
        return pd.DataFrame({"Variable": selected_vars, "Source": label})

    # Prepare X/y/weights
    y = df_woe[bad].values
    sample_weight = df_woe[weight].values

    selected_after_clu = forward_selection(df_woe[remaining_candidates].fillna(0), y, sample_weight, "After Clu")
    selected_all = forward_selection(df_woe[all_x_list].fillna(0), y, sample_weight, "All Forward")
    selected_all_by_grp = []
    for grp, df_grp in df_woe.groupby(data_subgrp):
        try:
            sel = forward_selection(df_grp[all_x_list].fillna(0),
                                     df_grp[bad].values,
                                     df_grp[weight].values,
                                     "All by Forward")
            selected_all_by_grp.append(sel)
        except Exception:
            continue
    selected_all_by_grp = pd.concat(selected_all_by_grp, ignore_index=True) if selected_all_by_grp else pd.DataFrame(columns=["Variable", "Source"])

    # Combine
    candidate_all = pd.concat([
        df_cluster_topn[["Variable", "Source"]],
        selected_after_clu,
        selected_all,
        selected_all_by_grp
    ]).drop_duplicates()

    selected_vars = candidate_all["Variable"].unique().tolist()
    df_dimR = df_woe[[uniq_id, bad, weight] + selected_vars].copy()

    # Summary
    summary = pd.DataFrame({"Variable": selected_vars})
    for src in ["Cluster", "After Clu", "All Forward", "All by Forward"]:
        summary[src.replace(" ", "_")] = summary["Variable"].isin(
            candidate_all[candidate_all["Source"] == src]["Variable"]
        ).astype(int)

    return df_dimR, summary

#call the function;
auto_df_dimR, autosel_summary=dr_woe_dim_reduction_more(
    df_woe=auto_model_data2,
    uniq_id= 'uniq_id',
    bad='bad',
    weight='weight',
    data_subgrp='tag',
    clus_cnt = 20,
    top_n= 2,
    sle  = 0.2,
    forward_stop_n  = 18
)

#function J:logit_full_report_to_excel generating logic plots for each potential model attribute candidates
import os
import pandas as pd
import statsmodels.api as sm
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows

def logit_full_report_to_excel(df_woe, output_excel='logit_full_report.xlsx'):
    os.makedirs("logit_full_report_imgs", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Logit_Report"

    current_row = 1

    for var in df_woe['x_nm'].dropna().unique():
        sub_df = df_woe[df_woe['x_nm'] == var].copy()
        sub_df_clean = sub_df.dropna(subset=['avg_x_value', 'logit_bad_rate', 'sum_wgt'])

        if sub_df_clean.shape[0] < 2:
            continue

        # regression modeling
        X = sm.add_constant(sub_df_clean['avg_x_value'])
        y = sub_df_clean['logit_bad_rate']
        weights = sub_df_clean['sum_wgt']
        model = sm.WLS(y, X, weights=weights).fit()

        # insert variable headers
        ws.cell(row=current_row, column=1, value=f"Variable: {var}")
        current_row += 1

        # Plot Figure
        plt.figure(figsize=(6, 4))
        sns.scatterplot(data=sub_df, x='avg_x_value', y='logit_bad_rate', size='sum_wgt', legend=False)
        plt.plot(sub_df_clean['avg_x_value'], model.predict(X), color='blue', linestyle='--')
        plt.title(f"{var} logit_bad_rate vs avg_x_value")
        plt.xlabel('avg_x_value')
        plt.ylabel('logit_bad_rate')
        plot_path = f"logit_full_report_imgs/{var}.png"
        plt.tight_layout()
        plt.savefig(plot_path)
        plt.close()

        # Insert Image
        img = ExcelImage(plot_path)
        img.anchor = f"A{current_row}"
        ws.add_image(img)
        current_row += 20

        # Insert regression summary information
        ws.cell(row=current_row, column=1, value=f"R-squared: {model.rsquared:.4f}")
        ws.cell(row=current_row + 1, column=1, value=f"Adj R-squared: {model.rsquared_adj:.4f}")
        ws.cell(row=current_row + 2, column=1, value=f"RMSE: {model.mse_resid**0.5:.4f}")
        current_row += 4

        # Insert coefficient table
        ws.cell(row=current_row, column=1, value="Coefficients:")
        coef_df = model.summary2().tables[1].reset_index()
        for row in dataframe_to_rows(coef_df, index=False, header=True):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=current_row, column=col_idx, value=val)
            current_row += 1
        current_row += 1

        # Insert raw data
        ws.cell(row=current_row, column=1, value="Raw Table:")
        current_row += 1
        for row in dataframe_to_rows(sub_df, index=False, header=True):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=current_row, column=col_idx, value=val)
            current_row += 1

        current_row += 3

    wb.save(output_excel)
    print(f"All logit outputs saved to one sheet: {output_excel}")
#call the plots function
logit_full_report_to_excel(stemp_mfc_1b, output_excel=r'D:\HA_study_plan_tasks\banking_financial_scorecard_dev\model_1\mfc\full_logit_report2.xlsx')


#function K:
#missing floor capping function;
def mfc(in_x, mrv, floor, cap):
    if pd.isna(in_x):
        return mrv
    elif in_x < floor:
        return floor
    elif in_x > cap:
        return cap
    else:
        return in_x

def mfc2(in_x, mv1, mrv1, mv2, mrv2, floor, cap):
    if in_x == mv1:
        return mrv1
    elif in_x == mv2:
        return mrv2
    else:
        return mfc(in_x, None, floor, cap)

def mfc3(in_x, mv1, mrv1, mv2, mrv2, mv3, mrv3, floor, cap):
    if in_x == mv1:
        return mrv1
    elif in_x == mv2:
        return mrv2
    elif in_x == mv3:
        return mrv3
    else:
        return mfc(in_x, None, floor, cap)

def mfc4(in_x, mv1, mrv1, mv2, mrv2, mv3, mrv3, mv4, mrv4, floor, cap):
    if in_x == mv1:
        return mrv1
    elif in_x == mv2:
        return mrv2
    elif in_x == mv3:
        return mrv3
    elif in_x == mv4:
        return mrv4
    else:
        return mfc(in_x, None, floor, cap)



#***************function L:************manully*selection criteria  **********************
# [1] Statistically significant
# [2] Marg_sign=model_sign
# [3] vif < 2.0
# [4] Business reasonable

#seed variable set
from statsmodels.stats.outliers_influence import variance_inflation_factor
def forward_logistic_selection(data, target, candidate_vars, weight_col=None, max_vars=10):
    selected = []
    remaining = candidate_vars.copy()
    y = data[target]
    weights = data[weight_col] if weight_col else None

    while len(selected) < max_vars and remaining:
        best_score = -float('inf')
        best_var = None
        for var in remaining:
            try:
                X = sm.add_constant(data[selected + [var]])
                model = sm.Logit(y, X, weights=weights).fit(disp=False)
                score = model.llf  # log-likelihood
                if score > best_score:
                    best_score = score
                    best_var = var
            except:
                continue
        if best_var is not None:
            selected.append(best_var)
            remaining.remove(best_var)
        else:
            break
    return selected

#call the function;
selected_vars = forward_logistic_selection(sel_model_dev, 'bad', m_list, weight_col='weight', max_vars=10)




#function M: weighted KS for checking lift KS of final score or key var **
def weighted_ks(in_data_pred: pd.DataFrame, bad: str, weight: str, pred: str) -> pd.DataFrame:
    # Step 1: add good_wgt / bad_wgt and sort by pred (DuckDB)
    duckdb.register("in_data_pred", in_data_pred)

    tempwh_ks_0 = duckdb.query(f"""
        SELECT *, 
               CASE WHEN {bad} = 1.0 THEN {weight} ELSE 0.0 END AS bad_wgt,
               CASE WHEN {bad} = 0.0 THEN {weight} ELSE 0.0 END AS good_wgt
        FROM in_data_pred
        ORDER BY {pred}
    """).to_df()

    # Step 2: cumulative weighted sum
    tempwh_ks_0["cum_bad_wgt"] = tempwh_ks_0["bad_wgt"].cumsum()
    tempwh_ks_0["cum_good_wgt"] = tempwh_ks_0["good_wgt"].cumsum()

    # Step 3: total weight
    total_bad_wgt = tempwh_ks_0["bad_wgt"].sum()
    total_good_wgt = tempwh_ks_0["good_wgt"].sum()

    tempwh_ks_0["bad_cdf"] = tempwh_ks_0["cum_bad_wgt"] / total_bad_wgt if total_bad_wgt > 0 else 0
    tempwh_ks_0["good_cdf"] = tempwh_ks_0["cum_good_wgt"] / total_good_wgt if total_good_wgt > 0 else 0

    # Step 4: calculate KS
    tempwh_ks_0["ks_gap"] = abs(tempwh_ks_0["bad_cdf"] - tempwh_ks_0["good_cdf"])
    ks_value = tempwh_ks_0["ks_gap"].max() * 100

    out_KS = pd.DataFrame({
        "variable": [f"KS of {pred}"],
        "KS": [ks_value]
    })

    return out_KS
#call the function;
sel_ks = weighted_ks(in_data_pred=sel_pred, bad="bad", weight="weight", pred="pred")


#function N:model lift table generation**
def weighted_lift_ks_roc(model_desc, in_data, indepvar, bad, weight, bin_size, descending_yn=1):
    # register data to DuckDB
    duckdb.register("in_data", in_data)

    sort_order = "DESC" if descending_yn == 1 else "ASC"

    # Step 1: Group by score → bad_wgt / good_wgt
    tempwh_xks_0s = duckdb.query(f"""
        SELECT {indepvar},
               SUM(CASE WHEN {bad}=1.0 THEN {weight} ELSE 0.0 END) AS bad_wgt,
               SUM(CASE WHEN {bad}=0.0 THEN {weight} ELSE 0.0 END) AS good_wgt
        FROM in_data
        WHERE {weight} IS NOT NULL AND {bad} IS NOT NULL AND {indepvar} IS NOT NULL
        GROUP BY {indepvar}
        ORDER BY {indepvar} {sort_order}
    """).to_df()

    # Step 2: Add total stats
    tempwh_ks_roc_0 = tempwh_xks_0s.copy()
    tempwh_ks_roc_0["tot_wgt"] = tempwh_ks_roc_0["bad_wgt"] + tempwh_ks_roc_0["good_wgt"]
    tot_bad_wgt = tempwh_ks_roc_0["bad_wgt"].sum()
    tot_good_wgt = tempwh_ks_roc_0["good_wgt"].sum()
    tot_wgt = tot_bad_wgt + tot_good_wgt

    # Step 3: Cumulative & binning
    df = tempwh_ks_roc_0.copy()
    df["cum_wgt"] = (df["bad_wgt"] + df["good_wgt"]).cumsum()
    df["cum_bad_wgt"] = df["bad_wgt"].cumsum()
    df["cum_good_wgt"] = df["good_wgt"].cumsum()
    bin_w = tot_wgt / bin_size
    df["bin"] = 1 + (df["cum_wgt"] / (1.0 + bin_w)).apply(np.floor).astype(int)

    # Step 4: KS
    df["bad_cdf"] = df["cum_bad_wgt"] / tot_bad_wgt
    df["good_cdf"] = df["cum_good_wgt"] / tot_good_wgt
    df["ks_gap"] = abs(df["bad_cdf"] - df["good_cdf"])
    ks_value = df["ks_gap"].max() * 100

    # Step 5: ROC - sensitivity vs 1-specificity
    df["id"] = range(1, len(df)+1)
    df["sensitivity"] = (tot_bad_wgt - df["cum_bad_wgt"]) / tot_bad_wgt
    df["specifity"] = df["cum_good_wgt"] / tot_good_wgt
    df["x"] = 1.0 - df["specifity"]
    df["y"] = df["sensitivity"]

    df["area"] = (df["x"].diff().fillna(0)) * (df["y"] + df["y"].shift(1).fillna(0)) / 2
    roc_auc = df["area"].sum()

    # Step 6: binning metrics output
    grouped = df.groupby("bin").agg(
        min_indepvar=(indepvar, "min"),
        max_indepvar=(indepvar, "max"),
        avg_indepvar=(indepvar, lambda x: np.average(x, weights=(df.loc[x.index, "bad_wgt"] + df.loc[x.index, "good_wgt"]))),
        marg_all=("cum_wgt", "max"),
        marg_bad=("cum_bad_wgt", "max"),
        marg_good=("cum_good_wgt", "max"),
        cum_wgt=("cum_wgt", "max"),
        cum_bad_wgt=("cum_bad_wgt", "max"),
        cum_good_wgt=("cum_good_wgt", "max"),
        one_minus_specifity=("x", "max"),
        sensitivity=("y", "max")
    ).reset_index()

    grouped["marg_dist_all_rate"] = grouped["marg_all"] / tot_wgt
    grouped["marg_dist_bad_rate"] = grouped["marg_bad"] / tot_bad_wgt
    grouped["marg_dist_good_rate"] = grouped["marg_good"] / tot_good_wgt
    grouped["marg_bad_rate"] = grouped["marg_bad"] / grouped["marg_all"]
    grouped["woe"] = np.log(grouped["marg_dist_bad_rate"] / grouped["marg_dist_good_rate"])
    grouped["cum_dist_all_rate"] = grouped["cum_wgt"] / tot_wgt
    grouped["cum_dist_bad_rate"] = grouped["cum_bad_wgt"] / tot_bad_wgt
    grouped["cum_dist_good_rate"] = grouped["cum_good_wgt"] / tot_good_wgt
    grouped["cum_bad_rate"] = grouped["cum_bad_wgt"] / grouped["cum_wgt"]

    out_ks_roc = pd.DataFrame({
        "model": [model_desc],
        "indepvar_order": [sort_order.lower()],
        "KS": [ks_value],
        "ROC": [roc_auc],
        "tot_wgt": [tot_wgt],
        "tot_bad_rate": [tot_bad_wgt / tot_wgt]
    })

    return out_ks_roc, grouped


#function O: ROC function;
def plot_roc(df_bins):
    df_plot = pd.concat([
        pd.DataFrame({"sensitivity": [0], "one_minus_specifity": [0]}),
        df_bins[["sensitivity", "one_minus_specifity"]]
    ])
    plt.plot(df_plot["one_minus_specifity"], df_plot["sensitivity"], marker="o")
    plt.title("ROC Curve")
    plt.xlabel("1 - Specificity")
    plt.ylabel("Sensitivity")
    plt.grid(True)
    plt.show()

#function P: plot_lorenz function;
def plot_lorenz(df_bins):
    df_plot = pd.concat([
        pd.DataFrame({"cum_dist_bad_rate": [0], "cum_dist_all_rate": [0]}),
        df_bins[["cum_dist_bad_rate", "cum_dist_all_rate"]]
    ])
    plt.plot(df_plot["cum_dist_all_rate"], df_plot["cum_dist_bad_rate"], marker="o", color="orange")
    plt.title("Lorenz Curve")
    plt.xlabel("Cumulative Distribution of Population")
    plt.ylabel("Cumulative Distribution of Bad")
    plt.grid(True)
    plt.show()

#call the function;    
sel_out_ks_roc, sel_out_bins = weighted_lift_ks_roc(
    model_desc="Model A",
    in_data=sel_pred,
    indepvar="pred",
    bad="bad",
    weight="weight",
    bin_size=10,
    descending_yn=1
)

print(sel_out_ks_roc)
print(sel_out_bins.head())

#call the function;
plot_roc(sel_out_bins)
plot_lorenz(sel_out_bins)

#function Q:
#sel_addone_bin  fine selection function;
import statsmodels.api as sm
import duckdb
from statsmodels.stats.outliers_influence import variance_inflation_factor

def sel_addone_bin(
    model_data: pd.DataFrame,
    uniq_id: str,
    bad: int,
    weight: int,
    bin_cnt: int,
    m_x_list: list,
    sel_x_list: list,
):
    # 1. temp_addOne_1: keep uniq_id, weight, bad, m_x_list; drop sel_x_list
    cols_to_keep = [uniq_id, weight, bad] + m_x_list
    temp_addOne_1 = model_data[cols_to_keep].copy()
    temp_addOne_1.drop(columns=sel_x_list, errors='ignore', inplace=True)

    # 2. sort by uniq_id, weight, bad (mimic proc sort)
    temp_addOne_1 = temp_addOne_1.sort_values(by=[uniq_id, weight, bad])

    # 3. transpose: convert from wide to long
    # SAS proc transpose by uniq_id, weight, bad, _name_ to x_nm, values to x_value
    temp_addOne_2 = (
        temp_addOne_1
        .set_index([uniq_id, weight, bad])
        .stack()
        .reset_index()
        .rename(columns={'level_3': 'x_nm', 0: 'x_value'})
    )

    # 4. temp_addOne_3: keep uniq_id, weight, bad, sel_x_list from original model_data
    temp_addOne_3 = model_data[[uniq_id, weight, bad] + sel_x_list].copy()

    # 5. SQL join temp_addOne_3 with temp_addOne_2 on uniq_id and bad (left join)
    con = duckdb.connect(database=':memory:')
    con.register('temp_addOne_3', temp_addOne_3)
    con.register('temp_addOne_2', temp_addOne_2)

    temp_addOne_4 = con.query(f"""
        SELECT b.x_nm, b.x_value, a.*
        FROM temp_addOne_3 a
        LEFT JOIN temp_addOne_2 b
          ON a.{uniq_id} = b.{uniq_id} AND a.{bad} = b.{bad}
        ORDER BY x_nm, a.{uniq_id}, a.{weight}, a.{bad}
    """).to_df()

    # 6. logistic regression for each x_nm group: bad ~ sel_x_list + x_value, weighted by weight
    # statistical paremeters，vif，predicted probabilities
    all_results = []
    all_vif = []
    all_preds = []

    for x_nm, group in temp_addOne_4.groupby('x_nm'):
        # dropna for needed columns
        cols_for_model = sel_x_list + ['x_value']
        temp_df = group[[bad, weight] + cols_for_model].dropna()

        # skip groups with too few samples
        if temp_df.shape[0] < 10:
            continue

        # independent variable X, add intercept term
        X = temp_df[cols_for_model]
        X = sm.add_constant(X, has_constant='add')

        y = temp_df[bad]
        w = temp_df[weight]

        try:
            # weighted logitistic model
            logit_model = sm.Logit(y, X, freq_weights=w).fit(disp=0)

            # parameters estimation
            params = logit_model.params.to_frame('Estimate').reset_index().rename(columns={'index': 'Variable'})

            # Wald Chi-square
            wald_chi = (logit_model.params / logit_model.bse)**2
            from scipy.stats import chi2
            prob_chi = chi2.sf(wald_chi, df=1)

            params['WaldChiSq'] = wald_chi.values
            params['ProbChiSq'] = prob_chi
            params['x_nm'] = x_nm

            # VIF calculation
            vif_data = pd.DataFrame()
            if len(cols_for_model) > 1:
                X_vif = temp_df[cols_for_model].copy()
                # calculate VIF for each variable
                vif_data = pd.DataFrame()
                vif_data['Variable'] = X_vif.columns
                vif_data['VarianceInflation'] = [variance_inflation_factor(X_vif.values, i) for i in range(X_vif.shape[1])]
            else:
                # When only one variable x_value，set VIF to nan
                vif_data = pd.DataFrame({'Variable': cols_for_model, 'VarianceInflation': [np.nan]*len(cols_for_model)})

            vif_data['x_nm'] = x_nm

            all_results.append(params)
            all_vif.append(vif_data)

            # predict probabilities
            pred = logit_model.predict(X)
            pred_df = temp_df[[bad, weight]].copy()
            pred_df['pred'] = pred
            pred_df['x_nm'] = x_nm

            all_preds.append(pred_df)

        except Exception as e:
            # skip for error
            print(f"Warning: logit failed for variable {x_nm}: {e}")
            continue

    # merge results
    res_params = pd.concat(all_results, ignore_index=True)
    res_vif = pd.concat(all_vif, ignore_index=True)
    pred_all = pd.concat(all_preds, ignore_index=True)

    # 7. Calculate KS (cumulative weighted difference between bad and good) by sorting cumulatively according to x_nm and pred
    con.register('temp_pred', pred_all)
    ks_df = con.query(f"""
        WITH ordered AS (
            SELECT
                x_nm, pred,
                SUM(CASE WHEN {bad} = 1 THEN {weight} ELSE 0 END) AS bad_wgt,
                SUM(CASE WHEN {bad} = 0 THEN {weight} ELSE 0 END) AS good_wgt
            FROM temp_pred
            GROUP BY x_nm, pred
        ),
        cum AS (
            SELECT
                x_nm, pred,
                SUM(bad_wgt) OVER (PARTITION BY x_nm ORDER BY pred) AS cum_bad_wgt,
                SUM(good_wgt) OVER (PARTITION BY x_nm ORDER BY pred) AS cum_good_wgt,
                SUM(bad_wgt) OVER (PARTITION BY x_nm) AS tot_bad_wgt,
                SUM(good_wgt) OVER (PARTITION BY x_nm) AS tot_good_wgt
            FROM ordered
        )
        SELECT
            x_nm,
            100 * MAX(ABS(cum_bad_wgt / tot_bad_wgt - cum_good_wgt / tot_good_wgt)) AS KS
        FROM cum
        GROUP BY x_nm
        ORDER BY KS DESC
    """).to_df()

    # 8. calculate bin_bad_pct for each bin
    # calculate each record’s weight and cumulative sum after sorting
    con.register('pred_bin', pred_all)
    bin_df = con.query(f"""
        SELECT
            x_nm, pred, {bad}, SUM({weight}) AS wgt
        FROM pred_bin
        GROUP BY x_nm, pred, {bad}
    """).to_df()

    # calcualte total weight by x_nm
    total_wgt = bin_df.groupby('x_nm')['wgt'].sum().rename('tot_wgt').reset_index()
    bin_df = bin_df.merge(total_wgt, on='x_nm', how='left')

    # calculate cumulative weights by x_nm, pred in descending order
    bin_df = bin_df.sort_values(['x_nm', 'pred'], ascending=[True, False])
    bin_df['cum_wgt'] = bin_df.groupby('x_nm')['wgt'].cumsum()

    # caculate width for each bin
    bin_df['bin_w'] = np.floor(0.5 + bin_df['tot_wgt'] / bin_cnt)
    bin_df['bin'] = 1 + np.floor(bin_df['cum_wgt'] / (1.0 + bin_df['bin_w']))

    # calculate bin_bad_pct
    bin_bad_pct_df = (
        bin_df.groupby(['x_nm', 'bin'])
        .apply(lambda x: np.sum(x['wgt'] * x[bad]) / np.sum(x['wgt']))
        .reset_index(name='bin_bad_pct')
    )

    # 9. transport bin_bad_pct，id=bin, row is x_nm，columns are bin_1, bin_2 ...
    bin_bad_pct_wide = bin_bad_pct_df.pivot(index='x_nm', columns='bin', values='bin_bad_pct')
    bin_bad_pct_wide.columns = [f'bin_{int(c)}' for c in bin_bad_pct_wide.columns]
    bin_bad_pct_wide = bin_bad_pct_wide.reset_index()

    # 10. summarize all results：merge KS，parameters，VIF，bin_bad_pct
    # add a flag int_yn in the coefficient table to mark the Intercept
    res_params['int_yn'] = res_params['Variable'].apply(lambda x: 1 if x.lower() == 'const' or x.lower() == 'intercept' else 0)

    # merge all tables
    final_df = (
        res_params
        .merge(res_vif, on=['x_nm', 'Variable'], how='left')
        .merge(ks_df, on='x_nm', how='left')
        .merge(bin_bad_pct_wide, on='x_nm', how='left')
    )

    # sorting by ks，x_nm，int_yn，WaldChiSq in descending order
    final_df = final_df.sort_values(by=['KS', 'x_nm', 'int_yn', 'WaldChiSq'], ascending=[False, True, False, False])

    # drop int_yn
    final_df = final_df.drop(columns=['int_yn'])

    # close connection
    con.close()

    return final_df.reset_index(drop=True)

#call the function;
result_df = sel_addone_bin(
    model_data=sel_model_dev,  # DataFrame
    uniq_id="uniq_id",         # 
    bad="bad",                 # 
    weight="weight",           # 
    bin_cnt=10,                # 
    m_x_list=m_list,           # 
    sel_x_list=sel_x_list      # 
)

